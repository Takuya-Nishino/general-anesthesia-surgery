# ============================================================
# Standalone offline IECV SHAP interaction analysis
#
# Project directory:
#   C:\Users\tears\Desktop\Reanalysis_20260714
#
# Required local inputs:
#   1) Revice_data_20260714.xlsx
#   2) Existing final IECV model bundles under a 03_models folder
#
# No internet connection is used.
# Model fitting, imputation, hyperparameter tuning, and calibration
# are NOT rerun. Existing fitted IECV model bundles are loaded.
#
# Analyses:
#   Preoperative / Perioperative × XGBoost / LightGBM
#   Pooled out-of-sample IECV observations
#   Mean absolute SHAP interaction values
#   Diagonal elements set to zero
#   Values below the within-model 80th percentile masked
# ============================================================

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Mapping, Sequence, Tuple
import hashlib
import importlib.metadata
import json
import platform
import re
import sys
import warnings

import cloudpickle
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import shap
import xgboost
import lightgbm

# Imports required when unpickling the saved preprocessing objects.
from lightgbm import LGBMRegressor
from sklearn.experimental import enable_iterative_imputer  # noqa: F401
from sklearn.impute import IterativeImputer
from sklearn.isotonic import IsotonicRegression
from sklearn.preprocessing import RobustScaler

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from PIL import Image
except ImportError:
    Image = None


# ============================================================
# 1. USER SETTINGS
# ============================================================

PROJECT_DIR = Path(
    r"C:\Users\tears\Desktop\Reanalysis_20260714"
)

RAW_DATA_PATH = (
    PROJECT_DIR
    / "Revice_data_20260714.xlsx"
)

RAW_DATA_SHEET: int | str = 0

# Usually leave this as None.
# The code automatically locates a 03_models folder under PROJECT_DIR.
# Set a specific path only when the model directory is elsewhere.
MODEL_DIR_OVERRIDE: Path | None = None

# New outputs are kept separate from the main analysis outputs.
OUTPUT_DIR = (
    PROJECT_DIR
    / "SHAP_interaction_outputs"
)

ID_COL = "INDEX"
FACILITY_COL = "付属_1"
OUTCOME_COL = "Event"
OUTCOME_LABEL = "Primary_composite"

RANDOM_STATE = 42

# Maximum number of held-out patients sampled per facility.
# This is identical to the final global SHAP setting.
MAX_PATIENTS_PER_FACILITY = 2000

# Reduce this value if memory is limited.
INTERACTION_BATCH_SIZE = 100

# Manuscript-specified masking threshold.
INTERACTION_PERCENTILE = 80.0

# Number of features displayed per heatmap panel.
# Use None to display every feature.
PLOT_TOP_N: int | None = 15

FIGURE_DPI = 180
PNG_MAX_WIDTH = 1200
PNG_MAX_HEIGHT = 1200
PNG_MAX_FILE_SIZE_MB = 5.0

# Reuse facility-level interaction caches after interruption.
REUSE_EXISTING_CACHE = True

COMBINATIONS: Tuple[Tuple[str, str], ...] = (
    ("Preoperative", "XGBoost"),
    ("Preoperative", "LightGBM"),
    ("Perioperative", "XGBoost"),
    ("Perioperative", "LightGBM"),
)

PANEL_LETTERS = ("A", "B", "C", "D")

MODEL_ABBREVIATION = {
    "XGBoost": "XGB",
    "LightGBM": "LGBM",
}

VARIANT_ABBREVIATION = {
    "Preoperative": "Preop",
    "Perioperative": "Periop",
}

SHAP_FEATURE_LABELS = {
    "ASA": "ASA-PS",
    "DeliMed": "Any diabetes medication",
    "AntiCa": "Calcium-channel blocker",
    "ResectNum": "Number of resections",
    "HighRiskProc": "High-risk procedure",
    "RBC Tx": "RBC transfusion",
    "FFP Tx": "FFP transfusion",
    "PLT Tx": "Platelet transfusion",
    "FluidBal": "Intraoperative fluid balance",
    "OpTime": "Operative duration",
    "HR at 6h": "Heart rate at 6 h",
    "MAP at 6h": "Mean arterial pressure at 6 h",
}


# ============================================================
# 2. PATHS
# ============================================================

CACHE_DIR = (
    OUTPUT_DIR
    / "01_facility_cache"
)
TABLE_DIR = (
    OUTPUT_DIR
    / "02_tables"
)
FIGURE_DIR = (
    OUTPUT_DIR
    / "03_figures"
)
LOG_DIR = (
    OUTPUT_DIR
    / "04_logs"
)

for directory in (
    CACHE_DIR,
    TABLE_DIR,
    FIGURE_DIR,
    LOG_DIR,
):
    directory.mkdir(
        parents=True,
        exist_ok=True,
    )

TABLE_PATH = (
    TABLE_DIR
    / "TableSx_SHAP_Interaction_Analysis.xlsx"
)

FIGURE_PATH = (
    FIGURE_DIR
    / "FigureS5_SHAP_Interaction_Heatmaps.png"
)

MANIFEST_PATH = (
    LOG_DIR
    / "SHAP_Interaction_Run_Manifest.json"
)


# ============================================================
# 3. GENERAL HELPERS
# ============================================================

def sanitize_filename(value: Any) -> str:
    text = str(value)
    text = re.sub(
        r"[^A-Za-z0-9_.-]+",
        "_",
        text,
    )
    return text.strip("_") or "value"


def feature_label(feature: str) -> str:
    return SHAP_FEATURE_LABELS.get(
        feature,
        feature,
    )


def installed_version(
    distribution_name: str,
) -> str:
    try:
        return importlib.metadata.version(
            distribution_name
        )
    except importlib.metadata.PackageNotFoundError:
        return "not installed"


def sha256_file(
    path: Path,
    chunk_size: int = 1024 * 1024,
) -> str:
    digest = hashlib.sha256()

    with path.open("rb") as file:
        while True:
            chunk = file.read(
                chunk_size
            )
            if not chunk:
                break
            digest.update(
                chunk
            )

    return digest.hexdigest()


def file_stat_signature(
    path: Path,
) -> str:
    stat = path.stat()
    return (
        f"{stat.st_size}:"
        f"{stat.st_mtime_ns}"
    )


def json_serializable(
    value: Any,
) -> Any:
    if isinstance(
        value,
        np.integer,
    ):
        return int(value)

    if isinstance(
        value,
        np.floating,
    ):
        return float(value)

    if isinstance(
        value,
        np.ndarray,
    ):
        return value.tolist()

    if isinstance(
        value,
        Path,
    ):
        return str(value)

    raise TypeError(
        f"Unsupported JSON type: {type(value)}"
    )


def save_json(
    payload: Mapping[str, Any],
    path: Path,
) -> None:
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with path.open(
        "w",
        encoding="utf-8",
    ) as file:
        json.dump(
            dict(payload),
            file,
            ensure_ascii=False,
            indent=2,
            default=json_serializable,
        )


def locate_model_dir(
    project_dir: Path,
) -> Path:
    if MODEL_DIR_OVERRIDE is not None:
        candidate = Path(
            MODEL_DIR_OVERRIDE
        )
        expected = (
            candidate
            / "primary"
            / "iecv"
            / OUTCOME_LABEL
        )

        if not expected.exists():
            raise FileNotFoundError(
                "MODEL_DIR_OVERRIDE does not contain "
                "the expected IECV model structure.\n"
                f"Expected:\n{expected}"
            )

        return candidate

    direct = (
        project_dir
        / "03_models"
    )

    if (
        direct
        / "primary"
        / "iecv"
        / OUTCOME_LABEL
    ).exists():
        return direct

    candidates = []

    for path in project_dir.rglob(
        "03_models"
    ):
        expected = (
            path
            / "primary"
            / "iecv"
            / OUTCOME_LABEL
        )
        if expected.exists():
            candidates.append(
                path
            )

    unique_candidates = sorted(
        {
            path.resolve()
            for path in candidates
        }
    )

    if len(unique_candidates) == 1:
        return unique_candidates[0]

    if len(unique_candidates) == 0:
        raise FileNotFoundError(
            "No final 03_models directory was found under PROJECT_DIR.\n\n"
            "The SHAP interaction analysis requires the fitted IECV model "
            "bundles created by the main analysis.\n\n"
            "Expected structure:\n"
            f"{project_dir}\\03_models\\primary\\iecv\\"
            f"{OUTCOME_LABEL}\\IECV_heldout_*\\"
            "Preoperative__XGBoost.pkl, etc.\n\n"
            "Copy the existing 03_models folder into PROJECT_DIR, "
            "or set MODEL_DIR_OVERRIDE to its current local path."
        )

    candidate_text = "\n".join(
        str(path)
        for path in unique_candidates
    )

    raise RuntimeError(
        "Multiple 03_models directories were found. "
        "Set MODEL_DIR_OVERRIDE explicitly.\n\n"
        f"Candidates:\n{candidate_text}"
    )


def load_bundle(
    path: Path,
) -> Mapping[str, Any]:
    with path.open(
        "rb"
    ) as file:
        bundle = cloudpickle.load(
            file
        )

    required_keys = {
        "feature_names",
        "preprocessor",
        "estimator",
    }
    missing_keys = (
        required_keys
        - set(bundle.keys())
    )

    if missing_keys:
        raise KeyError(
            "The saved model bundle is missing required keys: "
            f"{sorted(missing_keys)}\n"
            f"Bundle:\n{path}"
        )

    return bundle


def model_bundle_path(
    model_dir: Path,
    *,
    held_out_facility: str,
    variant: str,
    model_name: str,
) -> Path:
    outer_key = (
        f"IECV_heldout_{held_out_facility}"
    )

    return (
        model_dir
        / "primary"
        / "iecv"
        / OUTCOME_LABEL
        / sanitize_filename(
            outer_key
        )
        / (
            f"{sanitize_filename(variant)}__"
            f"{sanitize_filename(model_name)}.pkl"
        )
    )


def find_bundle_or_raise(
    model_dir: Path,
    *,
    held_out_facility: str,
    variant: str,
    model_name: str,
) -> Path:
    expected = model_bundle_path(
        model_dir,
        held_out_facility=held_out_facility,
        variant=variant,
        model_name=model_name,
    )

    if expected.exists():
        return expected

    filename = (
        f"{sanitize_filename(variant)}__"
        f"{sanitize_filename(model_name)}.pkl"
    )
    outcome_root = (
        model_dir
        / "primary"
        / "iecv"
        / OUTCOME_LABEL
    )

    candidates = sorted(
        outcome_root.rglob(
            filename
        )
    )

    outer_key_sanitized = sanitize_filename(
        f"IECV_heldout_{held_out_facility}"
    )

    facility_matches = [
        path
        for path in candidates
        if outer_key_sanitized
        in path.parent.name
    ]

    if len(facility_matches) == 1:
        return facility_matches[0]

    candidate_text = (
        "\n".join(
            str(path)
            for path in candidates[:30]
        )
        if candidates
        else "No candidates found."
    )

    raise FileNotFoundError(
        "The expected IECV model bundle was not found.\n"
        f"Held-out facility: {held_out_facility}\n"
        f"Feature set: {variant}\n"
        f"Model: {model_name}\n\n"
        f"Expected:\n{expected}\n\n"
        f"Bundles with the same filename:\n{candidate_text}"
    )


def transform_tree_model_input(
    bundle: Mapping[str, Any],
    X_raw: pd.DataFrame,
) -> np.ndarray:
    transformed = (
        bundle["preprocessor"]
        .transform(
            X_raw
        )
    )

    transformed = np.asarray(
        transformed,
        dtype=np.float32,
    )

    if transformed.ndim != 2:
        raise ValueError(
            "The preprocessor did not return a 2-dimensional matrix. "
            f"Observed shape: {transformed.shape}"
        )

    if not np.isfinite(
        transformed
    ).all():
        raise ValueError(
            "Nonfinite values remain after preprocessing."
        )

    return transformed


def stratified_sample_positions(
    y: np.ndarray,
    max_n: int,
    seed: int,
) -> np.ndarray:
    y_array = np.asarray(
        y,
        dtype=int,
    )
    n = len(
        y_array
    )

    if n <= max_n:
        return np.arange(
            n,
            dtype=int,
        )

    rng = np.random.default_rng(
        seed
    )

    positive = np.flatnonzero(
        y_array == 1
    )
    negative = np.flatnonzero(
        y_array == 0
    )

    if (
        len(positive) == 0
        or len(negative) == 0
    ):
        return np.sort(
            rng.choice(
                n,
                size=max_n,
                replace=False,
            )
        ).astype(int)

    target_positive = int(
        round(
            max_n
            * len(positive)
            / n
        )
    )

    target_positive = max(
        1,
        min(
            target_positive,
            len(positive),
        ),
    )

    target_negative = (
        max_n
        - target_positive
    )

    if target_negative > len(
        negative
    ):
        target_negative = len(
            negative
        )
        target_positive = min(
            max_n
            - target_negative,
            len(positive),
        )

    selected_positive = rng.choice(
        positive,
        size=target_positive,
        replace=False,
    )
    selected_negative = rng.choice(
        negative,
        size=target_negative,
        replace=False,
    )

    return np.sort(
        np.concatenate(
            [
                selected_positive,
                selected_negative,
            ]
        )
    ).astype(int)


def normalize_interaction_values(
    raw_values: Any,
    *,
    n_samples: int,
    n_features: int,
) -> np.ndarray:
    if hasattr(
        raw_values,
        "values",
    ):
        raw_values = raw_values.values

    if isinstance(
        raw_values,
        list,
    ):
        if len(raw_values) == 0:
            raise ValueError(
                "SHAP interaction output is empty."
            )

        raw_values = (
            raw_values[1]
            if len(raw_values) > 1
            else raw_values[0]
        )

    array = np.asarray(
        raw_values,
        dtype=float,
    )

    expected = (
        n_samples,
        n_features,
        n_features,
    )

    if array.shape == expected:
        return array

    if array.ndim == 4:
        # (n, p, p, classes)
        if array.shape[:3] == expected:
            class_index = (
                1
                if array.shape[3] > 1
                else 0
            )
            return array[
                :,
                :,
                :,
                class_index,
            ]

        # (classes, n, p, p)
        if array.shape[1:] == expected:
            class_index = (
                1
                if array.shape[0] > 1
                else 0
            )
            return array[
                class_index,
                :,
                :,
                :,
            ]

        # (n, classes, p, p)
        if (
            array.shape[0] == n_samples
            and array.shape[2] == n_features
            and array.shape[3] == n_features
        ):
            class_index = (
                1
                if array.shape[1] > 1
                else 0
            )
            return array[
                :,
                class_index,
                :,
                :,
            ]

    raise ValueError(
        "Unexpected SHAP interaction shape. "
        f"Observed: {array.shape}; expected: {expected}"
    )


def calculate_batch_interactions(
    estimator: Any,
    X_model: np.ndarray,
) -> np.ndarray:
    n_samples, n_features = (
        X_model.shape
    )

    explainer = shap.TreeExplainer(
        estimator,
        feature_perturbation="tree_path_dependent",
        model_output="raw",
    )

    absolute_sum = np.zeros(
        (
            n_features,
            n_features,
        ),
        dtype=np.float64,
    )

    for start in range(
        0,
        n_samples,
        INTERACTION_BATCH_SIZE,
    ):
        stop = min(
            start
            + INTERACTION_BATCH_SIZE,
            n_samples,
        )

        X_batch = X_model[
            start:stop
        ]

        raw_values = (
            explainer
            .shap_interaction_values(
                X_batch
            )
        )

        values = (
            normalize_interaction_values(
                raw_values,
                n_samples=len(
                    X_batch
                ),
                n_features=n_features,
            )
        )

        if not np.isfinite(
            values
        ).all():
            raise ValueError(
                "Nonfinite SHAP interaction values were detected."
            )

        absolute_sum += (
            np.abs(
                values
            )
            .sum(
                axis=0
            )
        )

        print(
            f"      batch "
            f"{start + 1:,}-"
            f"{stop:,} / "
            f"{n_samples:,}"
        )

        del raw_values
        del values

    return absolute_sum


def make_pair_table(
    matrix: np.ndarray,
    feature_names: Sequence[str],
    threshold: float,
) -> pd.DataFrame:
    rows: List[
        Dict[str, Any]
    ] = []

    for i in range(
        len(feature_names)
    ):
        for j in range(
            i + 1,
            len(feature_names),
        ):
            value = float(
                matrix[i, j]
            )

            rows.append(
                {
                    "Feature_1": (
                        feature_names[i]
                    ),
                    "Feature_1_label": (
                        feature_label(
                            feature_names[i]
                        )
                    ),
                    "Feature_2": (
                        feature_names[j]
                    ),
                    "Feature_2_label": (
                        feature_label(
                            feature_names[j]
                        )
                    ),
                    "Mean_absolute_SHAP_interaction": value,
                    "At_or_above_80th_percentile": bool(
                        value
                        >= threshold
                    ),
                }
            )

    table = (
        pd.DataFrame(
            rows
        )
        .sort_values(
            "Mean_absolute_SHAP_interaction",
            ascending=False,
            kind="stable",
        )
        .reset_index(
            drop=True
        )
    )

    table.insert(
        0,
        "Rank",
        np.arange(
            1,
            len(table) + 1,
        ),
    )

    return table


def style_workbook(
    path: Path,
) -> None:
    workbook = load_workbook(
        path
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="2F75B5",
    )
    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False

        if worksheet.max_row >= 1:
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        for row in worksheet.iter_rows(
            min_row=2
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            letter = get_column_letter(
                column_cells[0].column
            )
            max_length = max(
                len(
                    str(
                        cell.value
                    )
                )
                if cell.value is not None
                else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[
                letter
            ].width = min(
                max(
                    max_length + 2,
                    10,
                ),
                45,
            )

    workbook.save(
        path
    )


def save_png_compliant(
    fig: plt.Figure,
    output_path: Path,
) -> None:
    temporary_path = output_path.with_name(
        output_path.stem
        + "__temporary.png"
    )

    fig.savefig(
        temporary_path,
        dpi=FIGURE_DPI,
        bbox_inches="tight",
        pad_inches=0.04,
        facecolor="white",
        transparent=False,
    )

    if Image is None:
        temporary_path.replace(
            output_path
        )
        warnings.warn(
            "Pillow is not installed; PNG dimensions and file size "
            "were not automatically adjusted."
        )
        return

    max_bytes = int(
        PNG_MAX_FILE_SIZE_MB
        * 1024
        * 1024
    )

    with Image.open(
        temporary_path
    ) as image:
        image.load()
        final_image = image.convert(
            "RGB"
        ).copy()

    temporary_path.unlink(
        missing_ok=True
    )

    final_image.thumbnail(
        (
            PNG_MAX_WIDTH,
            PNG_MAX_HEIGHT,
        ),
        Image.Resampling.LANCZOS,
    )

    while True:
        final_image.save(
            output_path,
            format="PNG",
            optimize=True,
            compress_level=9,
            dpi=(
                FIGURE_DPI,
                FIGURE_DPI,
            ),
        )

        file_size = (
            output_path
            .stat()
            .st_size
        )

        if (
            final_image.width
            <= PNG_MAX_WIDTH
            and final_image.height
            <= PNG_MAX_HEIGHT
            and file_size
            <= max_bytes
        ):
            break

        new_width = max(
            300,
            int(
                round(
                    final_image.width
                    * 0.90
                )
            ),
        )
        new_height = max(
            300,
            int(
                round(
                    final_image.height
                    * 0.90
                )
            ),
        )

        if (
            new_width
            == final_image.width
            and new_height
            == final_image.height
        ):
            raise RuntimeError(
                "The PNG could not be reduced to the specified limits."
            )

        final_image = final_image.resize(
            (
                new_width,
                new_height,
            ),
            Image.Resampling.LANCZOS,
        )


# ============================================================
# 4. INPUT VALIDATION
# ============================================================

if not PROJECT_DIR.exists():
    raise FileNotFoundError(
        f"PROJECT_DIR was not found:\n{PROJECT_DIR}"
    )

if not RAW_DATA_PATH.exists():
    raise FileNotFoundError(
        "The raw data workbook was not found.\n"
        f"Expected:\n{RAW_DATA_PATH}"
    )

MODEL_DIR = locate_model_dir(
    PROJECT_DIR
)

df_raw = pd.read_excel(
    RAW_DATA_PATH,
    sheet_name=RAW_DATA_SHEET,
).copy()

required_columns = {
    ID_COL,
    FACILITY_COL,
    OUTCOME_COL,
}
missing_columns = (
    required_columns
    - set(
        df_raw.columns
    )
)

if missing_columns:
    raise KeyError(
        "Required columns are missing: "
        f"{sorted(missing_columns)}"
    )

if df_raw[ID_COL].isna().any():
    raise ValueError(
        f"{ID_COL} contains missing values."
    )

if df_raw[ID_COL].duplicated().any():
    duplicated = (
        df_raw.loc[
            df_raw[ID_COL].duplicated(),
            ID_COL,
        ]
        .head(
            10
        )
        .tolist()
    )
    raise ValueError(
        f"{ID_COL} contains duplicate values. "
        f"Examples: {duplicated}"
    )

if df_raw[FACILITY_COL].isna().any():
    raise ValueError(
        f"{FACILITY_COL} contains missing values."
    )

df_raw[FACILITY_COL] = (
    df_raw[FACILITY_COL]
    .astype(
        str
    )
)

df_raw[OUTCOME_COL] = (
    pd.to_numeric(
        df_raw[OUTCOME_COL],
        errors="raise",
    )
    .astype(
        int
    )
)

outcome_values = set(
    df_raw[OUTCOME_COL]
    .unique()
)

if not outcome_values.issubset(
    {
        0,
        1,
    }
):
    raise ValueError(
        f"{OUTCOME_COL} must contain only 0 and 1."
    )

facilities = sorted(
    df_raw[FACILITY_COL]
    .unique()
    .tolist()
)

if len(
    facilities
) != 3:
    raise ValueError(
        "The final analysis expects exactly 3 facilities. "
        f"Observed: {facilities}"
    )

facility_label_map = {
    facility: chr(
        ord(
            "A"
        )
        + index
    )
    for index, facility in enumerate(
        facilities
    )
}

iecv_splits: List[
    Dict[str, Any]
] = []

for facility in facilities:
    validation_mask = (
        df_raw[FACILITY_COL]
        .eq(
            facility
        )
        .to_numpy()
    )

    iecv_splits.append(
        {
            "held_out_facility": facility,
            "held_out_label": (
                facility_label_map[
                    facility
                ]
            ),
            "train_indices": (
                np.flatnonzero(
                    ~validation_mask
                )
            ),
            "validation_indices": (
                np.flatnonzero(
                    validation_mask
                )
            ),
        }
    )

RAW_DATA_SHA256 = sha256_file(
    RAW_DATA_PATH
)

print(
    "=" * 72
)
print(
    "Offline IECV SHAP interaction analysis"
)
print(
    "=" * 72
)
print(
    "Project directory:",
    PROJECT_DIR,
)
print(
    "Raw data:",
    RAW_DATA_PATH,
)
print(
    "Model directory:",
    MODEL_DIR,
)
print(
    "Output directory:",
    OUTPUT_DIR,
)
print(
    "Cohort N:",
    f"{len(df_raw):,}",
)
print(
    "Events:",
    f"{int(df_raw[OUTCOME_COL].sum()):,}",
    f"({df_raw[OUTCOME_COL].mean():.1%})",
)
print(
    "Facility mapping:",
    facility_label_map,
)


# ============================================================
# 5. MODEL INVENTORY
# ============================================================

model_inventory_rows: List[
    Dict[str, Any]
] = []

for split in iecv_splits:
    facility = str(
        split[
            "held_out_facility"
        ]
    )

    for (
        variant,
        model_name,
    ) in COMBINATIONS:
        path = find_bundle_or_raise(
            MODEL_DIR,
            held_out_facility=facility,
            variant=variant,
            model_name=model_name,
        )

        model_inventory_rows.append(
            {
                "Held_out_facility": facility,
                "Held_out_label": (
                    split[
                        "held_out_label"
                    ]
                ),
                "Feature_set": variant,
                "Model": model_name,
                "Bundle_path": str(
                    path
                ),
                "Bundle_signature": (
                    file_stat_signature(
                        path
                    )
                ),
            }
        )

model_inventory = pd.DataFrame(
    model_inventory_rows
)

print(
    "\nAll required IECV model bundles were found:"
)
print(
    model_inventory[
        [
            "Held_out_label",
            "Feature_set",
            "Model",
        ]
    ].to_string(
        index=False
    )
)


# ============================================================
# 6. POOLED OUT-OF-SAMPLE SHAP INTERACTIONS
# ============================================================

interaction_results: Dict[
    Tuple[str, str],
    Dict[str, Any],
] = {}

sample_summary_rows: List[
    Dict[str, Any]
] = []

for (
    variant,
    model_name,
) in COMBINATIONS:
    print(
        "\n"
        + "=" * 72
    )
    print(
        "SHAP interaction | "
        f"{variant} | "
        f"{model_name}"
    )

    reference_path = (
        find_bundle_or_raise(
            MODEL_DIR,
            held_out_facility=(
                facilities[0]
            ),
            variant=variant,
            model_name=model_name,
        )
    )

    reference_bundle = load_bundle(
        reference_path
    )

    feature_names = list(
        reference_bundle[
            "feature_names"
        ]
    )

    n_features = len(
        feature_names
    )

    missing_features = [
        feature
        for feature in feature_names
        if feature not in df_raw.columns
    ]

    if missing_features:
        raise KeyError(
            "Features required by the saved model are missing "
            f"from the raw workbook: {missing_features}"
        )

    pooled_absolute_sum = np.zeros(
        (
            n_features,
            n_features,
        ),
        dtype=np.float64,
    )

    pooled_sample_n = 0

    for facility_index, split in enumerate(
        iecv_splits,
        start=1,
    ):
        facility = str(
            split[
                "held_out_facility"
            ]
        )
        facility_label = str(
            split[
                "held_out_label"
            ]
        )

        path = find_bundle_or_raise(
            MODEL_DIR,
            held_out_facility=facility,
            variant=variant,
            model_name=model_name,
        )

        bundle = load_bundle(
            path
        )

        bundle_feature_names = list(
            bundle[
                "feature_names"
            ]
        )

        if (
            bundle_feature_names
            != feature_names
        ):
            raise ValueError(
                "Feature names or order differ between IECV folds.\n"
                f"Feature set: {variant}\n"
                f"Model: {model_name}\n"
                f"Held-out facility: {facility_label}"
            )

        validation_indices = np.asarray(
            split[
                "validation_indices"
            ],
            dtype=int,
        )

        y_validation = (
            df_raw.iloc[
                validation_indices
            ][OUTCOME_COL]
            .to_numpy(
                dtype=int
            )
        )

        sampled_positions = (
            stratified_sample_positions(
                y=y_validation,
                max_n=MAX_PATIENTS_PER_FACILITY,
                seed=(
                    RANDOM_STATE
                    + facility_index
                ),
            )
        )

        sampled_indices = (
            validation_indices[
                sampled_positions
            ]
        )

        cache_path = (
            CACHE_DIR
            / (
                f"{sanitize_filename(variant)}__"
                f"{sanitize_filename(model_name)}__"
                f"heldout_{sanitize_filename(facility_label)}.npz"
            )
        )

        bundle_signature = file_stat_signature(
            path
        )

        cache_is_valid = False
        facility_absolute_sum = None

        if (
            REUSE_EXISTING_CACHE
            and cache_path.exists()
        ):
            try:
                cached = np.load(
                    cache_path,
                    allow_pickle=False,
                )

                cached_features = (
                    cached[
                        "feature_names"
                    ]
                    .astype(
                        str
                    )
                    .tolist()
                )

                cache_is_valid = bool(
                    cached[
                        "raw_data_sha256"
                    ][0]
                    == RAW_DATA_SHA256
                    and cached[
                        "bundle_signature"
                    ][0]
                    == bundle_signature
                    and cached_features
                    == feature_names
                    and np.array_equal(
                        cached[
                            "sampled_indices"
                        ].astype(
                            int
                        ),
                        sampled_indices,
                    )
                )

                if cache_is_valid:
                    facility_absolute_sum = (
                        cached[
                            "absolute_sum"
                        ]
                        .astype(
                            np.float64
                        )
                    )
                    print(
                        "  Held-out institution "
                        f"{facility_label}: "
                        "reused cached interactions"
                    )

            except Exception as exc:
                warnings.warn(
                    "Existing interaction cache could not be used and "
                    f"will be recalculated: {cache_path.name}: {exc}"
                )

        if not cache_is_valid:
            X_raw = (
                df_raw.iloc[
                    sampled_indices
                ][feature_names]
            )

            X_model = (
                transform_tree_model_input(
                    bundle,
                    X_raw,
                )
            )

            print(
                "  Held-out institution "
                f"{facility_label}: "
                f"N={len(sampled_indices):,}; "
                f"events="
                f"{int(df_raw.iloc[sampled_indices][OUTCOME_COL].sum()):,}"
            )

            facility_absolute_sum = (
                calculate_batch_interactions(
                    estimator=(
                        bundle[
                            "estimator"
                        ]
                    ),
                    X_model=X_model,
                )
            )

            np.savez_compressed(
                cache_path,
                absolute_sum=(
                    facility_absolute_sum
                ),
                feature_names=np.asarray(
                    feature_names,
                    dtype=str,
                ),
                sampled_indices=np.asarray(
                    sampled_indices,
                    dtype=np.int64,
                ),
                raw_data_sha256=np.asarray(
                    [
                        RAW_DATA_SHA256
                    ],
                    dtype=str,
                ),
                bundle_signature=np.asarray(
                    [
                        bundle_signature
                    ],
                    dtype=str,
                ),
            )

            del X_model

        if facility_absolute_sum is None:
            raise RuntimeError(
                "Facility-level interaction calculation did not return a matrix."
            )

        pooled_absolute_sum += (
            facility_absolute_sum
        )
        pooled_sample_n += len(
            sampled_indices
        )

        sample_summary_rows.append(
            {
                "Feature_set": variant,
                "Model": model_name,
                "Held_out_facility": facility,
                "Held_out_label": facility_label,
                "Validation_N": len(
                    validation_indices
                ),
                "Validation_events": int(
                    y_validation.sum()
                ),
                "Validation_prevalence": float(
                    y_validation.mean()
                ),
                "Interaction_sample_N": len(
                    sampled_indices
                ),
                "Interaction_sample_events": int(
                    df_raw.iloc[
                        sampled_indices
                    ][OUTCOME_COL]
                    .sum()
                ),
                "Cache_file": str(
                    cache_path
                ),
            }
        )

        del facility_absolute_sum

    if pooled_sample_n <= 0:
        raise RuntimeError(
            "The pooled SHAP interaction sample is empty."
        )

    mean_absolute_matrix = (
        pooled_absolute_sum
        / pooled_sample_n
    )

    # Correct negligible asymmetry caused by floating-point operations.
    mean_absolute_matrix = 0.5 * (
        mean_absolute_matrix
        + mean_absolute_matrix.T
    )

    # Manuscript-specified procedure.
    np.fill_diagonal(
        mean_absolute_matrix,
        0.0,
    )

    threshold = float(
        np.percentile(
            mean_absolute_matrix,
            INTERACTION_PERCENTILE,
        )
    )

    thresholded_matrix = np.where(
        mean_absolute_matrix
        >= threshold,
        mean_absolute_matrix,
        0.0,
    )

    np.fill_diagonal(
        thresholded_matrix,
        0.0,
    )

    pair_table = make_pair_table(
        matrix=mean_absolute_matrix,
        feature_names=feature_names,
        threshold=threshold,
    )

    interaction_results[
        (
            variant,
            model_name,
        )
    ] = {
        "feature_names": feature_names,
        "mean_absolute_matrix": mean_absolute_matrix,
        "thresholded_matrix": thresholded_matrix,
        "threshold": threshold,
        "pair_table": pair_table,
        "sample_n": pooled_sample_n,
    }

    np.savez_compressed(
        CACHE_DIR
        / (
            f"POOLED__{sanitize_filename(variant)}__"
            f"{sanitize_filename(model_name)}.npz"
        ),
        mean_absolute_matrix=(
            mean_absolute_matrix
        ),
        thresholded_matrix=(
            thresholded_matrix
        ),
        threshold=np.asarray(
            [
                threshold
            ],
            dtype=float,
        ),
        feature_names=np.asarray(
            feature_names,
            dtype=str,
        ),
        pooled_sample_n=np.asarray(
            [
                pooled_sample_n
            ],
            dtype=np.int64,
        ),
    )

    print(
        "  Completed | "
        f"pooled N={pooled_sample_n:,} | "
        f"features={n_features} | "
        f"80th-percentile threshold="
        f"{threshold:.6g}"
    )


# ============================================================
# 7. EXCEL OUTPUT
# ============================================================

summary_rows: List[
    Dict[str, Any]
] = []

with pd.ExcelWriter(
    TABLE_PATH,
    engine="openpyxl",
) as writer:
    for (
        variant,
        model_name,
    ) in COMBINATIONS:
        result = interaction_results[
            (
                variant,
                model_name,
            )
        ]

        model_abbr = MODEL_ABBREVIATION[
            model_name
        ]
        variant_abbr = VARIANT_ABBREVIATION[
            variant
        ]
        sheet_prefix = (
            f"{model_abbr}_"
            f"{variant_abbr}"
        )

        feature_names = result[
            "feature_names"
        ]
        feature_labels = [
            feature_label(
                feature
            )
            for feature in feature_names
        ]

        full_matrix_df = pd.DataFrame(
            result[
                "mean_absolute_matrix"
            ],
            index=feature_labels,
            columns=feature_labels,
        )

        thresholded_matrix_df = pd.DataFrame(
            result[
                "thresholded_matrix"
            ],
            index=feature_labels,
            columns=feature_labels,
        )

        result[
            "pair_table"
        ].to_excel(
            writer,
            sheet_name=(
                f"{sheet_prefix}_Pairs"
            ),
            index=False,
        )

        full_matrix_df.to_excel(
            writer,
            sheet_name=(
                f"{sheet_prefix}_Matrix"
            ),
        )

        thresholded_matrix_df.to_excel(
            writer,
            sheet_name=(
                f"{sheet_prefix}_Masked"
            ),
        )

        top_pair = (
            result[
                "pair_table"
            ]
            .iloc[
                0
            ]
        )

        summary_rows.append(
            {
                "Feature_set": variant,
                "Model": model_name,
                "Pooled_interaction_sample_N": (
                    result[
                        "sample_n"
                    ]
                ),
                "Number_of_features": len(
                    feature_names
                ),
                "Percentile_threshold": (
                    INTERACTION_PERCENTILE
                ),
                "Interaction_threshold": (
                    result[
                        "threshold"
                    ]
                ),
                "Top_feature_1": (
                    top_pair[
                        "Feature_1_label"
                    ]
                ),
                "Top_feature_2": (
                    top_pair[
                        "Feature_2_label"
                    ]
                ),
                "Top_mean_absolute_SHAP_interaction": (
                    top_pair[
                        "Mean_absolute_SHAP_interaction"
                    ]
                ),
            }
        )

    pd.DataFrame(
        summary_rows
    ).to_excel(
        writer,
        sheet_name="Summary",
        index=False,
    )

    pd.DataFrame(
        sample_summary_rows
    ).to_excel(
        writer,
        sheet_name="Sample_summary",
        index=False,
    )

    model_inventory.to_excel(
        writer,
        sheet_name="Model_inventory",
        index=False,
    )

    pd.DataFrame(
        {
            "Item": [
                "Outcome",
                "Validation",
                "Models",
                "Feature sets",
                "Explained output",
                "Interaction calculation",
                "Thresholding",
                "Interpretation",
            ],
            "Description": [
                OUTCOME_LABEL,
                (
                    "Pooled out-of-sample leave-one-institution-out "
                    "internal-external cross-validation"
                ),
                "XGBoost and LightGBM",
                "Preoperative and perioperative",
                (
                    "Base-model raw margin before isotonic "
                    "probability calibration"
                ),
                (
                    "Absolute SHAP interaction values averaged across "
                    "sampled held-out patients; diagonal elements set to zero"
                ),
                (
                    "Values below the within-model 80th percentile were masked"
                ),
                (
                    "Predictive interaction, not a causal effect; "
                    "magnitudes should not be compared directly across algorithms"
                ),
            ],
        }
    ).to_excel(
        writer,
        sheet_name="Notes",
        index=False,
    )

style_workbook(
    TABLE_PATH
)

print(
    "\nExcel output:",
    TABLE_PATH,
)


# ============================================================
# 8. COMBINED 2 × 2 FIGURE
# ============================================================

fig, axes = plt.subplots(
    2,
    2,
    figsize=(
        6.4,
        6.4,
    ),
    dpi=FIGURE_DPI,
)

axes = axes.ravel()

for (
    ax,
    panel_letter,
    combination,
) in zip(
    axes,
    PANEL_LETTERS,
    COMBINATIONS,
):
    variant, model_name = (
        combination
    )

    result = interaction_results[
        combination
    ]

    full_matrix = result[
        "mean_absolute_matrix"
    ]
    thresholded_matrix = result[
        "thresholded_matrix"
    ]
    feature_names = result[
        "feature_names"
    ]

    if PLOT_TOP_N is None:
        selected_indices = np.arange(
            len(
                feature_names
            )
        )
    else:
        total_strength = full_matrix.sum(
            axis=1
        )
        selected_indices = (
            np.argsort(
                total_strength
            )[::-1][
                :min(
                    int(
                        PLOT_TOP_N
                    ),
                    len(
                        feature_names
                    ),
                )
            ]
        )

    plot_matrix = thresholded_matrix[
        np.ix_(
            selected_indices,
            selected_indices,
        )
    ]

    labels = [
        feature_label(
            feature_names[
                index
            ]
        )
        for index in selected_indices
    ]

    masked_matrix = np.ma.masked_where(
        plot_matrix <= 0,
        plot_matrix,
    )

    color_map = plt.get_cmap(
        "viridis"
    ).copy()
    color_map.set_bad(
        "white"
    )

    image = ax.imshow(
        masked_matrix,
        aspect="equal",
        interpolation="nearest",
        cmap=color_map,
    )

    tick_positions = np.arange(
        len(
            labels
        )
    )

    ax.set_xticks(
        tick_positions
    )
    ax.set_yticks(
        tick_positions
    )
    ax.set_xticklabels(
        labels,
        rotation=90,
        fontsize=5.2,
    )
    ax.set_yticklabels(
        labels,
        fontsize=5.2,
    )
    ax.tick_params(
        length=0,
        pad=1,
    )

    ax.set_title(
        f"({panel_letter}) "
        f"{variant}—"
        f"{MODEL_ABBREVIATION[model_name]}",
        fontsize=8.3,
        fontweight="bold",
        pad=5,
    )

    colorbar = fig.colorbar(
        image,
        ax=ax,
        fraction=0.046,
        pad=0.03,
    )
    colorbar.ax.tick_params(
        labelsize=5.0
    )
    colorbar.set_label(
        "Mean absolute SHAP interaction",
        fontsize=5.4,
    )

fig.text(
    0.5,
    0.006,
    (
        "Pooled out-of-sample IECV observations. "
        "Diagonal elements were set to zero, and values below "
        "the within-model 80th percentile were masked. "
        "Color scales are panel specific."
    ),
    ha="center",
    va="bottom",
    fontsize=5.8,
)

fig.tight_layout(
    rect=(
        0,
        0.038,
        1,
        1,
    )
)

save_png_compliant(
    fig,
    FIGURE_PATH,
)

plt.show()
plt.close(
    fig
)

print(
    "Figure output:",
    FIGURE_PATH,
)


# ============================================================
# 9. RUN MANIFEST AND FINAL CHECK
# ============================================================

manifest = {
    "project_directory": str(
        PROJECT_DIR
    ),
    "raw_data_path": str(
        RAW_DATA_PATH
    ),
    "raw_data_sheet": RAW_DATA_SHEET,
    "raw_data_sha256": RAW_DATA_SHA256,
    "model_directory": str(
        MODEL_DIR
    ),
    "output_directory": str(
        OUTPUT_DIR
    ),
    "cohort_n": int(
        len(
            df_raw
        )
    ),
    "events": int(
        df_raw[
            OUTCOME_COL
        ].sum()
    ),
    "prevalence": float(
        df_raw[
            OUTCOME_COL
        ].mean()
    ),
    "facility_label_map": (
        facility_label_map
    ),
    "combinations": [
        {
            "feature_set": variant,
            "model": model_name,
        }
        for (
            variant,
            model_name,
        ) in COMBINATIONS
    ],
    "max_patients_per_facility": (
        MAX_PATIENTS_PER_FACILITY
    ),
    "interaction_batch_size": (
        INTERACTION_BATCH_SIZE
    ),
    "interaction_percentile": (
        INTERACTION_PERCENTILE
    ),
    "plot_top_n": PLOT_TOP_N,
    "reuse_existing_cache": (
        REUSE_EXISTING_CACHE
    ),
    "table_path": str(
        TABLE_PATH
    ),
    "figure_path": str(
        FIGURE_PATH
    ),
    "python": sys.version,
    "platform": platform.platform(),
    "package_versions": {
        "numpy": installed_version(
            "numpy"
        ),
        "pandas": installed_version(
            "pandas"
        ),
        "scikit-learn": installed_version(
            "scikit-learn"
        ),
        "xgboost": installed_version(
            "xgboost"
        ),
        "lightgbm": installed_version(
            "lightgbm"
        ),
        "shap": installed_version(
            "shap"
        ),
        "matplotlib": installed_version(
            "matplotlib"
        ),
        "openpyxl": installed_version(
            "openpyxl"
        ),
        "cloudpickle": installed_version(
            "cloudpickle"
        ),
    },
}

save_json(
    manifest,
    MANIFEST_PATH,
)

if not TABLE_PATH.exists():
    raise RuntimeError(
        "The Excel output was not created."
    )

if not FIGURE_PATH.exists():
    raise RuntimeError(
        "The PNG output was not created."
    )

if Image is not None:
    with Image.open(
        FIGURE_PATH
    ) as image:
        width_px, height_px = (
            image.size
        )

    file_size_mb = (
        FIGURE_PATH
        .stat()
        .st_size
        / (
            1024 ** 2
        )
    )

    print(
        "PNG dimensions:",
        f"{width_px} × {height_px} px",
    )
    print(
        "PNG file size:",
        f"{file_size_mb:.3f} MB",
    )

    if (
        width_px > PNG_MAX_WIDTH
        or height_px > PNG_MAX_HEIGHT
        or file_size_mb > PNG_MAX_FILE_SIZE_MB
    ):
        raise RuntimeError(
            "The final PNG does not meet the configured JMIR limits."
        )

print(
    "Run manifest:",
    MANIFEST_PATH,
)
print(
    "\nSHAP interaction analysis completed."
)
