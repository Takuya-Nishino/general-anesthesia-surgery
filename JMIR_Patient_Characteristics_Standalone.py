# %% [markdown]
# # JMIR revision: standalone patient-characteristics tables
#
# This script reads the revision data set and creates:
#
# 1. Main Table 1: selected characteristics, overall and by institution
# 2. Supplementary Table S2: all 43 predictors and outcomes, overall and by institution
# 3. Supplementary Table S3: temporal development vs temporal validation cohorts
# 4. Supplementary Table S4: concise random hold-out split summary and full SMD audit
#
# The tables are generated from the raw, nonimputed data. No model fitting is performed.

# %%
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from hashlib import sha256
from itertools import combinations
from pathlib import Path
from typing import Callable, Iterable, Literal, Sequence

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.model_selection import train_test_split


# ============================================================
# 0. User settings
# ============================================================

RAW_DATA_PATH = Path(
    r"C:\Users\tears\Desktop\Study\2025\03_anesthesiology"
    r"\003_ML\005_JMIR_Revice\Revice_data_20260714.xlsx"
)
RAW_DATA_SHEET: int | str = 0

# Must be the same output root as the final model-analysis notebook.
MAIN_ANALYSIS_OUTPUT_DIR = Path(
    r"C:\Users\tears\OneDrive\Revice_JMIR"
    r"\Reanalysis_FINAL_20260716_Temporal"
)

OUTPUT_DIR = MAIN_ANALYSIS_OUTPUT_DIR / "11_patient_characteristics"
HOLDOUT_SPLIT_PATH = MAIN_ANALYSIS_OUTPUT_DIR / "fixed_holdout_split.csv"

ID_COL = "INDEX"
FACILITY_COL = "付属_1"
PRIMARY_OUTCOME_COL = "Event"
INFECTION_OUTCOME_COL = "PI"
NONINFECTION_OUTCOME_COL = "Hard_endpoint"
TEMPORAL_PERIOD_COL = "Period24_25"

TEMPORAL_DEVELOPMENT_VALUE = 0
TEMPORAL_VALIDATION_VALUE = 1
TEMPORAL_DEVELOPMENT_LABEL = "January 2021–December 2023"
TEMPORAL_VALIDATION_LABEL = "January 2024–March 2025"

RANDOM_STATE = 42
HOLDOUT_TEST_SIZE = 0.20
EXPECTED_N_FACILITIES = 3
SMD_THRESHOLD = 0.10

# Output files
COMBINED_WORKBOOK_NAME = "JMIR_Patient_Characteristics_Tables.xlsx"
TABLE1_WORKBOOK_NAME = "Table1_Selected_Characteristics_by_Facility.xlsx"
TABLES2_WORKBOOK_NAME = "TableS2_Full_Characteristics_by_Facility.xlsx"
TABLES3_WORKBOOK_NAME = "TableS3_Temporal_Cohort_Characteristics.xlsx"
TABLES4_WORKBOOK_NAME = "TableS4_Random_Holdout_Summary.xlsx"


# ============================================================
# 1. Model feature definitions: identical to the final analysis
# ============================================================

BINARY_PREOPERATIVE = (
    "Male",
    "Dialysis",
    "CHF",
    "Malig",
    "DeliMed",
    "β-blocker",
    "Oral steroids",
    "Antiplatelet",
    "Anticoag",
    "AntiCa",
    "Opioid",
    "Proc-Eye",
    "Proc-Face/Neck",
    "Proc-Thorax",
    "Proc-MSK",
    "Proc-ENT",
    "Proc-Neuro",
    "Proc-Genital",
    "Proc-Urinary",
    "Proc-Skin",
    "Proc-Abd",
    "HighRiskProc",
)

ORDINAL_PREOPERATIVE = ("ASA",)
COUNT_PREOPERATIVE = ("ResectNum",)

CONTINUOUS_PREOPERATIVE = (
    "Age",
    "BMI",
    "Alb",
    "BUN",
    "CRP",
    "Cre",
    "Hb",
    "K",
    "Na",
    "PLT",
    "T-Bil",
    "WBC",
)

CONTINUOUS_PERIOPERATIVE_ADDITIONAL = (
    "OpTime",
    "RBC Tx",
    "FFP Tx",
    "PLT Tx",
    "FluidBal",
    "HR at 6h",
    "MAP at 6h",
)

PREOPERATIVE_FEATURES = (
    "Age",
    "Male",
    "BMI",
    "ASA",
    "Dialysis",
    "CHF",
    "Malig",
    "Alb",
    "BUN",
    "CRP",
    "Cre",
    "Hb",
    "K",
    "Na",
    "PLT",
    "T-Bil",
    "WBC",
    "DeliMed",
    "β-blocker",
    "Oral steroids",
    "Antiplatelet",
    "Anticoag",
    "AntiCa",
    "Opioid",
    "Proc-Eye",
    "Proc-Face/Neck",
    "Proc-Thorax",
    "Proc-MSK",
    "Proc-ENT",
    "Proc-Neuro",
    "Proc-Genital",
    "Proc-Urinary",
    "Proc-Skin",
    "Proc-Abd",
    "ResectNum",
    "HighRiskProc",
)

PERIOPERATIVE_FEATURES = PREOPERATIVE_FEATURES + CONTINUOUS_PERIOPERATIVE_ADDITIONAL

if len(PREOPERATIVE_FEATURES) != 36:
    raise AssertionError(f"Expected 36 preoperative predictors, got {len(PREOPERATIVE_FEATURES)}")
if len(PERIOPERATIVE_FEATURES) != 43:
    raise AssertionError(f"Expected 43 perioperative predictors, got {len(PERIOPERATIVE_FEATURES)}")


# ============================================================
# 2. Publication labels and variable specifications
# ============================================================

VariableKind = Literal["continuous", "binary", "ordinal", "count"]


@dataclass(frozen=True)
class VariableSpec:
    key: str
    label: str
    section: str
    kind: VariableKind
    source: str
    decimals: int = 1
    transform: Callable[[pd.Series], pd.Series] | None = None


def identity(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def binary_positive(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    return (numeric > 0).astype(float).where(numeric.notna())


def asa_high(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    return (numeric >= 3).astype(float).where(numeric.notna())


FEATURE_LABELS: dict[str, tuple[str, str, VariableKind, int]] = {
    # Patient characteristics
    "Age": ("Age, years", "Demographic and preoperative status", "continuous", 0),
    "Male": ("Male sex, n (%)", "Demographic and preoperative status", "binary", 1),
    "BMI": ("Body mass index, kg/m²", "Demographic and preoperative status", "continuous", 1),
    "ASA": ("ASA Physical Status", "Demographic and preoperative status", "ordinal", 0),
    # Comorbidities
    "Dialysis": ("Maintenance dialysis, n (%)", "Comorbidities", "binary", 1),
    "CHF": ("Congestive heart failure, n (%)", "Comorbidities", "binary", 1),
    "Malig": ("Malignancy, n (%)", "Comorbidities", "binary", 1),
    # Laboratory data
    "Alb": ("Serum albumin, g/dL", "Preoperative laboratory values", "continuous", 1),
    "BUN": ("Blood urea nitrogen, mg/dL", "Preoperative laboratory values", "continuous", 1),
    "CRP": ("C-reactive protein, mg/dL", "Preoperative laboratory values", "continuous", 2),
    "Cre": ("Serum creatinine, mg/dL", "Preoperative laboratory values", "continuous", 2),
    "Hb": ("Hemoglobin, g/dL", "Preoperative laboratory values", "continuous", 1),
    "K": ("Serum potassium, mEq/L", "Preoperative laboratory values", "continuous", 1),
    "Na": ("Serum sodium, mEq/L", "Preoperative laboratory values", "continuous", 0),
    "PLT": ("Platelet count, ×10⁴/µL", "Preoperative laboratory values", "continuous", 1),
    "T-Bil": ("Total bilirubin, mg/dL", "Preoperative laboratory values", "continuous", 2),
    "WBC": ("White blood cell count, /µL", "Preoperative laboratory values", "continuous", 0),
    # Medications
    "DeliMed": ("Potentially deliriogenic medication, n (%)", "Preoperative medications", "binary", 1),
    "β-blocker": ("Beta-blocker, n (%)", "Preoperative medications", "binary", 1),
    "Oral steroids": ("Oral corticosteroid, n (%)", "Preoperative medications", "binary", 1),
    "Antiplatelet": ("Antiplatelet agent, n (%)", "Preoperative medications", "binary", 1),
    "Anticoag": ("Anticoagulant, n (%)", "Preoperative medications", "binary", 1),
    "AntiCa": ("Calcium-channel blocker, n (%)", "Preoperative medications", "binary", 1),
    "Opioid": ("Opioid, n (%)", "Preoperative medications", "binary", 1),
    # Procedure characteristics
    "Proc-Eye": ("Ophthalmic procedure, n (%)", "Procedure characteristics", "binary", 1),
    "Proc-Face/Neck": ("Face or neck procedure, n (%)", "Procedure characteristics", "binary", 1),
    "Proc-Thorax": ("Thoracic procedure, n (%)", "Procedure characteristics", "binary", 1),
    "Proc-MSK": ("Musculoskeletal procedure, n (%)", "Procedure characteristics", "binary", 1),
    "Proc-ENT": ("Otorhinolaryngologic procedure, n (%)", "Procedure characteristics", "binary", 1),
    "Proc-Neuro": ("Neurosurgical procedure, n (%)", "Procedure characteristics", "binary", 1),
    "Proc-Genital": ("Genital procedure, n (%)", "Procedure characteristics", "binary", 1),
    "Proc-Urinary": ("Urinary tract procedure, n (%)", "Procedure characteristics", "binary", 1),
    "Proc-Skin": ("Skin or soft-tissue procedure, n (%)", "Procedure characteristics", "binary", 1),
    "Proc-Abd": ("Abdominal procedure, n (%)", "Procedure characteristics", "binary", 1),
    "ResectNum": ("Number of planned resections", "Procedure characteristics", "count", 0),
    "HighRiskProc": ("High-risk procedure, n (%)", "Procedure characteristics", "binary", 1),
    # Perioperative data
    "OpTime": ("Operative duration, min", "Intraoperative characteristics", "continuous", 0),
    "RBC Tx": ("Red blood cell transfusion, units", "Intraoperative characteristics", "continuous", 0),
    "FFP Tx": ("Fresh frozen plasma transfusion, units", "Intraoperative characteristics", "continuous", 0),
    "PLT Tx": ("Platelet transfusion, units", "Intraoperative characteristics", "continuous", 0),
    "FluidBal": ("Intraoperative fluid balance, mL", "Intraoperative characteristics", "continuous", 0),
    "HR at 6h": ("Heart rate at 6 hours, beats/min", "Early postoperative physiology", "continuous", 0),
    "MAP at 6h": ("Mean arterial pressure at 6 hours, mmHg", "Early postoperative physiology", "continuous", 0),
}


def direct_spec(variable: str) -> VariableSpec:
    label, section, kind, decimals = FEATURE_LABELS[variable]
    return VariableSpec(
        key=variable,
        label=label,
        section=section,
        kind=kind,
        source=variable,
        decimals=decimals,
        transform=identity,
    )


def outcome_specs() -> list[VariableSpec]:
    return [
        VariableSpec(
            key=PRIMARY_OUTCOME_COL,
            label="Primary composite endpoint, n (%)",
            section="Outcomes",
            kind="binary",
            source=PRIMARY_OUTCOME_COL,
            decimals=1,
            transform=identity,
        ),
        VariableSpec(
            key=INFECTION_OUTCOME_COL,
            label="Postoperative infection, n (%)",
            section="Outcomes",
            kind="binary",
            source=INFECTION_OUTCOME_COL,
            decimals=1,
            transform=identity,
        ),
        VariableSpec(
            key=NONINFECTION_OUTCOME_COL,
            label="Noninfection composite endpoint, n (%)",
            section="Outcomes",
            kind="binary",
            source=NONINFECTION_OUTCOME_COL,
            decimals=1,
            transform=identity,
        ),
    ]


def build_full_specs() -> list[VariableSpec]:
    return outcome_specs() + [direct_spec(variable) for variable in PERIOPERATIVE_FEATURES]


def build_selected_specs(data: pd.DataFrame) -> list[VariableSpec]:
    asa_label = "ASA Physical Status III–V, n (%)" if (data["ASA"] == 5).any() else "ASA Physical Status III–IV, n (%)"

    return outcome_specs() + [
        direct_spec("Age"),
        direct_spec("Male"),
        direct_spec("BMI"),
        VariableSpec(
            key="ASA_high",
            label=asa_label,
            section="Demographic and preoperative status",
            kind="binary",
            source="ASA",
            decimals=1,
            transform=asa_high,
        ),
        direct_spec("Malig"),
        direct_spec("CHF"),
        direct_spec("Dialysis"),
        direct_spec("Hb"),
        direct_spec("Alb"),
        direct_spec("Cre"),
        direct_spec("CRP"),
        direct_spec("Proc-Abd"),
        direct_spec("Proc-Thorax"),
        direct_spec("Proc-MSK"),
        direct_spec("Proc-Urinary"),
        direct_spec("HighRiskProc"),
        direct_spec("ResectNum"),
        direct_spec("OpTime"),
        direct_spec("FluidBal"),
        VariableSpec(
            key="RBC_any",
            label="Red blood cell transfusion, n (%)",
            section="Intraoperative characteristics",
            kind="binary",
            source="RBC Tx",
            decimals=1,
            transform=binary_positive,
        ),
        direct_spec("HR at 6h"),
        direct_spec("MAP at 6h"),
    ]


# ============================================================
# 3. Data validation and split reconstruction
# ============================================================


def file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = sha256()
    with path.open("rb") as file:
        while True:
            chunk = file.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def read_and_validate_data(data_path: Path, sheet_name: int | str = 0) -> pd.DataFrame:
    if not data_path.exists():
        raise FileNotFoundError(f"Raw data file was not found: {data_path}")

    data = pd.read_excel(data_path, sheet_name=sheet_name)
    data.columns = [str(column).strip() for column in data.columns]

    required = {
        ID_COL,
        FACILITY_COL,
        PRIMARY_OUTCOME_COL,
        INFECTION_OUTCOME_COL,
        NONINFECTION_OUTCOME_COL,
        TEMPORAL_PERIOD_COL,
        *PERIOPERATIVE_FEATURES,
    }
    missing = sorted(required - set(data.columns))
    if missing:
        raise KeyError(f"Required columns are missing: {missing}")

    if data[ID_COL].isna().any():
        raise ValueError(f"{ID_COL} contains missing values.")
    if data[ID_COL].duplicated().any():
        examples = data.loc[data[ID_COL].duplicated(keep=False), ID_COL].head(10).tolist()
        raise ValueError(f"{ID_COL} must uniquely identify analytic episodes. Examples: {examples}")
    if data[FACILITY_COL].isna().any():
        raise ValueError(f"{FACILITY_COL} contains missing values.")

    numeric_columns = [
        PRIMARY_OUTCOME_COL,
        INFECTION_OUTCOME_COL,
        NONINFECTION_OUTCOME_COL,
        TEMPORAL_PERIOD_COL,
        *PERIOPERATIVE_FEATURES,
    ]
    for column in numeric_columns:
        original_nonmissing = data[column].notna()
        numeric = pd.to_numeric(data[column], errors="coerce")
        invalid = original_nonmissing & numeric.isna()
        if invalid.any():
            examples = data.loc[invalid, column].head(10).tolist()
            raise ValueError(f"Column {column} contains nonnumeric values. Examples: {examples}")
        if np.isinf(numeric.to_numpy(dtype=float, na_value=np.nan)).any():
            raise ValueError(f"Column {column} contains infinite values.")
        data[column] = numeric

    for column in [
        PRIMARY_OUTCOME_COL,
        INFECTION_OUTCOME_COL,
        NONINFECTION_OUTCOME_COL,
        *BINARY_PREOPERATIVE,
    ]:
        observed = set(data[column].dropna().unique().tolist())
        if not observed.issubset({0, 1}):
            raise ValueError(f"Binary column {column} contains values other than 0/1: {sorted(observed)[:10]}")

    for column in [PRIMARY_OUTCOME_COL, INFECTION_OUTCOME_COL, NONINFECTION_OUTCOME_COL]:
        if data[column].isna().any():
            raise ValueError(f"Outcome column {column} contains missing values.")
        data[column] = data[column].astype(int)

    temporal_values = set(data[TEMPORAL_PERIOD_COL].dropna().unique().tolist())
    expected_temporal = {TEMPORAL_DEVELOPMENT_VALUE, TEMPORAL_VALIDATION_VALUE}
    if temporal_values != expected_temporal:
        raise ValueError(
            f"{TEMPORAL_PERIOD_COL} must contain exactly {sorted(expected_temporal)}; "
            f"observed={sorted(temporal_values)}"
        )
    data[TEMPORAL_PERIOD_COL] = data[TEMPORAL_PERIOD_COL].astype(int)

    expected_primary = (
        (data[INFECTION_OUTCOME_COL] == 1)
        | (data[NONINFECTION_OUTCOME_COL] == 1)
    ).astype(int)
    mismatch = data[PRIMARY_OUTCOME_COL].ne(expected_primary)
    if mismatch.any():
        examples = data.loc[
            mismatch,
            [ID_COL, PRIMARY_OUTCOME_COL, INFECTION_OUTCOME_COL, NONINFECTION_OUTCOME_COL],
        ].head(10)
        raise ValueError(
            "Event is inconsistent with PI or Hard_endpoint. Examples:\n"
            + examples.to_string(index=False)
        )

    facilities = sorted(data[FACILITY_COL].unique().tolist(), key=str)
    if len(facilities) != EXPECTED_N_FACILITIES:
        raise ValueError(
            f"Expected {EXPECTED_N_FACILITIES} facilities, got {len(facilities)}: {facilities}"
        )

    return data.reset_index(drop=True)


def build_facility_label_map(data: pd.DataFrame) -> dict[str, str]:
    facilities = sorted(data[FACILITY_COL].unique().tolist(), key=str)
    return {
        str(facility): chr(ord("A") + index)
        for index, facility in enumerate(facilities)
    }


def load_or_create_holdout_split(
    data: pd.DataFrame,
    split_path: Path,
) -> pd.Series:
    """Return a Series containing Development/Holdout in the raw-data row order."""
    split_path.parent.mkdir(parents=True, exist_ok=True)

    if split_path.exists():
        split = pd.read_csv(split_path)
        required = {ID_COL, "Set"}
        missing = required - set(split.columns)
        if missing:
            raise KeyError(f"Existing hold-out split is missing columns: {sorted(missing)}")

        if split[ID_COL].duplicated().any():
            raise ValueError("Existing hold-out split contains duplicated IDs.")

        # String keys avoid dtype mismatch between Excel and CSV IDs.
        split_key = split[[ID_COL, "Set"]].copy()
        split_key["__id_key"] = split_key[ID_COL].astype(str)
        data_key = pd.DataFrame({"__id_key": data[ID_COL].astype(str)})
        merged = data_key.merge(
            split_key[["__id_key", "Set"]],
            on="__id_key",
            how="left",
            validate="one_to_one",
        )
        if merged["Set"].isna().any():
            missing_ids = data.loc[merged["Set"].isna(), ID_COL].head(10).tolist()
            raise ValueError(f"Existing hold-out split does not contain all analytic episodes: {missing_ids}")
        invalid_sets = set(merged["Set"].unique()) - {"Development", "Holdout"}
        if invalid_sets:
            raise ValueError(f"Existing hold-out split contains invalid labels: {sorted(invalid_sets)}")
        return merged["Set"].reset_index(drop=True)

    indices = np.arange(len(data), dtype=int)
    development_idx, holdout_idx = train_test_split(
        indices,
        test_size=HOLDOUT_TEST_SIZE,
        stratify=data[PRIMARY_OUTCOME_COL].to_numpy(dtype=int),
        random_state=RANDOM_STATE,
    )

    labels = np.full(len(data), "Development", dtype=object)
    labels[holdout_idx] = "Holdout"

    pd.DataFrame({ID_COL: data[ID_COL], "Set": labels}).to_csv(
        split_path,
        index=False,
        encoding="utf-8-sig",
    )
    return pd.Series(labels, index=data.index, name="Set")


# ============================================================
# 4. Summary statistics and standardized mean differences
# ============================================================


def series_for_spec(data: pd.DataFrame, spec: VariableSpec) -> pd.Series:
    raw = data[spec.source]
    transform = spec.transform or identity
    result = transform(raw)
    return pd.to_numeric(result, errors="coerce")


def format_number(value: float, decimals: int) -> str:
    if pd.isna(value):
        return "NA"
    return f"{value:,.{decimals}f}"


def summarize_series(series: pd.Series, spec: VariableSpec) -> str:
    values = pd.to_numeric(series, errors="coerce").dropna()
    if values.empty:
        return "NA"

    if spec.kind == "binary":
        denominator = len(values)
        numerator = int((values == 1).sum())
        percentage = 100.0 * numerator / denominator
        return f"{numerator:,} ({percentage:.1f})"

    median = float(values.median())
    q1 = float(values.quantile(0.25))
    q3 = float(values.quantile(0.75))
    return (
        f"{format_number(median, spec.decimals)} "
        f"[{format_number(q1, spec.decimals)}–{format_number(q3, spec.decimals)}]"
    )


def missing_summary(series: pd.Series) -> str:
    missing_n = int(series.isna().sum())
    missing_pct = 100.0 * missing_n / len(series) if len(series) else np.nan
    return f"{missing_n:,} ({missing_pct:.1f})"


def standardized_mean_difference(
    group_1: pd.Series,
    group_2: pd.Series,
    kind: VariableKind,
) -> float:
    x1 = pd.to_numeric(group_1, errors="coerce").dropna().to_numpy(dtype=float)
    x2 = pd.to_numeric(group_2, errors="coerce").dropna().to_numpy(dtype=float)

    if len(x1) == 0 or len(x2) == 0:
        return np.nan

    if kind == "binary":
        p1 = float(np.mean(x1 == 1))
        p2 = float(np.mean(x2 == 1))
        denominator = np.sqrt((p1 * (1 - p1) + p2 * (1 - p2)) / 2.0)
        difference = p1 - p2
    else:
        mean_1 = float(np.mean(x1))
        mean_2 = float(np.mean(x2))
        variance_1 = float(np.var(x1, ddof=1)) if len(x1) > 1 else 0.0
        variance_2 = float(np.var(x2, ddof=1)) if len(x2) > 1 else 0.0
        denominator = np.sqrt((variance_1 + variance_2) / 2.0)
        difference = mean_1 - mean_2

    if denominator <= np.finfo(float).eps:
        return 0.0 if abs(difference) <= np.finfo(float).eps else np.nan
    return float(difference / denominator)


def maximum_pairwise_smd(
    data: pd.DataFrame,
    spec: VariableSpec,
    facilities: Sequence[object],
) -> float:
    values: list[float] = []
    for facility_1, facility_2 in combinations(facilities, 2):
        s1 = series_for_spec(data.loc[data[FACILITY_COL] == facility_1], spec)
        s2 = series_for_spec(data.loc[data[FACILITY_COL] == facility_2], spec)
        smd = standardized_mean_difference(s1, s2, spec.kind)
        if np.isfinite(smd):
            values.append(abs(smd))
    return max(values) if values else np.nan


def build_sectioned_table(
    data: pd.DataFrame,
    specs: Sequence[VariableSpec],
    group_frames: Sequence[tuple[str, pd.DataFrame]],
    smd_function: Callable[[VariableSpec], float],
    smd_column_name: str,
    include_overall_missing: bool = False,
) -> tuple[pd.DataFrame, list[int]]:
    rows: list[dict[str, object]] = []
    section_rows: list[int] = []
    current_section: str | None = None

    for spec in specs:
        if spec.section != current_section:
            section_rows.append(len(rows))
            row: dict[str, object] = {"Variable": spec.section}
            for group_name, _ in group_frames:
                row[group_name] = ""
            if include_overall_missing:
                row["Overall missing, n (%)"] = ""
            row[smd_column_name] = np.nan
            rows.append(row)
            current_section = spec.section

        row = {"Variable": spec.label}
        for group_name, group_data in group_frames:
            row[group_name] = summarize_series(series_for_spec(group_data, spec), spec)
        if include_overall_missing:
            row["Overall missing, n (%)"] = missing_summary(data[spec.source])
        row[smd_column_name] = smd_function(spec)
        rows.append(row)

    return pd.DataFrame(rows), section_rows


def build_missingness_table(
    data: pd.DataFrame,
    specs: Sequence[VariableSpec],
    group_frames: Sequence[tuple[str, pd.DataFrame]],
) -> tuple[pd.DataFrame, list[int]]:
    rows: list[dict[str, object]] = []
    section_rows: list[int] = []
    current_section: str | None = None

    # Duplicate source columns are shown only once (eg, derived ASA row is not used here).
    seen_sources: set[str] = set()
    for spec in specs:
        if spec.source in seen_sources:
            continue
        seen_sources.add(spec.source)

        if spec.section != current_section:
            section_rows.append(len(rows))
            row = {"Variable": spec.section}
            for group_name, _ in group_frames:
                row[group_name] = ""
            rows.append(row)
            current_section = spec.section

        label = FEATURE_LABELS.get(spec.source, (spec.label, "", spec.kind, spec.decimals))[0]
        row = {"Variable": label.replace(", n (%)", "")}
        for group_name, group_data in group_frames:
            row[group_name] = missing_summary(group_data[spec.source])
        rows.append(row)

    return pd.DataFrame(rows), section_rows


# ============================================================
# 5. Build Table 1 and supplementary patient-background tables
# ============================================================


def facility_groups(
    data: pd.DataFrame,
    facility_map: dict[str, str],
) -> tuple[list[tuple[str, pd.DataFrame]], list[object]]:
    facilities = sorted(data[FACILITY_COL].unique().tolist(), key=str)
    groups: list[tuple[str, pd.DataFrame]] = [(f"Overall (n={len(data):,})", data)]
    for facility in facilities:
        subset = data.loc[data[FACILITY_COL] == facility]
        label = facility_map[str(facility)]
        groups.append((f"Facility {label} (n={len(subset):,})", subset))
    return groups, facilities


def build_facility_tables(
    data: pd.DataFrame,
    facility_map: dict[str, str],
) -> dict[str, object]:
    groups, facilities = facility_groups(data, facility_map)

    selected_specs = build_selected_specs(data)
    full_specs = build_full_specs()

    smd_fn = lambda spec: maximum_pairwise_smd(data, spec, facilities)

    table1, table1_sections = build_sectioned_table(
        data=data,
        specs=selected_specs,
        group_frames=groups,
        smd_function=smd_fn,
        smd_column_name="Maximum absolute pairwise SMD",
        include_overall_missing=False,
    )

    table_s2, table_s2_sections = build_sectioned_table(
        data=data,
        specs=full_specs,
        group_frames=groups,
        smd_function=smd_fn,
        smd_column_name="Maximum absolute pairwise SMD",
        include_overall_missing=True,
    )

    missing_s2, missing_s2_sections = build_missingness_table(
        data=data,
        specs=[direct_spec(variable) for variable in PERIOPERATIVE_FEATURES],
        group_frames=groups,
    )

    return {
        "table1": table1,
        "table1_sections": table1_sections,
        "table_s2": table_s2,
        "table_s2_sections": table_s2_sections,
        "missing_s2": missing_s2,
        "missing_s2_sections": missing_s2_sections,
    }


def build_temporal_tables(data: pd.DataFrame) -> dict[str, object]:
    development = data.loc[data[TEMPORAL_PERIOD_COL] == TEMPORAL_DEVELOPMENT_VALUE]
    validation = data.loc[data[TEMPORAL_PERIOD_COL] == TEMPORAL_VALIDATION_VALUE]

    groups = [
        (f"Development: {TEMPORAL_DEVELOPMENT_LABEL} (n={len(development):,})", development),
        (f"Temporal validation: {TEMPORAL_VALIDATION_LABEL} (n={len(validation):,})", validation),
    ]

    specs = build_full_specs()

    def smd_fn(spec: VariableSpec) -> float:
        return abs(
            standardized_mean_difference(
                series_for_spec(development, spec),
                series_for_spec(validation, spec),
                spec.kind,
            )
        )

    table, section_rows = build_sectioned_table(
        data=data,
        specs=specs,
        group_frames=groups,
        smd_function=smd_fn,
        smd_column_name="Absolute SMD",
        include_overall_missing=False,
    )

    missing, missing_sections = build_missingness_table(
        data=data,
        specs=[direct_spec(variable) for variable in PERIOPERATIVE_FEATURES],
        group_frames=groups,
    )

    return {
        "table": table,
        "sections": section_rows,
        "missing": missing,
        "missing_sections": missing_sections,
    }


def build_holdout_tables(
    data: pd.DataFrame,
    split_labels: pd.Series,
) -> dict[str, pd.DataFrame]:
    if len(split_labels) != len(data):
        raise ValueError("Hold-out split length does not match the analytic data set.")

    working = data.copy()
    working["__Set"] = split_labels.to_numpy()
    development = working.loc[working["__Set"] == "Development"]
    holdout = working.loc[working["__Set"] == "Holdout"]

    summary_rows: list[dict[str, object]] = [
        {
            "Measure": "Surgical admissions, n",
            "Development set": f"{len(development):,}",
            "Hold-out test set": f"{len(holdout):,}",
        }
    ]

    for spec in outcome_specs():
        summary_rows.append(
            {
                "Measure": spec.label,
                "Development set": summarize_series(series_for_spec(development, spec), spec),
                "Hold-out test set": summarize_series(series_for_spec(holdout, spec), spec),
            }
        )

    detail_rows: list[dict[str, object]] = []
    for variable in PERIOPERATIVE_FEATURES:
        spec = direct_spec(variable)
        smd = standardized_mean_difference(
            series_for_spec(development, spec),
            series_for_spec(holdout, spec),
            spec.kind,
        )
        detail_rows.append(
            {
                "Source variable": variable,
                "Predictor": spec.label,
                "Type": spec.kind,
                "Absolute SMD": abs(smd) if np.isfinite(smd) else np.nan,
            }
        )

    detail = pd.DataFrame(detail_rows).sort_values(
        "Absolute SMD",
        ascending=False,
        na_position="last",
    ).reset_index(drop=True)

    finite_smd = detail["Absolute SMD"].dropna()
    max_smd = float(finite_smd.max()) if not finite_smd.empty else np.nan
    max_variable = (
        str(detail.loc[detail["Absolute SMD"].idxmax(), "Predictor"])
        if not finite_smd.empty
        else "NA"
    )

    balance = pd.DataFrame(
        [
            {"Balance measure": "Predictors assessed", "Value": len(PERIOPERATIVE_FEATURES)},
            {"Balance measure": "Maximum absolute SMD", "Value": max_smd},
            {"Balance measure": "Predictor with maximum absolute SMD", "Value": max_variable},
            {
                "Balance measure": f"Predictors with absolute SMD ≥{SMD_THRESHOLD:.2f}",
                "Value": int((finite_smd >= SMD_THRESHOLD).sum()),
            },
            {
                "Balance measure": "Predictors with absolute SMD ≥0.20",
                "Value": int((finite_smd >= 0.20).sum()),
            },
        ]
    )

    return {
        "summary": pd.DataFrame(summary_rows),
        "balance": balance,
        "detail": detail,
    }


# ============================================================
# 6. Excel formatting
# ============================================================

TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="2F75B5")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
LIGHT_FILL = PatternFill("solid", fgColor="F7FAFC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=13)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SECTION_FONT = Font(color="1F1F1F", bold=True)
BODY_FONT = Font(size=10)
FOOTNOTE_FONT = Font(size=9, italic=True, color="404040")
THIN_GRAY = Side(style="thin", color="D9E1F2")
BODY_BORDER = Border(bottom=THIN_GRAY)


@dataclass(frozen=True)
class SheetPayload:
    sheet_name: str
    title: str
    table: pd.DataFrame
    section_rows: Sequence[int] = ()
    footnotes: Sequence[str] = ()
    smd_columns: Sequence[str] = ()


def write_readme_sheet(
    writer: pd.ExcelWriter,
    data_path: Path,
    source_hash: str,
    facility_map: dict[str, str],
    output_dir: Path,
) -> None:
    rows: list[dict[str, str]] = [
        {"Item": "Purpose", "Value": "Publication-ready patient-characteristics tables for the JMIR revision"},
        {"Item": "Source data", "Value": str(data_path)},
        {"Item": "Source SHA-256", "Value": source_hash},
        {"Item": "Generated", "Value": datetime.now().isoformat(timespec="seconds")},
        {"Item": "Output directory", "Value": str(output_dir)},
        {"Item": "Statistical summaries", "Value": "Median [25th–75th percentile] or number (%)"},
        {"Item": "Missing data", "Value": "Raw, nonimputed data; percentages use nonmissing denominators"},
        {"Item": "Facility SMD", "Value": "Maximum absolute pairwise SMD across A vs B, A vs C, and B vs C"},
        {"Item": "Temporal SMD", "Value": "Absolute SMD between 2021–2023 development and 2024–March 2025 validation periods"},
        {"Item": "Random hold-out", "Value": f"Stratified {int((1-HOLDOUT_TEST_SIZE)*100)}:{int(HOLDOUT_TEST_SIZE*100)} split; random_state={RANDOM_STATE}"},
    ]
    for raw_value, anonymous_label in facility_map.items():
        rows.append(
            {
                "Item": f"Facility {anonymous_label}",
                "Value": f"Raw facility value: {raw_value}",
            }
        )

    pd.DataFrame(rows).to_excel(writer, sheet_name="README", index=False)


def style_readme(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    if "README" not in workbook.sheetnames:
        workbook.close()
        return
    sheet = workbook["README"]
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 95
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BODY_BORDER
    workbook.save(workbook_path)
    workbook.close()


def style_table_sheet(
    workbook_path: Path,
    payload: SheetPayload,
) -> None:
    workbook = load_workbook(workbook_path)
    sheet = workbook[payload.sheet_name]

    n_columns = payload.table.shape[1]
    last_column_letter = get_column_letter(n_columns)
    header_row = 3
    first_data_row = 4
    last_data_row = first_data_row + len(payload.table) - 1

    # Title
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_columns)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = payload.title
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 24

    # Header
    for cell in sheet[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=THIN_GRAY, bottom=THIN_GRAY, left=THIN_GRAY, right=THIN_GRAY)
    sheet.row_dimensions[header_row].height = 38

    # Body
    for row in sheet.iter_rows(min_row=first_data_row, max_row=max(last_data_row, first_data_row), min_col=1, max_col=n_columns):
        for index, cell in enumerate(row, start=1):
            cell.font = BODY_FONT
            cell.alignment = Alignment(
                horizontal="left" if index == 1 else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = BODY_BORDER

    # Section headings
    for zero_based_index in payload.section_rows:
        excel_row = first_data_row + int(zero_based_index)
        sheet.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=n_columns)
        section_cell = sheet.cell(row=excel_row, column=1)
        section_cell.fill = SECTION_FILL
        section_cell.font = SECTION_FONT
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[excel_row].height = 20

    # SMD columns
    for column_name in payload.smd_columns:
        if column_name not in payload.table.columns:
            continue
        column_index = payload.table.columns.get_loc(column_name) + 1
        for row_number in range(first_data_row, last_data_row + 1):
            cell = sheet.cell(row=row_number, column=column_index)
            cell.number_format = "0.000"
        sheet.column_dimensions[get_column_letter(column_index)].width = 18

    # Column widths
    sheet.column_dimensions["A"].width = 42
    for column_index in range(2, n_columns + 1):
        letter = get_column_letter(column_index)
        if sheet.column_dimensions[letter].width is None or sheet.column_dimensions[letter].width < 18:
            sheet.column_dimensions[letter].width = 23

    sheet.freeze_panes = "B4"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A{header_row}:{last_column_letter}{last_data_row}"

    # Footnotes
    footnote_start = last_data_row + 2
    for offset, note in enumerate(payload.footnotes):
        row_number = footnote_start + offset
        sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=n_columns)
        cell = sheet.cell(row=row_number, column=1)
        cell.value = note
        cell.font = FOOTNOTE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        sheet.row_dimensions[row_number].height = 24

    workbook.save(workbook_path)
    workbook.close()


def write_workbook(
    path: Path,
    payloads: Sequence[SheetPayload],
    *,
    data_path: Path | None = None,
    source_hash: str | None = None,
    facility_map: dict[str, str] | None = None,
    output_dir: Path | None = None,
    include_readme: bool = False,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if include_readme:
            if data_path is None or source_hash is None or facility_map is None or output_dir is None:
                raise ValueError("README metadata is incomplete.")
            write_readme_sheet(writer, data_path, source_hash, facility_map, output_dir)

        for payload in payloads:
            payload.table.to_excel(
                writer,
                sheet_name=payload.sheet_name,
                index=False,
                startrow=2,
            )

    if include_readme:
        style_readme(path)
    for payload in payloads:
        style_table_sheet(path, payload)


# ============================================================
# 7. Main execution
# ============================================================


def run_patient_characteristics(
    *,
    raw_data_path: Path = RAW_DATA_PATH,
    raw_data_sheet: int | str = RAW_DATA_SHEET,
    main_analysis_output_dir: Path = MAIN_ANALYSIS_OUTPUT_DIR,
    output_dir: Path = OUTPUT_DIR,
) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    split_path = main_analysis_output_dir / HOLDOUT_SPLIT_PATH.name

    data = read_and_validate_data(raw_data_path, raw_data_sheet)
    source_hash = file_sha256(raw_data_path)
    facility_map = build_facility_label_map(data)
    split_labels = load_or_create_holdout_split(data, split_path)

    facility_outputs = build_facility_tables(data, facility_map)
    temporal_outputs = build_temporal_tables(data)
    holdout_outputs = build_holdout_tables(data, split_labels)

    common_facility_notes = (
        "Values are median [25th–75th percentile] or number (%).",
        "Percentages were calculated among nonmissing observations; no imputation was applied to descriptive summaries.",
        "The maximum absolute pairwise standardized mean difference (SMD) is the largest value among Facility A vs B, A vs C, and B vs C.",
        "Facility labels were anonymized. Outcome categories were not necessarily mutually exclusive.",
    )
    temporal_notes = (
        "Values are median [25th–75th percentile] or number (%).",
        "Percentages were calculated among nonmissing observations; no imputation was applied to descriptive summaries.",
        f"Development period: {TEMPORAL_DEVELOPMENT_LABEL}; temporal validation period: {TEMPORAL_VALIDATION_LABEL}.",
        "Absolute SMD values describe differences in case mix and should not be interpreted as hypothesis tests.",
    )
    missing_notes = (
        "Values are missing observations, n (%), calculated within each displayed cohort.",
        "These values were calculated before any missing-value imputation.",
    )
    holdout_notes = (
        f"The random hold-out split was stratified by the primary endpoint with test_size={HOLDOUT_TEST_SIZE:.2f} and random_state={RANDOM_STATE}.",
        "Predictor balance was assessed across all 43 model predictors using absolute standardized mean differences.",
        "No full random hold-out baseline table was generated because the split was a secondary within-cohort validation; the full SMD audit is provided in a separate sheet.",
    )

    table1_payload = SheetPayload(
        sheet_name="Table1_Main",
        title="Table 1. Selected Demographic and Clinical Characteristics of the Overall Cohort and Participating Institutions",
        table=facility_outputs["table1"],
        section_rows=facility_outputs["table1_sections"],
        footnotes=common_facility_notes,
        smd_columns=("Maximum absolute pairwise SMD",),
    )
    table_s2_payload = SheetPayload(
        sheet_name="TableS2_Facility",
        title="Supplementary Table S2. Complete Predictor and Outcome Distributions in the Overall Cohort and by Institution",
        table=facility_outputs["table_s2"],
        section_rows=facility_outputs["table_s2_sections"],
        footnotes=common_facility_notes,
        smd_columns=("Maximum absolute pairwise SMD",),
    )
    missing_s2_payload = SheetPayload(
        sheet_name="TableS2_Missing",
        title="Supplementary Table S2B. Predictor Missingness in the Overall Cohort and by Institution",
        table=facility_outputs["missing_s2"],
        section_rows=facility_outputs["missing_s2_sections"],
        footnotes=missing_notes,
    )
    table_s3_payload = SheetPayload(
        sheet_name="TableS3_Temporal",
        title="Supplementary Table S3. Characteristics of the Temporal Development and Validation Cohorts",
        table=temporal_outputs["table"],
        section_rows=temporal_outputs["sections"],
        footnotes=temporal_notes,
        smd_columns=("Absolute SMD",),
    )
    missing_s3_payload = SheetPayload(
        sheet_name="TableS3_Missing",
        title="Supplementary Table S3B. Predictor Missingness by Temporal Period",
        table=temporal_outputs["missing"],
        section_rows=temporal_outputs["missing_sections"],
        footnotes=missing_notes,
    )
    holdout_summary_payload = SheetPayload(
        sheet_name="TableS4_Holdout",
        title="Supplementary Table S4. Summary of the Secondary Random Hold-out Split",
        table=holdout_outputs["summary"],
        footnotes=holdout_notes,
    )
    holdout_balance_payload = SheetPayload(
        sheet_name="TableS4_Balance",
        title="Supplementary Table S4B. Predictor-Balance Summary for the Secondary Random Hold-out Split",
        table=holdout_outputs["balance"],
        footnotes=("SMD indicates standardized mean difference.",),
        smd_columns=("Value",),
    )
    holdout_detail_payload = SheetPayload(
        sheet_name="TableS4_SMD_Detail",
        title="Supplementary Table S4C. Full Standardized Mean Difference Audit for the Secondary Random Hold-out Split",
        table=holdout_outputs["detail"],
        footnotes=("Absolute SMD values were calculated on the raw, nonimputed predictor data.",),
        smd_columns=("Absolute SMD",),
    )

    output_paths = {
        "combined": output_dir / COMBINED_WORKBOOK_NAME,
        "table1": output_dir / TABLE1_WORKBOOK_NAME,
        "table_s2": output_dir / TABLES2_WORKBOOK_NAME,
        "table_s3": output_dir / TABLES3_WORKBOOK_NAME,
        "table_s4": output_dir / TABLES4_WORKBOOK_NAME,
    }

    all_payloads = [
        table1_payload,
        table_s2_payload,
        missing_s2_payload,
        table_s3_payload,
        missing_s3_payload,
        holdout_summary_payload,
        holdout_balance_payload,
        holdout_detail_payload,
    ]

    write_workbook(
        output_paths["combined"],
        all_payloads,
        data_path=raw_data_path,
        source_hash=source_hash,
        facility_map=facility_map,
        output_dir=output_dir,
        include_readme=True,
    )
    write_workbook(output_paths["table1"], [table1_payload])
    write_workbook(output_paths["table_s2"], [table_s2_payload, missing_s2_payload])
    write_workbook(output_paths["table_s3"], [table_s3_payload, missing_s3_payload])
    write_workbook(
        output_paths["table_s4"],
        [holdout_summary_payload, holdout_balance_payload, holdout_detail_payload],
    )

    # Machine-readable audit files
    facility_mapping_path = output_dir / "Facility_Label_Mapping.csv"
    pd.DataFrame(
        [
            {"Raw facility value": raw, "Anonymized label": f"Facility {anonymous}"}
            for raw, anonymous in facility_map.items()
        ]
    ).to_csv(facility_mapping_path, index=False, encoding="utf-8-sig")

    holdout_smd_path = output_dir / "Random_Holdout_SMD_Detail.csv"
    holdout_outputs["detail"].to_csv(holdout_smd_path, index=False, encoding="utf-8-sig")

    print("Patient-characteristics tables were generated successfully.")
    print(f"Source data: {raw_data_path}")
    print(f"Source SHA-256: {source_hash}")
    print(f"N: {len(data):,}")
    print(
        f"Primary events: {int(data[PRIMARY_OUTCOME_COL].sum()):,} "
        f"({100 * data[PRIMARY_OUTCOME_COL].mean():.1f}%)"
    )
    print("Facility mapping:")
    for raw, anonymous in facility_map.items():
        print(f"  Facility {anonymous}: raw value {raw}")
    print("Outputs:")
    for name, path in output_paths.items():
        print(f"  {name}: {path}")

    return output_paths


# %%
if __name__ == "__main__":
    run_patient_characteristics()
