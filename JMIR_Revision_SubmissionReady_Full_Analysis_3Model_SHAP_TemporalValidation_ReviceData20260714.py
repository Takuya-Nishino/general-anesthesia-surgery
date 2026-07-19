from __future__ import annotations

# %% [markdown]
# Cell 0
# # JMIR Revision：投稿用最終統合解析Notebook
# 
# 本Notebookは、未補完の原データから以下を一括実行します。
# 
# 1. 最終特徴量辞書（Supplementary Table S1）の作成  
# 2. 主解析：`Event`を対象とした施設単位internal-external cross-validation（IECV、主要検証）  
# 3. 二次的検証：層化random hold-out（80:20）  
# 4. 時間的検証：2021年1月–2023年12月で開発し、2024年1月–2025年3月で評価  
# 5. 3モデル比較：Logistic regression、XGBoost、LightGBM  
# 6. 感度分析：`PI`（術後感染）および`Hard_endpoint`（感染除外複合アウトカム）  
# 7. 多重代入感度分析：`Event`、LightGBM、IECV、10補完データセット  
# 8. AUROC、PR-AUC（average precision）、Brier score、scaled Brier score、較正、DCA  
# 9. IECV out-of-sample SHAP：3モデル × 術前・周術期特徴量セット  
# 10. 投稿用Table、Figure、caption、PNG適合性レポートの出力  
# 
# ## 最終解析の固定仕様
# 
# - Optuna：Logistic regression 20 trial、XGBoost 50 trial、LightGBM 50 trial
# - IECV、random hold-out、temporal validationの各development data内で個別に最適化
# - Complete-case analysisは実施しない
# - Warm startを使用しない
# - 前処理、ハイパーパラメータ最適化、較正は各development data内に限定
# - 旧キャッシュを再利用せず、指定出力先へ最終解析を再構築
# - 個別構成要素別解析は実施しない
# - 投稿FigureはNotebook最終セルの統一コードのみで作成
# 
# 出力先：
# 
# `C:\Users\tears\OneDrive\Revice_JMIR\Reanalysis_FINAL_20260716_Temporal`
# 
# ## 2026-07-16 final code audit
# 
# - 施設ラベルは原データの施設コードからA/B/Cへ一貫して匿名化
# - 最終Figureセルは解析ルート`OUTPUT_DIR`を上書きしない
# - Figureセルの症例ID参照は`ID_COL`へ統一
# - 投稿用TableとFigureは同一の施設ラベル定義を使用

# %%
# Cell 1
# ============================================================
# 0. 解析設定
# ============================================================

from pathlib import Path

ANALYSIS_VERSION = "JMIR_revision_submission_ready_v4_temporal_revice_data_20260714"

# 原データ
RAW_DATA_PATH = Path(
    r"C:\Users\tears\Desktop\Study\2025\03_anesthesiology"
    r"\003_ML\005_JMIR_Revice\Revice_data_20260714.xlsx"
)
RAW_DATA_SHEET = 0

# 最終解析専用の新規出力先
OUTPUT_DIR = Path(
    r"C:\Users\tears\OneDrive\Revice_JMIR"
    r"\Reanalysis_FINAL_20260716_Temporal"
)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 出力サブフォルダ
FEATURE_DIR = OUTPUT_DIR / "00_feature_dictionary"
CACHE_DIR = OUTPUT_DIR / "01_cache"
PARAMETER_DIR = OUTPUT_DIR / "02_hyperparameters"
MODEL_DIR = OUTPUT_DIR / "03_models"
PREDICTION_DIR = OUTPUT_DIR / "04_predictions"
METRIC_DIR = OUTPUT_DIR / "05_metrics"
MI_DIR = OUTPUT_DIR / "06_multiple_imputation"
SHAP_DIR = OUTPUT_DIR / "07_shap"
PUBLICATION_TABLE_DIR = OUTPUT_DIR / "08_publication_tables"
PUBLICATION_FIGURE_DIR = OUTPUT_DIR / "09_publication_figures"
LOG_DIR = OUTPUT_DIR / "10_logs"

for directory in (
    FEATURE_DIR,
    CACHE_DIR,
    PARAMETER_DIR,
    MODEL_DIR,
    PREDICTION_DIR,
    METRIC_DIR,
    MI_DIR,
    SHAP_DIR,
    PUBLICATION_TABLE_DIR,
    PUBLICATION_FIGURE_DIR,
    LOG_DIR,
):
    directory.mkdir(parents=True, exist_ok=True)

FEATURE_TABLE_PATH = FEATURE_DIR / "TableS1_Feature_Dictionary.xlsx"

ID_COL = "INDEX"
FACILITY_COL = "付属_1"
EXPECTED_N_FACILITIES = 3

PRIMARY_OUTCOME_COL = "Event"
INFECTION_OUTCOME_COL = "PI"
NONINFECTION_OUTCOME_COL = "Hard_endpoint"

# Temporal validation indicator:
#   0 = January 2021 through December 2023 (development period)
#   1 = January 2024 through March 2025 (temporal validation period)
TEMPORAL_PERIOD_COL = "Period24_25"
TEMPORAL_DEVELOPMENT_VALUE = 0
TEMPORAL_VALIDATION_VALUE = 1
TEMPORAL_DEVELOPMENT_LABEL = "January 2021–December 2023"
TEMPORAL_VALIDATION_LABEL = "January 2024–March 2025"

PRIMARY_OUTCOME_LABEL = "Primary_composite"
INFECTION_OUTCOME_LABEL = "Postoperative_infection"
NONINFECTION_OUTCOME_LABEL = "Noninfection_composite"

OUTCOME_SPECS = {
    PRIMARY_OUTCOME_LABEL: PRIMARY_OUTCOME_COL,
    INFECTION_OUTCOME_LABEL: INFECTION_OUTCOME_COL,
    NONINFECTION_OUTCOME_LABEL: NONINFECTION_OUTCOME_COL,
}

PRIMARY_MODELS = ("Logistic", "XGBoost", "LightGBM")
SENSITIVITY_MODELS = ("LightGBM",)
VARIANTS = ("Preoperative", "Perioperative")

# 解析実行スイッチ
RUN_PRIMARY_IECV = True
RUN_PRIMARY_HOLDOUT = True
RUN_PRIMARY_TEMPORAL = True
RUN_SENSITIVITY_IECV = True
RUN_SENSITIVITY_HOLDOUT = True
RUN_MULTIPLE_IMPUTATION = True
RUN_PUBLICATION_TABLES = True
RUN_SHAP = True
RUN_FINAL_FIGURES = True

# 乱数・交差検証
RANDOM_STATE = 42
IMPUTER_RANDOM_STATE = 0

N_SPLITS_TUNING = 5
N_SPLITS_CALIBRATION = 3
N_SPLITS_MI_CALIBRATION = 3

# 最終解析のモデル別Optuna予算
N_TRIALS_IECV_BY_MODEL = {
    "Logistic": 20,
    "XGBoost": 50,
    "LightGBM": 50,
}
N_TRIALS_HOLDOUT_BY_MODEL = {
    "Logistic": 20,
    "XGBoost": 50,
    "LightGBM": 50,
}
N_TRIALS_TEMPORAL_BY_MODEL = {
    "Logistic": 20,
    "XGBoost": 50,
    "LightGBM": 50,
}

# 探索収束の診断（自動的なtrial追加は行わず、ログに記録する）
OPTUNA_CONVERGENCE_WINDOW = 10
OPTUNA_LATE_IMPROVEMENT_WARNING = 0.001

N_BOOTSTRAP_PRIMARY = 1000
N_BOOTSTRAP_SENSITIVITY = 500
N_BOOTSTRAP_MI_POOLED = 1000

N_IMPUTATIONS = 10

# 主解析の単一反復補完
SINGLE_IMPUTER_MAX_ITER = 5
SINGLE_IMPUTER_N_ESTIMATORS = 50
SINGLE_IMPUTER_LEARNING_RATE = 0.1

# 多重代入
MI_IMPUTER_MAX_ITER = 5
MI_N_NEAREST_FEATURES = 20

# CPU並列：Optuna trialは逐次実行し、各モデル内部を並列化する
MODEL_N_JOBS = -1
OPTUNA_N_JOBS = 1

# 最終解析では旧キャッシュ・旧モデルを再利用しない。
# 中断後に同一設定で再開する場合のみTrueへ変更する。
REUSE_EXISTING_OUTPUTS = False
SKIP_COMPLETED = REUSE_EXISTING_OUTPUTS
FORCE_RETUNE = not REUSE_EXISTING_OUTPUTS
FORCE_REBUILD_PREPROCESSING = not REUSE_EXISTING_OUTPUTS

# Warm startは最終解析で使用しない
USE_WARM_START = False
WARM_START_PARAMETER_DIR = Path(
    r"C:\Users\tears\Desktop\Study\2025\03_anesthesiology"
    r"\003_ML\005_JMIR_Revice\Results\best_parameters"
)

# SHAP：主要アウトカム、術前・周術期、3アルゴリズム
SHAP_MODELS = ("Logistic", "XGBoost", "LightGBM")
SHAP_VARIANTS = ("Preoperative", "Perioperative")
SHAP_MAX_PER_FACILITY = 2000
SHAP_BACKGROUND_N = 200
SHAP_TOP_N = 20
SHAP_SAVE_VALUES = True

# JMIR投稿システムのPNG制約
PNG_MAX_WIDTH = 1200
PNG_MAX_HEIGHT = 1200
PNG_MAX_FILE_SIZE_MB = 5.0
PNG_EXPORT_DPI = 180

# 動作確認時のみTrue
FAST_TEST_MODE = False

if FAST_TEST_MODE:
    N_TRIALS_IECV_BY_MODEL = {
        model_name: 2 for model_name in PRIMARY_MODELS
    }
    N_TRIALS_HOLDOUT_BY_MODEL = {
        model_name: 2 for model_name in PRIMARY_MODELS
    }
    N_TRIALS_TEMPORAL_BY_MODEL = {
        model_name: 2 for model_name in PRIMARY_MODELS
    }
    N_BOOTSTRAP_PRIMARY = 100
    N_BOOTSTRAP_SENSITIVITY = 100
    N_BOOTSTRAP_MI_POOLED = 100
    N_IMPUTATIONS = 2
    SHAP_MAX_PER_FACILITY = 300

# %%
# Cell 2
# ============================================================
# 1. Imports
# ============================================================


import hashlib
import importlib.metadata
import json
import math
import os
import platform
import random
import re
import sys
import time
import warnings
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import cloudpickle
import joblib
import lightgbm as lgb
import matplotlib.pyplot as plt
import numpy as np
import optuna
import pandas as pd
import shap
import xgboost as xgb

from IPython.display import display
from PIL import Image
from lightgbm import LGBMRegressor
from scipy.special import expit
from scipy.stats import norm

from sklearn.base import BaseEstimator, TransformerMixin
from sklearn.calibration import calibration_curve
from sklearn.experimental import enable_iterative_imputer  # noqa: F401
from sklearn.impute import IterativeImputer
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import BayesianRidge, LogisticRegression
from sklearn.metrics import (
    average_precision_score,
    brier_score_loss,
    precision_recall_curve,
    roc_auc_score,
    roc_curve,
)
from sklearn.model_selection import StratifiedKFold, train_test_split
from sklearn.preprocessing import RobustScaler

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

np.random.seed(RANDOM_STATE)
random.seed(RANDOM_STATE)

# 重要な収束警告等を隠さない。将来互換性に関する警告のみ抑制する。
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings(
    "ignore",
    message=r"X does not have valid feature names, but LGBM.* was fitted with feature names",
    category=UserWarning,
)
optuna.logging.set_verbosity(optuna.logging.WARNING)

plt.rcParams.update(
    {
        "font.family": "sans-serif",
        "font.sans-serif": ["DejaVu Sans", "Arial", "Helvetica"],
        "axes.titlesize": 11,
        "axes.labelsize": 10,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "legend.fontsize": 8,
    }
)

# %%
# Cell 3
# ============================================================
# 2. 原データ読み込み・基本検証
# ============================================================


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        while True:
            chunk = file.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


if not RAW_DATA_PATH.exists():
    raise FileNotFoundError(f"原データが見つかりません: {RAW_DATA_PATH}")

SOURCE_FILE_SHA256 = sha256_file(RAW_DATA_PATH)

df_raw = pd.read_excel(
    RAW_DATA_PATH,
    sheet_name=RAW_DATA_SHEET,
).copy()

required_core_columns = {
    ID_COL,
    FACILITY_COL,
    PRIMARY_OUTCOME_COL,
    INFECTION_OUTCOME_COL,
    NONINFECTION_OUTCOME_COL,
}
if RUN_PRIMARY_TEMPORAL:
    required_core_columns.add(TEMPORAL_PERIOD_COL)

missing_core_columns = sorted(required_core_columns - set(df_raw.columns))
if missing_core_columns:
    raise KeyError(f"必須列が不足しています: {missing_core_columns}")

if df_raw[ID_COL].isna().any():
    raise ValueError(f"{ID_COL}に欠測があります。")

if df_raw[ID_COL].duplicated().any():
    duplicated_ids = (
        df_raw.loc[df_raw[ID_COL].duplicated(), ID_COL]
        .head(10)
        .tolist()
    )
    raise ValueError(f"{ID_COL}に重複があります。例: {duplicated_ids}")

if df_raw[FACILITY_COL].isna().any():
    raise ValueError(f"{FACILITY_COL}に欠測があります。")

df_raw[FACILITY_COL] = df_raw[FACILITY_COL].astype(str)

for outcome_col in (
    PRIMARY_OUTCOME_COL,
    INFECTION_OUTCOME_COL,
    NONINFECTION_OUTCOME_COL,
):
    numeric_outcome = pd.to_numeric(df_raw[outcome_col], errors="coerce")

    if numeric_outcome.isna().any():
        raise ValueError(
            f"アウトカム列 {outcome_col} に欠測または非数値があります。"
        )

    values = set(numeric_outcome.unique())
    if not values.issubset({0, 1}):
        raise ValueError(
            f"アウトカム列 {outcome_col} に0/1以外の値があります: "
            f"{sorted(values)[:10]}"
        )

    df_raw[outcome_col] = numeric_outcome.astype(int)

if RUN_PRIMARY_TEMPORAL:
    temporal_period = pd.to_numeric(
        df_raw[TEMPORAL_PERIOD_COL],
        errors="coerce",
    )
    if temporal_period.isna().any():
        raise ValueError(
            f"時間的検証列 {TEMPORAL_PERIOD_COL} に欠測または非数値があります。"
        )
    temporal_values = set(temporal_period.unique())
    expected_temporal_values = {
        TEMPORAL_DEVELOPMENT_VALUE,
        TEMPORAL_VALIDATION_VALUE,
    }
    if temporal_values != expected_temporal_values:
        raise ValueError(
            f"時間的検証列 {TEMPORAL_PERIOD_COL} は"
            f"{sorted(expected_temporal_values)}の2値である必要があります: "
            f"observed={sorted(temporal_values)}"
        )
    df_raw[TEMPORAL_PERIOD_COL] = temporal_period.astype(int)

# Eventは、感染単独または感染除外複合アウトカムのいずれかを満たす症例と整合する必要がある。
expected_primary = (
    (df_raw[INFECTION_OUTCOME_COL] == 1)
    | (df_raw[NONINFECTION_OUTCOME_COL] == 1)
).astype(int)
outcome_mismatch = df_raw[PRIMARY_OUTCOME_COL].ne(expected_primary)
if outcome_mismatch.any():
    mismatch_examples = df_raw.loc[
        outcome_mismatch,
        [
            ID_COL,
            PRIMARY_OUTCOME_COL,
            INFECTION_OUTCOME_COL,
            NONINFECTION_OUTCOME_COL,
        ],
    ].head(10)
    raise ValueError(
        "EventとPI/Hard_endpointの論理関係が一致しません。例:\n"
        + mismatch_examples.to_string(index=False)
    )

facilities = sorted(df_raw[FACILITY_COL].unique().tolist())
if len(facilities) != EXPECTED_N_FACILITIES:
    raise ValueError(
        f"施設数が想定と異なります: observed={len(facilities)}, "
        f"expected={EXPECTED_N_FACILITIES}, facilities={facilities}"
    )

# 原データの施設コードを、投稿原稿・Figureで用いる匿名化ラベルA/B/Cへ固定する。
# 並び順は施設コードの昇順とし、Table、Figure、manifestで共通利用する。
FACILITY_LABEL_MAP = {
    str(facility): chr(ord("A") + index)
    for index, facility in enumerate(facilities)
}


def facility_display_label(value: Any) -> str:
    text = str(value)
    return FACILITY_LABEL_MAP.get(text, text)


SOURCE_SIGNATURE = SOURCE_FILE_SHA256[:16]
CACHE_RUN_DIR = CACHE_DIR / SOURCE_SIGNATURE
CACHE_RUN_DIR.mkdir(parents=True, exist_ok=True)

print(f"N = {len(df_raw):,}")
print(f"Facilities = {facilities}")
print(
    f"Primary events = {int(df_raw[PRIMARY_OUTCOME_COL].sum()):,} "
    f"({df_raw[PRIMARY_OUTCOME_COL].mean() * 100:.1f}%)"
)
if RUN_PRIMARY_TEMPORAL:
    temporal_development_n = int(
        (df_raw[TEMPORAL_PERIOD_COL] == TEMPORAL_DEVELOPMENT_VALUE).sum()
    )
    temporal_validation_n = int(
        (df_raw[TEMPORAL_PERIOD_COL] == TEMPORAL_VALIDATION_VALUE).sum()
    )
    print(
        "Temporal periods: "
        f"development={temporal_development_n:,} "
        f"({TEMPORAL_DEVELOPMENT_LABEL}); "
        f"validation={temporal_validation_n:,} "
        f"({TEMPORAL_VALIDATION_LABEL})"
    )
print(f"Source SHA256 = {SOURCE_FILE_SHA256}")
print(f"Output directory = {OUTPUT_DIR}")

# %% [markdown]
# Cell 4
# ## 3. 最終モデルの特徴量定義
# 
# - 術前モデル：36特徴量
# - 周術期モデル：43特徴量
# - 施設識別子、患者識別子、アウトカム列は予測因子に含めない
# - 二値変数にTarget Encodingは使用しない
# - `ASA`は順序変数、`ResectNum`はカウント変数として扱う

# %%
# Cell 5
# ============================================================
# 3. 特徴量セット
# ============================================================

BINARY_PREOPERATIVE = (
    "Male",
    "Dialysis",
    "CHF",
    "Malig",
    "DeliMed",
    "β-blocker",
    "Oral steroids",
    "Antiplatelet",
    "Anticoag",
    "AntiCa",
    "Opioid",
    "Proc-Eye",
    "Proc-Face/Neck",
    "Proc-Thorax",
    "Proc-MSK",
    "Proc-ENT",
    "Proc-Neuro",
    "Proc-Genital",
    "Proc-Urinary",
    "Proc-Skin",
    "Proc-Abd",
    "HighRiskProc",
)

ORDINAL_PREOPERATIVE = ("ASA",)
COUNT_PREOPERATIVE = ("ResectNum",)

CONTINUOUS_PREOPERATIVE = (
    "Age",
    "BMI",
    "Alb",
    "BUN",
    "CRP",
    "Cre",
    "Hb",
    "K",
    "Na",
    "PLT",
    "T-Bil",
    "WBC",
)

CONTINUOUS_PERIOPERATIVE_ADDITIONAL = (
    "OpTime",
    "RBC Tx",
    "FFP Tx",
    "PLT Tx",
    "FluidBal",
    "HR at 6h",
    "MAP at 6h",
)

PREOPERATIVE_FEATURES = (
    "Age",
    "Male",
    "BMI",
    "ASA",
    "Dialysis",
    "CHF",
    "Malig",
    "Alb",
    "BUN",
    "CRP",
    "Cre",
    "Hb",
    "K",
    "Na",
    "PLT",
    "T-Bil",
    "WBC",
    "DeliMed",
    "β-blocker",
    "Oral steroids",
    "Antiplatelet",
    "Anticoag",
    "AntiCa",
    "Opioid",
    "Proc-Eye",
    "Proc-Face/Neck",
    "Proc-Thorax",
    "Proc-MSK",
    "Proc-ENT",
    "Proc-Neuro",
    "Proc-Genital",
    "Proc-Urinary",
    "Proc-Skin",
    "Proc-Abd",
    "ResectNum",
    "HighRiskProc",
)

PERIOPERATIVE_FEATURES = PREOPERATIVE_FEATURES + CONTINUOUS_PERIOPERATIVE_ADDITIONAL

ALL_BINARY_FEATURES = tuple(BINARY_PREOPERATIVE)
ALL_ORDINAL_FEATURES = tuple(ORDINAL_PREOPERATIVE)
ALL_COUNT_FEATURES = tuple(COUNT_PREOPERATIVE)
ALL_CONTINUOUS_FEATURES = (
    *CONTINUOUS_PREOPERATIVE,
    *CONTINUOUS_PERIOPERATIVE_ADDITIONAL,
)

VARIANT_FEATURES = {
    "Preoperative": PREOPERATIVE_FEATURES,
    "Perioperative": PERIOPERATIVE_FEATURES,
}

EXCLUDED_COLUMNS = {
    ID_COL: "Patient or surgical-case identifier",
    FACILITY_COL: "Institutional identifier",
    PRIMARY_OUTCOME_COL: "Primary composite outcome",
    INFECTION_OUTCOME_COL: "Postoperative infection outcome",
    NONINFECTION_OUTCOME_COL: "Noninfection composite outcome",
    TEMPORAL_PERIOD_COL: "Temporal validation period indicator",
}

if len(PREOPERATIVE_FEATURES) != 36:
    raise AssertionError(
        f"術前特徴量数が36ではありません: {len(PREOPERATIVE_FEATURES)}"
    )

if len(PERIOPERATIVE_FEATURES) != 43:
    raise AssertionError(
        f"周術期特徴量数が43ではありません: {len(PERIOPERATIVE_FEATURES)}"
    )

if len(set(PERIOPERATIVE_FEATURES)) != len(PERIOPERATIVE_FEATURES):
    raise AssertionError("特徴量名に重複があります。")

if set(PERIOPERATIVE_FEATURES) & set(OUTCOME_SPECS.values()):
    raise AssertionError("特徴量セットにアウトカム列が含まれています。")

missing_feature_columns = sorted(set(PERIOPERATIVE_FEATURES) - set(df_raw.columns))
if missing_feature_columns:
    raise KeyError(f"原データに存在しない特徴量があります: {missing_feature_columns}")

# すべてのモデル入力を数値化し、非数値・無限値を解析前に検出する。
for column in PERIOPERATIVE_FEATURES:
    original_nonmissing = df_raw[column].notna()
    numeric = pd.to_numeric(df_raw[column], errors="coerce")
    invalid_nonmissing = original_nonmissing & numeric.isna()
    if invalid_nonmissing.any():
        examples = df_raw.loc[invalid_nonmissing, column].head(10).tolist()
        raise ValueError(f"特徴量 {column} に非数値があります。例: {examples}")
    if np.isinf(numeric.to_numpy(dtype=float, na_value=np.nan)).any():
        raise ValueError(f"特徴量 {column} に無限値があります。")
    df_raw[column] = numeric

for column in ALL_BINARY_FEATURES:
    values = set(df_raw[column].dropna().unique())
    if not values.issubset({0, 1}):
        raise ValueError(
            f"二値特徴量 {column} に0/1以外の値があります: "
            f"{sorted(values)[:10]}"
        )

asa_values = df_raw["ASA"].dropna()
if not asa_values.between(1, 5).all():
    raise ValueError("ASAに1–5の範囲外の値があります。")

resect_values = df_raw["ResectNum"].dropna()
if (resect_values < 0).any():
    raise ValueError("ResectNumに負値があります。")

print(f"Preoperative features: {len(PREOPERATIVE_FEATURES)}")
print(f"Perioperative features: {len(PERIOPERATIVE_FEATURES)}")

# %%
# Cell 6
# ============================================================
# 4. Supplementary Table用メタデータ
# ============================================================

PREOP_LAB_WINDOW_EN = (
    "Within 60 days before surgery; the measurement closest to the admission date was used"
)
PREOP_LAB_WINDOW_JA = (
    "手術前60日以内に測定された値のうち、入院日に最も近い値"
)

PREOP_MED_WINDOW_EN = "Medication use documented within 60 days before surgery"
PREOP_MED_WINDOW_JA = "手術前60日以内に確認された薬剤使用"

PREPROCESSING_EN = {
    "Continuous": (
        "Winsorization at the 0.5th and 99.5th percentiles; log1p transformation "
        "when all observed values were nonnegative and absolute skewness exceeded 1; "
        "single iterative multivariable imputation using IterativeImputer with an "
        "LGBMRegressor; robust scaling for logistic regression only"
    ),
    "Binary": (
        "Entered as 0/1 without scaling; included as auxiliary predictors in the "
        "iterative imputation model; any imputed values were constrained to 0/1"
    ),
    "Ordinal": (
        "Entered as an ordinal numeric variable; single iterative multivariable "
        "imputation if required; constrained to the admissible integer range; "
        "robust scaling for logistic regression only"
    ),
    "Count": (
        "Entered as a nonnegative count; single iterative multivariable imputation "
        "if required; rounded to a nonnegative integer; robust scaling for "
        "logistic regression only"
    ),
}

PREPROCESSING_JA = {
    "Continuous": (
        "0.5および99.5パーセンタイルでWinsorizationを実施し、観測値がすべて非負かつ"
        "絶対歪度が1を超える場合にlog1p変換を実施。IterativeImputerとLGBMRegressorによる"
        "単一の多変量反復補完を行い、ロジスティック回帰でのみRobust scalingを実施"
    ),
    "Binary": (
        "0/1で投入し、スケーリングは実施しない。反復補完モデルの補助予測因子として使用し、"
        "欠測補完が生じた場合は0/1に制約"
    ),
    "Ordinal": (
        "順序尺度の数値変数として投入。必要時に多変量反復補完を行い、許容整数範囲に制約。"
        "ロジスティック回帰でのみRobust scalingを実施"
    ),
    "Count": (
        "非負のカウント変数として投入。必要時に多変量反復補完を行い、非負整数に丸める。"
        "ロジスティック回帰でのみRobust scalingを実施"
    ),
}

feature_specs: Dict[str, Dict[str, str]] = {}


def add_feature_spec(
    variable: str,
    name_en: str,
    name_ja: str,
    group_en: str,
    group_ja: str,
    definition_en: str,
    definition_ja: str,
    variable_type: str,
    unit_or_coding_en: str,
    unit_or_coding_ja: str,
    window_en: str,
    window_ja: str,
) -> None:
    feature_specs[variable] = {
        "Predictor": name_en,
        "予測因子": name_ja,
        "Feature group": group_en,
        "特徴量群": group_ja,
        "Definition": definition_en,
        "定義": definition_ja,
        "Type": variable_type,
        "Unit or coding": unit_or_coding_en,
        "単位または符号化": unit_or_coding_ja,
        "Measurement window": window_en,
        "測定時間窓": window_ja,
    }


# 患者背景
add_feature_spec(
    "Age", "Age", "年齢",
    "Preoperative patient characteristics", "術前患者背景",
    "Age on the date of the index surgery.", "対象手術日の年齢。",
    "Continuous", "years", "歳",
    "At the index surgery", "対象手術時",
)
add_feature_spec(
    "Male", "Male sex", "男性",
    "Preoperative patient characteristics", "術前患者背景",
    "Sex recorded as male in the electronic health record.",
    "電子診療録上の性別が男性。",
    "Binary", "0 = female; 1 = male", "0＝女性、1＝男性",
    "Demographic information available before surgery", "手術前に利用可能な基本情報",
)
add_feature_spec(
    "BMI", "Body mass index", "体格指数",
    "Preoperative patient characteristics", "術前患者背景",
    "Body weight divided by height squared.", "体重を身長の二乗で除した値。",
    "Continuous", "kg/m²", "kg/m²",
    "Most recent height and weight available before surgery",
    "手術前に利用可能な直近の身長および体重",
)
add_feature_spec(
    "ASA", "ASA Physical Status", "ASA身体状態分類",
    "Preoperative patient characteristics", "術前患者背景",
    "Preoperative physical status according to the American Society of "
    "Anesthesiologists Physical Status classification.",
    "米国麻酔科学会身体状態分類に基づく術前全身状態。",
    "Ordinal", "ASA-PS I–V", "ASA-PS I–V",
    "At the preoperative anesthesia assessment", "術前麻酔評価時",
)

# 併存疾患
for variable, name_en, name_ja, definition_en, definition_ja in [
    (
        "Dialysis", "Maintenance dialysis", "維持透析",
        "Receipt of maintenance dialysis before the index surgery.",
        "対象手術前の維持透析施行。",
    ),
    (
        "CHF", "Congestive heart failure", "うっ血性心不全",
        "Documented diagnosis of congestive heart failure before the index surgery.",
        "対象手術前に記録されたうっ血性心不全。",
    ),
    (
        "Malig", "Malignancy", "悪性腫瘍",
        "Documented diagnosis of malignancy before the index surgery.",
        "対象手術前に記録された悪性腫瘍。",
    ),
]:
    add_feature_spec(
        variable, name_en, name_ja,
        "Preoperative comorbidities", "術前併存疾患",
        definition_en, definition_ja,
        "Binary", "0 = no; 1 = yes", "0＝なし、1＝あり",
        "Status documented before the index surgery", "対象手術前に確認された状態",
    )

# 術前検査値
laboratory_specs = {
    "Alb": ("Serum albumin", "血清アルブミン", "g/dL"),
    "BUN": ("Blood urea nitrogen", "血中尿素窒素", "mg/dL"),
    "CRP": ("C-reactive protein", "C反応性蛋白", "mg/dL"),
    "Cre": ("Serum creatinine", "血清クレアチニン", "mg/dL"),
    "Hb": ("Hemoglobin", "ヘモグロビン", "g/dL"),
    "K": ("Serum potassium", "血清カリウム", "mEq/L"),
    "Na": ("Serum sodium", "血清ナトリウム", "mEq/L"),
    "PLT": ("Platelet count", "血小板数", "×10⁴/µL"),
    "T-Bil": ("Total bilirubin", "総ビリルビン", "mg/dL"),
    "WBC": ("White blood cell count", "白血球数", "/µL"),
}

for variable, (name_en, name_ja, unit) in laboratory_specs.items():
    add_feature_spec(
        variable, name_en, name_ja,
        "Preoperative laboratory data", "術前検査値",
        f"{name_en} measured before surgery.",
        f"手術前に測定された{name_ja}。",
        "Continuous", unit, unit,
        PREOP_LAB_WINDOW_EN, PREOP_LAB_WINDOW_JA,
    )

# 術前薬剤
medication_specs = {
    "DeliMed": (
        "Potentially deliriogenic medication",
        "せん妄関連薬",
        "Use of any medication included in the prespecified potentially "
        "deliriogenic medication list.",
        "事前規定したせん妄誘発可能性のある薬剤リストに含まれる薬剤の使用。",
    ),
    "β-blocker": (
        "Beta-blocker", "β遮断薬",
        "Preoperative use of a beta-blocker.", "術前のβ遮断薬使用。",
    ),
    "Oral steroids": (
        "Oral corticosteroid", "経口ステロイド",
        "Preoperative use of an oral corticosteroid.", "術前の経口副腎皮質ステロイド使用。",
    ),
    "Antiplatelet": (
        "Antiplatelet agent", "抗血小板薬",
        "Preoperative use of an antiplatelet agent.", "術前の抗血小板薬使用。",
    ),
    "Anticoag": (
        "Anticoagulant", "抗凝固薬",
        "Preoperative use of an anticoagulant.", "術前の抗凝固薬使用。",
    ),
    "AntiCa": (
        "Calcium-channel blocker", "カルシウム拮抗薬",
        "Preoperative use of a calcium-channel blocker.", "術前のカルシウム拮抗薬使用。",
    ),
    "Opioid": (
        "Opioid", "オピオイド",
        "Preoperative use of an opioid.", "術前のオピオイド使用。",
    ),
}

for variable, (name_en, name_ja, definition_en, definition_ja) in medication_specs.items():
    add_feature_spec(
        variable, name_en, name_ja,
        "Preoperative medications", "術前薬剤",
        definition_en, definition_ja,
        "Binary", "0 = no; 1 = yes", "0＝使用なし、1＝使用あり",
        PREOP_MED_WINDOW_EN, PREOP_MED_WINDOW_JA,
    )

# 手術分類
procedure_specs = {
    "Proc-Eye": ("Ophthalmic procedure", "眼科手術"),
    "Proc-Face/Neck": ("Face or neck procedure", "顔面・頸部手術"),
    "Proc-Thorax": ("Thoracic procedure", "胸部手術"),
    "Proc-MSK": ("Musculoskeletal procedure", "筋骨格系手術"),
    "Proc-ENT": ("Otorhinolaryngologic procedure", "耳鼻咽喉科手術"),
    "Proc-Neuro": ("Neurosurgical procedure", "脳神経外科手術"),
    "Proc-Genital": ("Genital procedure", "生殖器手術"),
    "Proc-Urinary": ("Urinary tract procedure", "尿路系手術"),
    "Proc-Skin": ("Skin or soft-tissue procedure", "皮膚・軟部組織手術"),
    "Proc-Abd": ("Abdominal procedure", "腹部手術"),
}

for variable, (name_en, name_ja) in procedure_specs.items():
    article = "an" if name_en[0].lower() in "aeiou" else "a"
    add_feature_spec(
        variable, name_en, name_ja,
        "Procedure characteristics", "手術特性",
        f"Indicator that the index surgery was classified as {article} {name_en.lower()}.",
        f"対象手術が{name_ja}に分類されたことを示す指標。",
        "Binary", "0 = no; 1 = yes", "0＝非該当、1＝該当",
        "Planned index procedure documented before surgery", "手術前に記録された予定手術",
    )

add_feature_spec(
    "ResectNum", "Number of resections", "切除数",
    "Procedure characteristics", "手術特性",
    "Planned number of resection procedures documented for the index surgery.",
    "対象手術について手術前に記録された予定切除手技数。",
    "Count", "count", "個",
    "Planned index procedure documented before surgery", "手術前に記録された予定手術",
)
add_feature_spec(
    "HighRiskProc", "High-risk procedure", "高リスク手術",
    "Procedure characteristics", "手術特性",
    "Indicator that the index surgery met the prespecified definition of a "
    "high-risk procedure.",
    "対象手術が事前規定した高リスク手術の定義を満たすことを示す指標。",
    "Binary", "0 = no; 1 = yes", "0＝非該当、1＝該当",
    "Planned index procedure documented before surgery", "手術前に記録された予定手術",
)

# 術中情報
intraoperative_specs = {
    "OpTime": (
        "Operative duration", "手術時間",
        "Time from the start to the end of the surgical procedure.",
        "手術開始から手術終了までの時間。",
        "min", "分",
    ),
    "RBC Tx": (
        "Red blood cell transfusion", "赤血球輸血",
        "Total amount of red blood cells transfused during the index surgery.",
        "対象手術中の赤血球輸血量。",
        "units", "単位",
    ),
    "FFP Tx": (
        "Fresh frozen plasma transfusion", "新鮮凍結血漿輸血",
        "Total amount of fresh frozen plasma transfused during the index surgery.",
        "対象手術中の新鮮凍結血漿輸血量。",
        "units", "単位",
    ),
    "PLT Tx": (
        "Platelet transfusion", "血小板輸血",
        "Total amount of platelets transfused during the index surgery.",
        "対象手術中の血小板輸血量。",
        "units", "単位",
    ),
    "FluidBal": (
        "Intraoperative fluid balance", "術中水分出納",
        "Net intraoperative fluid balance calculated from administered fluid "
        "and measured output.",
        "投与液量と測定排出量から算出した術中水分出納。",
        "mL", "mL",
    ),
}

for variable, (
    name_en,
    name_ja,
    definition_en,
    definition_ja,
    unit_en,
    unit_ja,
) in intraoperative_specs.items():
    add_feature_spec(
        variable, name_en, name_ja,
        "Intraoperative data", "術中情報",
        definition_en, definition_ja,
        "Continuous", unit_en, unit_ja,
        "From the start to the end of the index surgery",
        "対象手術の開始から終了まで",
    )

# 術後6時間
add_feature_spec(
    "HR at 6h", "Heart rate at 6 hours", "術後6時間の心拍数",
    "Early postoperative physiological data", "術後早期生理指標",
    "Heart rate recorded 6 hours after the end of surgery.",
    "手術終了6時間後に記録された心拍数。",
    "Continuous", "beats/min", "回/分",
    "6 hours after the end of surgery", "手術終了6時間後",
)
add_feature_spec(
    "MAP at 6h", "Mean arterial pressure at 6 hours", "術後6時間の平均動脈圧",
    "Early postoperative physiological data", "術後早期生理指標",
    "Mean arterial pressure recorded 6 hours after the end of surgery.",
    "手術終了6時間後に記録された平均動脈圧。",
    "Continuous", "mmHg", "mmHg",
    "6 hours after the end of surgery", "手術終了6時間後",
)

missing_specs = sorted(set(PERIOPERATIVE_FEATURES) - set(feature_specs))
if missing_specs:
    raise KeyError(f"メタデータが未定義の特徴量があります: {missing_specs}")


def build_feature_tables(data: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    rows_en: List[Dict[str, Any]] = []
    rows_ja: List[Dict[str, Any]] = []

    for number, variable in enumerate(PERIOPERATIVE_FEATURES, start=1):
        spec = feature_specs[variable]
        missing_n = int(data[variable].isna().sum())
        missing_pct = float(data[variable].isna().mean() * 100)
        variable_type = spec["Type"]

        rows_en.append(
            {
                "No.": number,
                "Source variable": variable,
                "Predictor": spec["Predictor"],
                "Feature group": spec["Feature group"],
                "Definition": spec["Definition"],
                "Type": variable_type,
                "Unit or coding": spec["Unit or coding"],
                "Measurement window": spec["Measurement window"],
                "Missing, n": missing_n,
                "Missing, %": round(missing_pct, 2),
                "Preoperative model": "Yes" if variable in PREOPERATIVE_FEATURES else "No",
                "Perioperative model": "Yes",
                "Preprocessing": PREPROCESSING_EN[variable_type],
            }
        )

        rows_ja.append(
            {
                "No.": number,
                "原変数名": variable,
                "予測因子": spec["予測因子"],
                "特徴量群": spec["特徴量群"],
                "定義": spec["定義"],
                "変数型": variable_type,
                "単位または符号化": spec["単位または符号化"],
                "測定時間窓": spec["測定時間窓"],
                "欠測数": missing_n,
                "欠測率（%）": round(missing_pct, 2),
                "術前モデル": "Yes" if variable in PREOPERATIVE_FEATURES else "No",
                "周術期モデル": "Yes",
                "前処理": PREPROCESSING_JA[variable_type],
            }
        )

    return pd.DataFrame(rows_en), pd.DataFrame(rows_ja)


feature_table_en, feature_table_ja = build_feature_tables(df_raw)
feature_table_en.head()

# %%
# Cell 7
# ============================================================
# 5. 最終特徴量辞書をExcel出力
# ============================================================

model_feature_rows: List[Dict[str, Any]] = []
for model_name, variables in [
    ("Preoperative model", PREOPERATIVE_FEATURES),
    ("Perioperative model", PERIOPERATIVE_FEATURES),
]:
    for order, variable in enumerate(variables, start=1):
        model_feature_rows.append(
            {
                "Model": model_name,
                "Order": order,
                "Source variable": variable,
                "Predictor": feature_specs[variable]["Predictor"],
            }
        )

model_feature_table = pd.DataFrame(model_feature_rows)

excluded_table = pd.DataFrame(
    [
        {
            "Column": column,
            "Reason for exclusion": reason,
        }
        for column, reason in EXCLUDED_COLUMNS.items()
        if column in df_raw.columns
    ]
)

with pd.ExcelWriter(FEATURE_TABLE_PATH, engine="openpyxl") as writer:
    feature_table_en.to_excel(
        writer,
        sheet_name="Table_S1_EN",
        index=False,
        startrow=3,
    )
    feature_table_ja.to_excel(
        writer,
        sheet_name="Table_S1_JA",
        index=False,
        startrow=3,
    )
    model_feature_table.to_excel(
        writer,
        sheet_name="Model_Feature_Lists",
        index=False,
    )
    excluded_table.to_excel(
        writer,
        sheet_name="Excluded_Columns",
        index=False,
    )

workbook = load_workbook(FEATURE_TABLE_PATH)

title_fill = PatternFill(fill_type="solid", fgColor="17365D")
header_fill = PatternFill(fill_type="solid", fgColor="2F75B5")
subheader_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
light_fill = PatternFill(fill_type="solid", fgColor="F7FAFC")

title_font = Font(bold=True, color="FFFFFF", size=13)
header_font = Font(bold=True, color="FFFFFF", size=10)

titles = {
    "Table_S1_EN": (
        "Supplementary Table S1. Definitions, measurement windows, missingness, "
        "and preprocessing of predictor variables included in the final models"
    ),
    "Table_S1_JA": (
        "Supplementary Table S1. 最終モデルに含めた予測因子の定義、測定時間窓、"
        "欠測状況および前処理"
    ),
}

subtitles = {
    "Table_S1_EN": (
        "The preoperative model included 36 predictors, and the perioperative "
        "model included all 43 predictors."
    ),
    "Table_S1_JA": (
        "術前モデルは36特徴量、周術期モデルは43特徴量を使用した。"
    ),
}

footnotes_en = [
    "Missingness was calculated in the full study cohort before imputation.",
    (
        "The primary analysis used single iterative multivariable imputation "
        "implemented with scikit-learn IterativeImputer. An LGBMRegressor "
        "(n_estimators=50, learning_rate=0.1, random_state=0) was used as the "
        "conditional estimator; mean imputation was used only for initialization, "
        "and the maximum number of iterations was five."
    ),
    (
        "Within every hold-out, IECV, calibration, and tuning split, the "
        "preprocessing parameters and imputer were fitted using the corresponding "
        "development/training data only and were then applied to the validation data."
    ),
    (
        "Patient identifiers, institutional identifiers, and outcome variables "
        "were not used as predictors or as inputs to the imputation model."
    ),
]

footnotes_ja = [
    "欠測数および欠測率は、欠測補完前の研究対象集団全体で算出した。",
    (
        "主解析ではscikit-learnのIterativeImputerによる単一の多変量反復補完を使用した。"
        "条件付き推定器にはLGBMRegressor（n_estimators=50、learning_rate=0.1、"
        "random_state=0）を用い、平均値補完は初期化のみに使用し、最大反復回数は5回とした。"
    ),
    (
        "hold-out、IECV、較正およびチューニングの各分割で、前処理パラメータおよび補完器は"
        "対応する開発・学習データのみで学習し、その後に検証データへ適用した。"
    ),
    (
        "患者識別子、施設識別子およびアウトカム変数は、予測因子および欠測補完モデルへの"
        "入力に使用しなかった。"
    ),
]

for sheet_name in ["Table_S1_EN", "Table_S1_JA"]:
    ws = workbook[sheet_name]
    max_col = ws.max_column
    max_col_letter = get_column_letter(max_col)

    ws.merge_cells(f"A1:{max_col_letter}1")
    ws["A1"] = titles[sheet_name]
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{max_col_letter}2")
    ws["A2"] = subtitles[sheet_name]
    ws["A2"].fill = subheader_fill
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    header_row = 4
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    data_start = 5
    data_end = data_start + len(feature_table_en) - 1

    for row in range(data_start, data_end + 1):
        if (row - data_start) % 2 == 1:
            for cell in ws[row]:
                cell.fill = light_fill
        for cell in ws[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{max_col_letter}{data_end}"
    ws.sheet_view.showGridLines = False

    header_values = {
        ws.cell(header_row, col).value: col
        for col in range(1, max_col + 1)
    }
    missing_pct_header = "Missing, %" if sheet_name == "Table_S1_EN" else "欠測率（%）"
    missing_pct_col = header_values[missing_pct_header]
    for row in range(data_start, data_end + 1):
        ws.cell(row, missing_pct_col).number_format = "0.00"

    footnotes = footnotes_en if sheet_name == "Table_S1_EN" else footnotes_ja
    footnote_start = data_end + 3
    for idx, text in enumerate(footnotes, start=1):
        row = footnote_start + idx - 1
        ws.merge_cells(f"A{row}:{max_col_letter}{row}")
        ws.cell(row, 1).value = f"{idx}. {text}"
        ws.cell(row, 1).font = Font(italic=True, size=9)
        ws.cell(row, 1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 32

    widths = {
        1: 6,
        2: 20,
        3: 26,
        4: 28,
        5: 48,
        6: 14,
        7: 22,
        8: 38,
        9: 12,
        10: 12,
        11: 17,
        12: 17,
        13: 60,
    }
    for col_idx, width in widths.items():
        if col_idx <= max_col:
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in range(data_start, data_end + 1):
        ws.row_dimensions[row].height = 54

for sheet_name in ["Model_Feature_Lists", "Excluded_Columns"]:
    ws = workbook[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 45)

workbook.save(FEATURE_TABLE_PATH)

print(f"Saved: {FEATURE_TABLE_PATH}")

# %% [markdown]
# Cell 8
# ## 6. リークを避けた前処理
# 
# - 連続変数のWinsorizationおよびlog1p変換を学習データ内で決定
# - `IterativeImputer` + `LGBMRegressor`による単一反復補完
# - 平均値は反復補完の初期値としてのみ使用
# - hold-out、IECV、チューニング、較正foldごとに、補完器を学習fold内のみでfit
# - ロジスティック回帰のみ、連続・順序・カウント変数をRobust scaling
# - 木構造モデルではスケーリングを行わない

# %%
# Cell 9
# ============================================================
# 6. リークを避けた前処理Transformer
# ============================================================

class LeakageFreePreprocessor(BaseEstimator, TransformerMixin):
    """
    continuous featuresのWinsorization/log1p、全特徴量を用いた反復補完、
    変数型制約、および選択的スケーリングを行う。
    """

    def __init__(
        self,
        feature_columns: Sequence[str],
        continuous_columns: Sequence[str],
        binary_columns: Sequence[str],
        ordinal_columns: Sequence[str],
        count_columns: Sequence[str],
        scale_for_logistic: bool = False,
        imputer_kind: str = "lgbm_single",
        random_state: int = 0,
        p_low: float = 0.005,
        p_high: float = 0.995,
        skew_threshold: float = 1.0,
        imputer_n_estimators: int = 50,
        imputer_learning_rate: float = 0.1,
        imputer_max_iter: int = 5,
        imputer_n_nearest_features: Optional[int] = None,
    ) -> None:
        self.feature_columns = feature_columns
        self.continuous_columns = continuous_columns
        self.binary_columns = binary_columns
        self.ordinal_columns = ordinal_columns
        self.count_columns = count_columns
        self.scale_for_logistic = scale_for_logistic
        self.imputer_kind = imputer_kind
        self.random_state = random_state
        self.p_low = p_low
        self.p_high = p_high
        self.skew_threshold = skew_threshold
        self.imputer_n_estimators = imputer_n_estimators
        self.imputer_learning_rate = imputer_learning_rate
        self.imputer_max_iter = imputer_max_iter
        self.imputer_n_nearest_features = imputer_n_nearest_features

    def _as_dataframe(self, X: Any) -> pd.DataFrame:
        columns = list(self.feature_columns)
        if isinstance(X, pd.DataFrame):
            missing = [column for column in columns if column not in X.columns]
            if missing:
                raise KeyError(f"前処理入力に不足列があります: {missing}")
            return X.loc[:, columns].copy()

        array = np.asarray(X)
        if array.ndim != 2 or array.shape[1] != len(columns):
            raise ValueError(
                f"入力形状が不正です: {array.shape}; expected columns={len(columns)}"
            )
        return pd.DataFrame(array, columns=columns)

    def _fit_winsor_log(self, X: pd.DataFrame) -> None:
        self.lower_bounds_: Dict[str, float] = {}
        self.upper_bounds_: Dict[str, float] = {}
        self.log_columns_: List[str] = []

        for column in self.continuous_columns_:
            series = pd.to_numeric(X[column], errors="coerce")
            lower = float(series.quantile(self.p_low))
            upper = float(series.quantile(self.p_high))

            if not np.isfinite(lower) or not np.isfinite(upper):
                raise ValueError(f"{column}の分位点を推定できません。")

            self.lower_bounds_[column] = lower
            self.upper_bounds_[column] = upper

            clipped = series.clip(lower, upper)
            observed = clipped.dropna()

            if (
                len(observed) > 2
                and (observed >= 0).all()
                and abs(float(observed.skew())) > self.skew_threshold
            ):
                self.log_columns_.append(column)

    def _apply_winsor_log(self, X: pd.DataFrame) -> pd.DataFrame:
        transformed = X.copy()

        for column in self.continuous_columns_:
            transformed[column] = pd.to_numeric(
                transformed[column], errors="coerce"
            ).clip(
                self.lower_bounds_[column],
                self.upper_bounds_[column],
            )

            if column in self.log_columns_:
                transformed[column] = np.log1p(transformed[column])

        for column in (
            *self.binary_columns_,
            *self.ordinal_columns_,
            *self.count_columns_,
        ):
            transformed[column] = pd.to_numeric(
                transformed[column], errors="coerce"
            )

        return transformed

    def _build_imputer(self) -> IterativeImputer:
        if self.imputer_kind == "lgbm_single":
            estimator = LGBMRegressor(
                n_estimators=self.imputer_n_estimators,
                learning_rate=self.imputer_learning_rate,
                n_jobs=1,
                random_state=self.random_state,
                verbosity=-1,
            )
            sample_posterior = False

        elif self.imputer_kind == "bayesian_multiple":
            estimator = BayesianRidge()
            sample_posterior = True

        else:
            raise ValueError(
                "imputer_kindは'lgbm_single'または'bayesian_multiple'を指定してください。"
            )

        return IterativeImputer(
            estimator=estimator,
            max_iter=self.imputer_max_iter,
            initial_strategy="mean",
            imputation_order="ascending",
            sample_posterior=sample_posterior,
            skip_complete=True,
            keep_empty_features=True,
            n_nearest_features=self.imputer_n_nearest_features,
            random_state=self.random_state,
        )

    def _constrain_types(self, X: pd.DataFrame) -> pd.DataFrame:
        constrained = X.copy()

        for column in self.binary_columns_:
            constrained[column] = constrained[column].clip(0, 1).round()

        for column in self.ordinal_columns_:
            constrained[column] = constrained[column].round()
            if column == "ASA":
                constrained[column] = constrained[column].clip(1, 5)

        for column in self.count_columns_:
            constrained[column] = constrained[column].clip(lower=0).round()

        return constrained

    def fit(self, X: Any, y: Optional[np.ndarray] = None) -> "LeakageFreePreprocessor":
        self.feature_columns_ = list(self.feature_columns)
        self.continuous_columns_ = [
            column for column in self.continuous_columns
            if column in self.feature_columns_
        ]
        self.binary_columns_ = [
            column for column in self.binary_columns
            if column in self.feature_columns_
        ]
        self.ordinal_columns_ = [
            column for column in self.ordinal_columns
            if column in self.feature_columns_
        ]
        self.count_columns_ = [
            column for column in self.count_columns
            if column in self.feature_columns_
        ]

        X_df = self._as_dataframe(X)
        self._fit_winsor_log(X_df)
        X_transformed = self._apply_winsor_log(X_df)

        self.imputer_ = self._build_imputer()
        imputed_array = self.imputer_.fit_transform(
            X_transformed.loc[:, self.feature_columns_]
        )

        X_imputed = pd.DataFrame(
            imputed_array,
            columns=self.feature_columns_,
            index=X_df.index,
        )
        X_imputed = self._constrain_types(X_imputed)

        self.scaled_columns_ = [
            column
            for column in (
                *self.continuous_columns_,
                *self.ordinal_columns_,
                *self.count_columns_,
            )
            if column in self.feature_columns_
        ]

        if self.scale_for_logistic:
            self.scaler_ = RobustScaler()
            self.scaler_.fit(X_imputed.loc[:, self.scaled_columns_])
        else:
            self.scaler_ = None

        return self

    def transform(self, X: Any) -> np.ndarray:
        X_df = self._as_dataframe(X)
        X_transformed = self._apply_winsor_log(X_df)

        imputed_array = self.imputer_.transform(
            X_transformed.loc[:, self.feature_columns_]
        )

        X_imputed = pd.DataFrame(
            imputed_array,
            columns=self.feature_columns_,
            index=X_df.index,
        )
        X_imputed = self._constrain_types(X_imputed)

        if self.scaler_ is not None:
            X_imputed.loc[:, self.scaled_columns_] = self.scaler_.transform(
                X_imputed.loc[:, self.scaled_columns_]
            )

        output = X_imputed.loc[:, self.feature_columns_].to_numpy(dtype=float)

        if not np.isfinite(output).all():
            raise ValueError(
                "前処理後のモデル入力に非有限値が残っています。"
            )

        return output

    def get_feature_names_out(
        self,
        input_features: Optional[Sequence[str]] = None,
    ) -> np.ndarray:
        return np.asarray(self.feature_columns_, dtype=object)

# %% [markdown]
# Cell 10
# ## 7. 評価指標

# %%
# Cell 11
# ============================================================
# 7. 評価指標・信頼区間・ペア比較
# ============================================================

METRIC_NAMES = (
    "AUROC",
    "AUPRC",
    "Brier",
    "Scaled_Brier",
    "Calibration_slope",
    "Calibration_intercept",
)


def calibration_slope_intercept(
    y_true: np.ndarray,
    y_prob: np.ndarray,
    max_iter: int = 50,
    tolerance: float = 1e-8,
) -> Tuple[float, float]:
    """
    logit(P[Y=1]) = intercept + slope * logit(predicted probability)
    を2変数Newton法で推定する。Bootstrap内で多数回呼ぶため、
    sklearnモデルfitより軽量な実装とする。
    """
    y = np.asarray(y_true, dtype=float)
    p = np.asarray(y_prob, dtype=float)

    if len(y) == 0 or np.unique(y).size < 2:
        return np.nan, np.nan

    eps = 1e-12
    p = np.clip(p, eps, 1 - eps)
    z = np.log(p / (1 - p))
    X = np.column_stack([np.ones(len(z)), z])

    prevalence = np.clip(y.mean(), eps, 1 - eps)
    beta = np.array(
        [
            math.log(prevalence / (1 - prevalence)),
            1.0,
        ],
        dtype=float,
    )

    for _ in range(max_iter):
        eta = X @ beta
        mu = expit(eta)
        weights = np.clip(mu * (1 - mu), 1e-9, None)

        information = X.T @ (weights[:, None] * X)
        information += np.eye(2) * 1e-9
        score = X.T @ (y - mu)

        try:
            step = np.linalg.solve(information, score)
        except np.linalg.LinAlgError:
            return np.nan, np.nan

        beta_new = beta + step

        if np.max(np.abs(beta_new - beta)) < tolerance:
            beta = beta_new
            break

        beta = beta_new

    return float(beta[1]), float(beta[0])


def compute_metrics(
    y_true: np.ndarray,
    y_prob: np.ndarray,
) -> Dict[str, float]:
    y = np.asarray(y_true, dtype=int)
    p = np.asarray(y_prob, dtype=float)

    if len(y) != len(p):
        raise ValueError("y_trueとy_probの長さが一致しません。")

    if len(y) == 0:
        return {
            "N": 0,
            "Events": 0,
            "Prevalence": np.nan,
            **{metric: np.nan for metric in METRIC_NAMES},
        }

    p = np.clip(p, 0.0, 1.0)
    n = len(y)
    events = int(y.sum())
    prevalence = float(y.mean())

    if np.unique(y).size < 2:
        auroc = np.nan
        auprc = np.nan
    else:
        auroc = float(roc_auc_score(y, p))
        auprc = float(average_precision_score(y, p))

    brier = float(brier_score_loss(y, p))
    null_brier = prevalence * (1 - prevalence)
    scaled_brier = (
        float(1 - brier / null_brier)
        if null_brier > 0
        else np.nan
    )

    slope, intercept = calibration_slope_intercept(y, p)

    return {
        "N": n,
        "Events": events,
        "Prevalence": prevalence,
        "AUROC": auroc,
        "AUPRC": auprc,
        "Brier": brier,
        "Scaled_Brier": scaled_brier,
        "Calibration_slope": slope,
        "Calibration_intercept": intercept,
    }


def stratified_bootstrap_indices(
    y_true: np.ndarray,
    rng: np.random.Generator,
) -> np.ndarray:
    y = np.asarray(y_true, dtype=int)
    positive = np.flatnonzero(y == 1)
    negative = np.flatnonzero(y == 0)

    if len(positive) == 0 or len(negative) == 0:
        raise ValueError(
            "両クラスが存在しないため層化ブートストラップを実行できません。"
        )

    sampled_positive = rng.choice(
        positive,
        size=len(positive),
        replace=True,
    )
    sampled_negative = rng.choice(
        negative,
        size=len(negative),
        replace=True,
    )
    sampled = np.concatenate(
        [sampled_positive, sampled_negative]
    )
    rng.shuffle(sampled)
    return sampled


def bootstrap_metric_summary(
    y_true: np.ndarray,
    y_prob: np.ndarray,
    n_bootstrap: int,
    seed: int = RANDOM_STATE,
    alpha: float = 0.05,
) -> Dict[str, float]:
    point = compute_metrics(y_true, y_prob)
    result: Dict[str, float] = {
        "N": point["N"],
        "Events": point["Events"],
        "Prevalence": point["Prevalence"],
    }

    if np.unique(y_true).size < 2:
        for metric in METRIC_NAMES:
            result[metric] = point[metric]
            result[f"{metric}_lo"] = np.nan
            result[f"{metric}_hi"] = np.nan
        return result

    rng = np.random.default_rng(seed)
    bootstrap_values: Dict[str, List[float]] = {
        metric: [] for metric in METRIC_NAMES
    }

    y = np.asarray(y_true, dtype=int)
    p = np.asarray(y_prob, dtype=float)

    for _ in range(n_bootstrap):
        sample_index = stratified_bootstrap_indices(y, rng)
        sample_metrics = compute_metrics(
            y[sample_index],
            p[sample_index],
        )
        for metric in METRIC_NAMES:
            bootstrap_values[metric].append(
                sample_metrics[metric]
            )

    lower_q = 100 * alpha / 2
    upper_q = 100 * (1 - alpha / 2)

    for metric in METRIC_NAMES:
        values = np.asarray(
            bootstrap_values[metric],
            dtype=float,
        )
        valid = values[np.isfinite(values)]

        result[metric] = point[metric]

        if len(valid) == 0:
            result[f"{metric}_lo"] = np.nan
            result[f"{metric}_hi"] = np.nan
        else:
            result[f"{metric}_lo"] = float(
                np.percentile(valid, lower_q)
            )
            result[f"{metric}_hi"] = float(
                np.percentile(valid, upper_q)
            )

    return result


def compute_midrank(x: np.ndarray) -> np.ndarray:
    x = np.asarray(x, dtype=float)
    order = np.argsort(x)
    sorted_x = x[order]
    n = len(x)
    midranks = np.zeros(n, dtype=float)

    i = 0
    while i < n:
        j = i
        while j < n and sorted_x[j] == sorted_x[i]:
            j += 1
        midranks[i:j] = 0.5 * (i + j - 1)
        i = j

    restored = np.empty(n, dtype=float)
    restored[order] = midranks + 1
    return restored


def fast_delong(
    predictions_sorted_transposed: np.ndarray,
    positive_count: int,
) -> Tuple[np.ndarray, np.ndarray]:
    m = positive_count
    n = predictions_sorted_transposed.shape[1] - m
    k = predictions_sorted_transposed.shape[0]

    positive_examples = predictions_sorted_transposed[:, :m]
    negative_examples = predictions_sorted_transposed[:, m:]

    tx = np.empty((k, m), dtype=float)
    ty = np.empty((k, n), dtype=float)
    tz = np.empty((k, m + n), dtype=float)

    for classifier_index in range(k):
        tx[classifier_index] = compute_midrank(
            positive_examples[classifier_index]
        )
        ty[classifier_index] = compute_midrank(
            negative_examples[classifier_index]
        )
        tz[classifier_index] = compute_midrank(
            predictions_sorted_transposed[classifier_index]
        )

    aucs = (
        tz[:, :m].sum(axis=1) / (m * n)
        - (m + 1) / (2 * n)
    )
    v01 = (tz[:, :m] - tx) / n
    v10 = 1 - (tz[:, m:] - ty) / m

    sx = np.cov(v01)
    sy = np.cov(v10)
    covariance = sx / m + sy / n
    return aucs, covariance


def paired_delong_test(
    y_true: np.ndarray,
    prediction_1: np.ndarray,
    prediction_2: np.ndarray,
) -> Dict[str, float]:
    y = np.asarray(y_true, dtype=int)
    p1 = np.asarray(prediction_1, dtype=float)
    p2 = np.asarray(prediction_2, dtype=float)

    if np.unique(y).size < 2:
        return {
            "Estimate_1": np.nan,
            "Estimate_2": np.nan,
            "Difference_2_minus_1": np.nan,
            "Difference_lo": np.nan,
            "Difference_hi": np.nan,
            "P_value": np.nan,
        }

    order = np.argsort(-y)
    positive_count = int(y.sum())
    predictions = np.vstack([p1, p2])[:, order]

    aucs, covariance = fast_delong(
        predictions,
        positive_count,
    )

    contrast = np.array([-1.0, 1.0])
    difference = float(aucs[1] - aucs[0])
    variance = float(
        contrast @ covariance @ contrast.T
    )
    variance = max(variance, 0.0)
    standard_error = math.sqrt(variance)

    if standard_error > 0:
        z_value = difference / standard_error
        p_value = float(
            2 * (1 - norm.cdf(abs(z_value)))
        )
        lower = difference - 1.96 * standard_error
        upper = difference + 1.96 * standard_error
    else:
        p_value = 1.0 if difference == 0 else 0.0
        lower = difference
        upper = difference

    return {
        "Estimate_1": float(aucs[0]),
        "Estimate_2": float(aucs[1]),
        "Difference_2_minus_1": difference,
        "Difference_lo": lower,
        "Difference_hi": upper,
        "P_value": p_value,
    }


def paired_bootstrap_difference(
    y_true: np.ndarray,
    prediction_1: np.ndarray,
    prediction_2: np.ndarray,
    metric: str,
    n_bootstrap: int,
    seed: int = RANDOM_STATE,
    alpha: float = 0.05,
) -> Dict[str, float]:
    y = np.asarray(y_true, dtype=int)
    p1 = np.asarray(prediction_1, dtype=float)
    p2 = np.asarray(prediction_2, dtype=float)

    point_1 = compute_metrics(y, p1)[metric]
    point_2 = compute_metrics(y, p2)[metric]
    point_difference = point_2 - point_1

    if np.unique(y).size < 2:
        return {
            "Estimate_1": point_1,
            "Estimate_2": point_2,
            "Difference_2_minus_1": point_difference,
            "Difference_lo": np.nan,
            "Difference_hi": np.nan,
            "P_value": np.nan,
        }

    rng = np.random.default_rng(seed)
    differences: List[float] = []

    for _ in range(n_bootstrap):
        sample_index = stratified_bootstrap_indices(
            y,
            rng,
        )
        metric_1 = compute_metrics(
            y[sample_index],
            p1[sample_index],
        )[metric]
        metric_2 = compute_metrics(
            y[sample_index],
            p2[sample_index],
        )[metric]

        if np.isfinite(metric_1) and np.isfinite(metric_2):
            differences.append(metric_2 - metric_1)

    difference_array = np.asarray(
        differences,
        dtype=float,
    )

    if len(difference_array) == 0:
        lower = np.nan
        upper = np.nan
        p_value = np.nan
    else:
        lower = float(
            np.percentile(
                difference_array,
                100 * alpha / 2,
            )
        )
        upper = float(
            np.percentile(
                difference_array,
                100 * (1 - alpha / 2),
            )
        )
        denominator = len(difference_array) + 1
        lower_tail = (
            int(np.sum(difference_array <= 0)) + 1
        ) / denominator
        upper_tail = (
            int(np.sum(difference_array >= 0)) + 1
        ) / denominator
        p_value = min(
            float(2 * min(lower_tail, upper_tail)),
            1.0,
        )

    return {
        "Estimate_1": point_1,
        "Estimate_2": point_2,
        "Difference_2_minus_1": point_difference,
        "Difference_lo": lower,
        "Difference_hi": upper,
        "P_value": p_value,
    }


def compare_preoperative_vs_perioperative(
    y_true: np.ndarray,
    preoperative_prediction: np.ndarray,
    perioperative_prediction: np.ndarray,
    n_bootstrap: int,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    rows.append(
        {
            "Metric": "AUROC",
            "Method": "Paired DeLong",
            **paired_delong_test(
                y_true,
                preoperative_prediction,
                perioperative_prediction,
            ),
        }
    )

    for metric in (
        "AUPRC",
        "Brier",
        "Scaled_Brier",
    ):
        rows.append(
            {
                "Metric": metric,
                "Method": "Paired stratified bootstrap",
                **paired_bootstrap_difference(
                    y_true=y_true,
                    prediction_1=preoperative_prediction,
                    prediction_2=perioperative_prediction,
                    metric=metric,
                    n_bootstrap=n_bootstrap,
                ),
            }
        )

    return pd.DataFrame(rows)

# %% [markdown]
# Cell 12
# ## 8. 共通ユーティリティ

# %%
# Cell 13
# ============================================================
# 8. 共通ユーティリティ・Excel書式・PNG制約
# ============================================================

def sanitize_filename(value: Any) -> str:
    text = str(value)
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("_") or "value"



def stable_seed(*parts: Any, base_seed: int = RANDOM_STATE) -> int:
    text = "||".join(str(part) for part in parts)
    digest = hashlib.sha256(text.encode("utf-8")).digest()
    offset = int.from_bytes(digest[:4], byteorder="little", signed=False)
    return int((base_seed + offset) % (2**31 - 1))


def installed_version(distribution_name: str) -> str:
    try:
        return importlib.metadata.version(distribution_name)
    except importlib.metadata.PackageNotFoundError:
        return "not installed"


PREPROCESSING_CONFIG_PAYLOAD = {
    "single_imputer_max_iter": SINGLE_IMPUTER_MAX_ITER,
    "single_imputer_n_estimators": SINGLE_IMPUTER_N_ESTIMATORS,
    "single_imputer_learning_rate": SINGLE_IMPUTER_LEARNING_RATE,
    "mi_imputer_max_iter": MI_IMPUTER_MAX_ITER,
    "mi_n_nearest_features": MI_N_NEAREST_FEATURES,
    "tuning_splits": N_SPLITS_TUNING,
    "calibration_splits": N_SPLITS_CALIBRATION,
    "mi_calibration_splits": N_SPLITS_MI_CALIBRATION,
    "features": list(PERIOPERATIVE_FEATURES),
}
PREPROCESSING_SIGNATURE = hashlib.sha256(
    json.dumps(
        PREPROCESSING_CONFIG_PAYLOAD,
        ensure_ascii=False,
        sort_keys=True,
    ).encode("utf-8")
).hexdigest()[:16]


def json_serializable(value: Any) -> Any:
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, Path):
        return str(value)
    raise TypeError(
        f"JSONに変換できない型です: {type(value)}"
    )


def save_json(
    payload: Mapping[str, Any],
    path: Path,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(
            dict(payload),
            file,
            ensure_ascii=False,
            indent=2,
            default=json_serializable,
        )


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def dump_object(
    payload: Any,
    path: Path,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("wb") as file:
        cloudpickle.dump(
            payload,
            file,
            protocol=5,
        )


def load_object(path: Path) -> Any:
    with path.open("rb") as file:
        return cloudpickle.load(file)


def style_excel_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="2F75B5",
    )
    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for worksheet in workbook.worksheets:
        if worksheet.max_row < 1:
            continue

        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False

        if worksheet.max_column > 0:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in worksheet.columns:
            letter = get_column_letter(
                column_cells[0].column
            )
            max_length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[letter].width = min(
                max(max_length + 2, 11),
                45,
            )

    workbook.save(path)


def save_pil_png_compliant(
    image: Image.Image,
    output_path: Path,
    *,
    dpi: int = PNG_EXPORT_DPI,
    max_width: int = PNG_MAX_WIDTH,
    max_height: int = PNG_MAX_HEIGHT,
    max_file_size_mb: float = PNG_MAX_FILE_SIZE_MB,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    max_bytes = int(max_file_size_mb * 1024 * 1024)

    img = image.convert("RGB").copy()
    img.thumbnail(
        (max_width, max_height),
        Image.Resampling.LANCZOS,
    )

    while True:
        img.save(
            output_path,
            format="PNG",
            optimize=True,
            compress_level=9,
            dpi=(dpi, dpi),
        )

        file_size = output_path.stat().st_size
        if (
            img.width <= max_width
            and img.height <= max_height
            and file_size <= max_bytes
        ):
            break

        new_width = max(
            300,
            int(round(img.width * 0.90)),
        )
        new_height = max(
            300,
            int(round(img.height * 0.90)),
        )

        if (
            new_width == img.width
            and new_height == img.height
        ):
            raise RuntimeError(
                f"PNGを投稿制約内に縮小できません: {output_path}"
            )

        img = img.resize(
            (new_width, new_height),
            Image.Resampling.LANCZOS,
        )

    return output_path


def save_figure_png_compliant(
    fig: plt.Figure,
    output_path: Path,
    *,
    dpi: int = PNG_EXPORT_DPI,
    bbox_inches: str = "tight",
    pad_inches: float = 0.05,
) -> Path:
    temporary_path = output_path.with_name(
        f"{output_path.stem}__temporary.png"
    )

    fig.savefig(
        temporary_path,
        dpi=dpi,
        bbox_inches=bbox_inches,
        pad_inches=pad_inches,
        facecolor="white",
    )

    with Image.open(temporary_path) as image:
        image.load()
        copied_image = image.copy()

    temporary_path.unlink(missing_ok=True)

    return save_pil_png_compliant(
        copied_image,
        output_path,
        dpi=dpi,
    )


def validate_png_file(
    png_path: Path,
) -> Dict[str, Any]:
    with Image.open(png_path) as image:
        width, height = image.size
        file_format = image.format

    file_size_mb = (
        png_path.stat().st_size / (1024 ** 2)
    )

    result = {
        "File": png_path.name,
        "Path": str(png_path),
        "Format": file_format,
        "Width_px": width,
        "Height_px": height,
        "File_size_MB": round(file_size_mb, 3),
        "Extension_OK": png_path.suffix.lower() == ".png",
        "Format_OK": file_format == "PNG",
        "Width_OK": width <= PNG_MAX_WIDTH,
        "Height_OK": height <= PNG_MAX_HEIGHT,
        "File_size_OK": (
            file_size_mb <= PNG_MAX_FILE_SIZE_MB
        ),
    }
    result["All_requirements_met"] = all(
        result[key]
        for key in (
            "Extension_OK",
            "Format_OK",
            "Width_OK",
            "Height_OK",
            "File_size_OK",
        )
    )
    return result


def validate_publication_pngs() -> pd.DataFrame:
    png_paths = sorted(
        PUBLICATION_FIGURE_DIR.glob("*.png")
    )
    if not png_paths:
        return pd.DataFrame()

    report = pd.DataFrame(
        [validate_png_file(path) for path in png_paths]
    )
    report_path = (
        PUBLICATION_FIGURE_DIR
        / "PNG_Submission_Compliance_Report.xlsx"
    )
    report.to_excel(report_path, index=False)
    style_excel_workbook(report_path)

    violations = report.loc[
        ~report["All_requirements_met"]
    ]
    if not violations.empty:
        raise RuntimeError(
            "投稿PNG制約を満たしていないファイルがあります:\n"
            + violations[
                [
                    "File",
                    "Width_px",
                    "Height_px",
                    "File_size_MB",
                ]
            ].to_string(index=False)
        )

    return report


def write_run_manifest() -> Path:
    manifest = {
        "analysis_version": ANALYSIS_VERSION,
        "raw_data_path": str(RAW_DATA_PATH),
        "raw_data_sheet": RAW_DATA_SHEET,
        "source_file_sha256": SOURCE_FILE_SHA256,
        "source_signature": SOURCE_SIGNATURE,
        "preprocessing_signature": PREPROCESSING_SIGNATURE,
        "output_directory": str(OUTPUT_DIR),
        "n_rows": int(len(df_raw)),
        "n_columns": int(len(df_raw.columns)),
        "facilities": sorted(df_raw[FACILITY_COL].astype(str).unique().tolist()),
        "facility_label_map": dict(FACILITY_LABEL_MAP),
        "primary_events": int(df_raw[PRIMARY_OUTCOME_COL].sum()),
        "primary_prevalence": float(df_raw[PRIMARY_OUTCOME_COL].mean()),
        "primary_models": list(PRIMARY_MODELS),
        "sensitivity_models": list(SENSITIVITY_MODELS),
        "variants": list(VARIANTS),
        "n_trials_iecv_by_model": dict(N_TRIALS_IECV_BY_MODEL),
        "n_trials_holdout_by_model": dict(N_TRIALS_HOLDOUT_BY_MODEL),
        "n_trials_temporal_by_model": dict(N_TRIALS_TEMPORAL_BY_MODEL),
        "run_primary_temporal": RUN_PRIMARY_TEMPORAL,
        "temporal_period_column": TEMPORAL_PERIOD_COL,
        "temporal_development_value": TEMPORAL_DEVELOPMENT_VALUE,
        "temporal_validation_value": TEMPORAL_VALIDATION_VALUE,
        "temporal_development_label": TEMPORAL_DEVELOPMENT_LABEL,
        "temporal_validation_label": TEMPORAL_VALIDATION_LABEL,
        "n_splits_tuning": N_SPLITS_TUNING,
        "n_splits_calibration": N_SPLITS_CALIBRATION,
        "n_bootstrap_primary": N_BOOTSTRAP_PRIMARY,
        "n_bootstrap_sensitivity": N_BOOTSTRAP_SENSITIVITY,
        "n_imputations": N_IMPUTATIONS,
        "shap_models": list(SHAP_MODELS),
        "shap_variants": list(SHAP_VARIANTS),
        "shap_max_per_facility": SHAP_MAX_PER_FACILITY,
        "random_state": RANDOM_STATE,
        "imputer_random_state": IMPUTER_RANDOM_STATE,
        "reuse_existing_outputs": REUSE_EXISTING_OUTPUTS,
        "use_warm_start": USE_WARM_START,
        "python": sys.version,
        "platform": platform.platform(),
        "package_versions": {
            "numpy": installed_version("numpy"),
            "pandas": installed_version("pandas"),
            "scikit-learn": installed_version("scikit-learn"),
            "scipy": installed_version("scipy"),
            "xgboost": installed_version("xgboost"),
            "lightgbm": installed_version("lightgbm"),
            "optuna": installed_version("optuna"),
            "shap": installed_version("shap"),
            "matplotlib": installed_version("matplotlib"),
            "openpyxl": installed_version("openpyxl"),
        },
    }

    path = LOG_DIR / "analysis_manifest.json"
    save_json(manifest, path)
    return path

# %% [markdown]
# Cell 14
# ## 9. 前処理キャッシュ

# %%
# Cell 15
# ============================================================
# 9. 前処理キャッシュ
# ============================================================

def variant_columns(
    variant: str,
) -> Tuple[
    Tuple[str, ...],
    Tuple[str, ...],
    Tuple[str, ...],
    Tuple[str, ...],
    Tuple[str, ...],
]:
    if variant not in VARIANT_FEATURES:
        raise KeyError(f"未知のvariantです: {variant}")

    features = tuple(VARIANT_FEATURES[variant])
    continuous = tuple(
        column
        for column in ALL_CONTINUOUS_FEATURES
        if column in features
    )
    binary = tuple(
        column
        for column in ALL_BINARY_FEATURES
        if column in features
    )
    ordinal = tuple(
        column
        for column in ALL_ORDINAL_FEATURES
        if column in features
    )
    count = tuple(
        column
        for column in ALL_COUNT_FEATURES
        if column in features
    )
    return features, continuous, binary, ordinal, count


def safe_cv_splits(
    y: np.ndarray,
    desired_splits: int,
) -> int:
    y_array = np.asarray(y, dtype=int)
    counts = np.bincount(y_array, minlength=2)
    minimum_class_count = int(counts.min())

    if minimum_class_count < 2:
        raise ValueError(
            f"少数クラスが{minimum_class_count}例のためCVを実行できません。"
        )

    return max(
        2,
        min(desired_splits, minimum_class_count),
    )


def scaled_feature_indices(
    variant: str,
) -> np.ndarray:
    features, continuous, _, ordinal, count = variant_columns(
        variant
    )
    scale_columns = set(
        (*continuous, *ordinal, *count)
    )
    return np.asarray(
        [
            index
            for index, feature in enumerate(features)
            if feature in scale_columns
        ],
        dtype=int,
    )


def fit_logistic_scaler(
    X_train_base: np.ndarray,
    X_validation_base: np.ndarray,
    variant: str,
) -> Tuple[np.ndarray, np.ndarray, RobustScaler]:
    indices = scaled_feature_indices(variant)
    scaler = RobustScaler()

    X_train = np.asarray(
        X_train_base,
        dtype=np.float32,
    ).copy()
    X_validation = np.asarray(
        X_validation_base,
        dtype=np.float32,
    ).copy()

    scaler.fit(X_train[:, indices])
    X_train[:, indices] = scaler.transform(
        X_train[:, indices]
    ).astype(np.float32)
    X_validation[:, indices] = scaler.transform(
        X_validation[:, indices]
    ).astype(np.float32)

    return X_train, X_validation, scaler


def build_base_preprocessor(
    variant: str,
    *,
    imputer_kind: str,
    imputer_seed: int,
    imputer_max_iter: int,
    imputer_n_nearest_features: Optional[int] = None,
) -> LeakageFreePreprocessor:
    features, continuous, binary, ordinal, count = variant_columns(
        variant
    )

    return LeakageFreePreprocessor(
        feature_columns=features,
        continuous_columns=continuous,
        binary_columns=binary,
        ordinal_columns=ordinal,
        count_columns=count,
        scale_for_logistic=False,
        imputer_kind=imputer_kind,
        random_state=imputer_seed,
        imputer_n_estimators=SINGLE_IMPUTER_N_ESTIMATORS,
        imputer_learning_rate=SINGLE_IMPUTER_LEARNING_RATE,
        imputer_max_iter=imputer_max_iter,
        imputer_n_nearest_features=imputer_n_nearest_features,
    )


def cache_file(
    *parts: Any,
    extension: str = ".joblib",
) -> Path:
    safe_parts = [
        sanitize_filename(part)
        for part in parts
    ]
    return (
        CACHE_RUN_DIR
        / PREPROCESSING_SIGNATURE
    ).joinpath(*safe_parts).with_suffix(extension)


def build_or_load_cv_preprocessing(
    data: pd.DataFrame,
    outer_train_indices: np.ndarray,
    outcome_column: str,
    variant: str,
    *,
    outer_key: str,
    purpose: str,
    n_splits: int,
    seed: int,
    imputer_kind: str = "lgbm_single",
    imputation_index: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    各CV foldの前処理を1回だけ実行し、全Optuna trial・全モデルで共有する。
    Logistic用のRobust scaling済み配列と、tree model用の未scaled配列を保存する。
    """
    cache_path = cache_file(
        "cv",
        outer_key,
        outcome_column,
        variant,
        purpose,
        f"splits_{n_splits}",
        f"seed_{seed}",
        imputer_kind,
        f"imputation_{imputation_index}"
        if imputation_index is not None
        else "single",
    )

    if (
        cache_path.exists()
        and not FORCE_REBUILD_PREPROCESSING
    ):
        return joblib.load(cache_path)

    features = list(VARIANT_FEATURES[variant])
    X_outer = data.iloc[
        outer_train_indices
    ][features].reset_index(drop=True)
    y_outer = data.iloc[
        outer_train_indices
    ][outcome_column].to_numpy(dtype=int)

    actual_splits = safe_cv_splits(
        y_outer,
        n_splits,
    )
    cv = StratifiedKFold(
        n_splits=actual_splits,
        shuffle=True,
        random_state=seed,
    )

    fold_records: List[Dict[str, Any]] = []

    for fold_number, (
        train_position,
        validation_position,
    ) in enumerate(
        cv.split(X_outer, y_outer),
        start=1,
    ):
        if imputer_kind == "bayesian_multiple":
            imputer_seed = (
                IMPUTER_RANDOM_STATE
                + 10000 * int(imputation_index or 0)
                + 100 * fold_number
                + seed
            )
            nearest_features = MI_N_NEAREST_FEATURES
            max_iter = MI_IMPUTER_MAX_ITER
        else:
            imputer_seed = (
                IMPUTER_RANDOM_STATE
                + fold_number
                + seed
            )
            nearest_features = None
            max_iter = SINGLE_IMPUTER_MAX_ITER

        preprocessor = build_base_preprocessor(
            variant=variant,
            imputer_kind=imputer_kind,
            imputer_seed=imputer_seed,
            imputer_max_iter=max_iter,
            imputer_n_nearest_features=nearest_features,
        )

        X_train_base = preprocessor.fit_transform(
            X_outer.iloc[train_position]
        ).astype(np.float32)
        X_validation_base = preprocessor.transform(
            X_outer.iloc[validation_position]
        ).astype(np.float32)

        (
            X_train_logistic,
            X_validation_logistic,
            _,
        ) = fit_logistic_scaler(
            X_train_base,
            X_validation_base,
            variant,
        )

        fold_records.append(
            {
                "fold": fold_number,
                "train_position": train_position.astype(int),
                "validation_position": validation_position.astype(int),
                "X_train_tree": X_train_base,
                "X_validation_tree": X_validation_base,
                "X_train_logistic": X_train_logistic,
                "X_validation_logistic": X_validation_logistic,
                "y_train": y_outer[train_position].astype(np.int8),
                "y_validation": y_outer[
                    validation_position
                ].astype(np.int8),
            }
        )

    cache_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    joblib.dump(
        fold_records,
        cache_path,
        compress=3,
    )
    return fold_records


def build_or_load_final_preprocessing(
    data: pd.DataFrame,
    train_indices: np.ndarray,
    validation_indices: np.ndarray,
    variant: str,
    *,
    outer_key: str,
    imputer_kind: str = "lgbm_single",
    imputation_index: Optional[int] = None,
) -> Dict[str, Any]:
    """
    outer development全体で前処理をfitし、outer validationへtransformする。
    アウトカムに依存しないため、主解析・感度分析で共有する。
    """
    cache_path = cache_file(
        "final",
        outer_key,
        variant,
        imputer_kind,
        f"imputation_{imputation_index}"
        if imputation_index is not None
        else "single",
        extension=".pkl",
    )

    if (
        cache_path.exists()
        and not FORCE_REBUILD_PREPROCESSING
    ):
        return load_object(cache_path)

    features = list(VARIANT_FEATURES[variant])
    X_train_raw = data.iloc[
        train_indices
    ][features]
    X_validation_raw = data.iloc[
        validation_indices
    ][features]

    if imputer_kind == "bayesian_multiple":
        imputer_seed = (
            IMPUTER_RANDOM_STATE
            + 10000 * int(imputation_index or 0)
            + RANDOM_STATE
        )
        nearest_features = MI_N_NEAREST_FEATURES
        max_iter = MI_IMPUTER_MAX_ITER
    else:
        imputer_seed = IMPUTER_RANDOM_STATE
        nearest_features = None
        max_iter = SINGLE_IMPUTER_MAX_ITER

    preprocessor = build_base_preprocessor(
        variant=variant,
        imputer_kind=imputer_kind,
        imputer_seed=imputer_seed,
        imputer_max_iter=max_iter,
        imputer_n_nearest_features=nearest_features,
    )

    X_train_tree = preprocessor.fit_transform(
        X_train_raw
    ).astype(np.float32)
    X_validation_tree = preprocessor.transform(
        X_validation_raw
    ).astype(np.float32)

    (
        X_train_logistic,
        X_validation_logistic,
        logistic_scaler,
    ) = fit_logistic_scaler(
        X_train_tree,
        X_validation_tree,
        variant,
    )

    payload = {
        "preprocessor": preprocessor,
        "logistic_scaler": logistic_scaler,
        "X_train_tree": X_train_tree,
        "X_validation_tree": X_validation_tree,
        "X_train_logistic": X_train_logistic,
        "X_validation_logistic": X_validation_logistic,
        "feature_names": features,
        "train_indices": np.asarray(
            train_indices,
            dtype=int,
        ),
        "validation_indices": np.asarray(
            validation_indices,
            dtype=int,
        ),
    }

    dump_object(
        payload,
        cache_path,
    )
    return payload


def model_arrays_from_fold(
    fold: Mapping[str, Any],
    model_name: str,
) -> Tuple[np.ndarray, np.ndarray]:
    if model_name == "Logistic":
        return (
            fold["X_train_logistic"],
            fold["X_validation_logistic"],
        )

    return (
        fold["X_train_tree"],
        fold["X_validation_tree"],
    )


def model_arrays_from_final(
    final_data: Mapping[str, Any],
    model_name: str,
) -> Tuple[np.ndarray, np.ndarray]:
    if model_name == "Logistic":
        return (
            final_data["X_train_logistic"],
            final_data["X_validation_logistic"],
        )

    return (
        final_data["X_train_tree"],
        final_data["X_validation_tree"],
    )

# %% [markdown]
# Cell 16
# ## 10. モデル構築・チューニング・較正

# %%
# Cell 17
# ============================================================
# 10. モデル構築・効率化Optuna・較正
# ============================================================

def positive_class_weight(
    y: np.ndarray,
) -> float:
    y_array = np.asarray(y, dtype=int)
    positive_count = int((y_array == 1).sum())
    negative_count = int((y_array == 0).sum())
    return negative_count / max(positive_count, 1)


def suggest_model_parameters(
    trial: optuna.Trial,
    model_name: str,
    positive_weight: float,
) -> Dict[str, Any]:
    if model_name == "Logistic":
        return {
            "C": trial.suggest_float(
                "C",
                1e-3,
                1e1,
                log=True,
            ),
            "class_weight": trial.suggest_categorical(
                "class_weight",
                [None, "balanced"],
            ),
        }

    if model_name == "XGBoost":
        return {
            "n_estimators": trial.suggest_int(
                "n_estimators",
                300,
                800,
                step=100,
            ),
            "max_depth": trial.suggest_int(
                "max_depth",
                3,
                9,
            ),
            "learning_rate": trial.suggest_float(
                "learning_rate",
                1e-3,
                3e-2,
                log=True,
            ),
            "min_child_weight": trial.suggest_int(
                "min_child_weight",
                1,
                10,
            ),
            "gamma": trial.suggest_float(
                "gamma",
                0.0,
                5.0,
            ),
            "subsample": trial.suggest_float(
                "subsample",
                0.5,
                0.9,
            ),
            "colsample_bytree": trial.suggest_float(
                "colsample_bytree",
                0.5,
                0.9,
            ),
            "reg_alpha": trial.suggest_float(
                "reg_alpha",
                1e-3,
                10.0,
                log=True,
            ),
            "reg_lambda": trial.suggest_float(
                "reg_lambda",
                1e-3,
                10.0,
                log=True,
            ),
            "scale_pos_weight": trial.suggest_float(
                "scale_pos_weight",
                0.5 * positive_weight,
                1.5 * positive_weight,
            ),
        }

    if model_name == "LightGBM":
        return {
            "n_estimators": trial.suggest_int(
                "n_estimators",
                300,
                800,
                step=100,
            ),
            "num_leaves": trial.suggest_int(
                "num_leaves",
                31,
                255,
            ),
            "learning_rate": trial.suggest_float(
                "learning_rate",
                1e-3,
                3e-2,
                log=True,
            ),
            "feature_fraction": trial.suggest_float(
                "feature_fraction",
                0.5,
                0.9,
            ),
            "bagging_fraction": trial.suggest_float(
                "bagging_fraction",
                0.5,
                0.9,
            ),
            "bagging_freq": trial.suggest_int(
                "bagging_freq",
                1,
                7,
            ),
            "lambda_l1": trial.suggest_float(
                "lambda_l1",
                1e-3,
                10.0,
                log=True,
            ),
            "lambda_l2": trial.suggest_float(
                "lambda_l2",
                1e-3,
                10.0,
                log=True,
            ),
            "scale_pos_weight": trial.suggest_float(
                "scale_pos_weight",
                0.5 * positive_weight,
                1.5 * positive_weight,
            ),
        }

    raise KeyError(f"未知のモデルです: {model_name}")


def build_estimator(
    model_name: str,
    parameters: Mapping[str, Any],
) -> BaseEstimator:
    params = dict(parameters)

    if model_name == "Logistic":
        return LogisticRegression(
            solver="liblinear",
            max_iter=2000,
            random_state=RANDOM_STATE,
            C=float(params["C"]),
            class_weight=params.get("class_weight"),
        )

    if model_name == "XGBoost":
        return xgb.XGBClassifier(
            objective="binary:logistic",
            eval_metric="logloss",
            tree_method="hist",
            n_jobs=MODEL_N_JOBS,
            random_state=RANDOM_STATE,
            verbosity=0,
            **params,
        )

    if model_name == "LightGBM":
        return lgb.LGBMClassifier(
            objective="binary",
            metric="binary_logloss",
            n_jobs=MODEL_N_JOBS,
            random_state=RANDOM_STATE,
            verbosity=-1,
            deterministic=True,
            force_col_wise=True,
            **params,
        )

    raise KeyError(f"未知のモデルです: {model_name}")


def clamp_warm_start_parameters(
    model_name: str,
    parameters: Mapping[str, Any],
    positive_weight: float,
) -> Dict[str, Any]:
    """
    旧解析のbest parametersを現在の探索範囲へ収める。
    """
    params = dict(parameters)

    if model_name == "Logistic":
        if "C" in params:
            params["C"] = float(
                np.clip(float(params["C"]), 1e-3, 1e1)
            )
        if params.get("class_weight") not in {
            None,
            "balanced",
        }:
            params["class_weight"] = None
        return {
            key: params[key]
            for key in ("C", "class_weight")
            if key in params
        }

    if model_name == "XGBoost":
        integer_bounds = {
            "n_estimators": (300, 800),
            "max_depth": (3, 9),
            "min_child_weight": (1, 10),
        }
        float_bounds = {
            "learning_rate": (1e-3, 3e-2),
            "gamma": (0.0, 5.0),
            "subsample": (0.5, 0.9),
            "colsample_bytree": (0.5, 0.9),
            "reg_alpha": (1e-3, 10.0),
            "reg_lambda": (1e-3, 10.0),
            "scale_pos_weight": (
                0.5 * positive_weight,
                1.5 * positive_weight,
            ),
        }
    elif model_name == "LightGBM":
        integer_bounds = {
            "n_estimators": (300, 800),
            "num_leaves": (31, 255),
            "bagging_freq": (1, 7),
        }
        float_bounds = {
            "learning_rate": (1e-3, 3e-2),
            "feature_fraction": (0.5, 0.9),
            "bagging_fraction": (0.5, 0.9),
            "lambda_l1": (1e-3, 10.0),
            "lambda_l2": (1e-3, 10.0),
            "scale_pos_weight": (
                0.5 * positive_weight,
                1.5 * positive_weight,
            ),
        }
    else:
        return {}

    cleaned: Dict[str, Any] = {}

    for key, (lower, upper) in integer_bounds.items():
        if key in params:
            value = int(
                round(
                    np.clip(
                        float(params[key]),
                        lower,
                        upper,
                    )
                )
            )
            if key == "n_estimators":
                value = int(
                    round(value / 100) * 100
                )
                value = int(
                    np.clip(value, lower, upper)
                )
            cleaned[key] = value

    for key, (lower, upper) in float_bounds.items():
        if key in params:
            cleaned[key] = float(
                np.clip(
                    float(params[key]),
                    lower,
                    upper,
                )
            )

    return cleaned


def previous_parameter_path(
    validation: str,
    outcome_label: str,
    outer_key: str,
    variant: str,
    model_name: str,
) -> Optional[Path]:
    if (
        not USE_WARM_START
        or not WARM_START_PARAMETER_DIR.exists()
    ):
        return None

    if validation == "IECV":
        facility = outer_key.replace(
            "IECV_heldout_",
            "",
        )
        filename = (
            f"{sanitize_filename(outcome_label)}__"
            f"heldout_{sanitize_filename(facility)}__"
            f"{sanitize_filename(variant)}__"
            f"{sanitize_filename(model_name)}.json"
        )
        path = (
            WARM_START_PARAMETER_DIR
            / "iecv"
            / filename
        )
    else:
        filename = (
            f"{sanitize_filename(outcome_label)}__"
            f"{sanitize_filename(variant)}__"
            f"{sanitize_filename(model_name)}.json"
        )
        path = (
            WARM_START_PARAMETER_DIR
            / "holdout"
            / filename
        )

    return path if path.exists() else None


def current_parameter_path(
    validation: str,
    outcome_label: str,
    outer_key: str,
    variant: str,
    model_name: str,
) -> Path:
    return (
        PARAMETER_DIR
        / validation.lower()
        / sanitize_filename(outcome_label)
        / sanitize_filename(outer_key)
        / (
            f"{sanitize_filename(variant)}__"
            f"{sanitize_filename(model_name)}.json"
        )
    )


def current_trials_path(
    validation: str,
    outcome_label: str,
    outer_key: str,
    variant: str,
    model_name: str,
) -> Path:
    return (
        PARAMETER_DIR
        / validation.lower()
        / sanitize_filename(outcome_label)
        / sanitize_filename(outer_key)
        / (
            f"{sanitize_filename(variant)}__"
            f"{sanitize_filename(model_name)}__trials.csv"
        )
    )


def tune_model_parameters_cached(
    *,
    validation: str,
    outcome_label: str,
    outer_key: str,
    variant: str,
    model_name: str,
    fold_cache: Sequence[Mapping[str, Any]],
    y_outer: np.ndarray,
    n_trials: int,
) -> Dict[str, Any]:
    parameter_path = current_parameter_path(
        validation,
        outcome_label,
        outer_key,
        variant,
        model_name,
    )

    if parameter_path.exists() and not FORCE_RETUNE:
        return dict(
            load_json(parameter_path)["best_parameters"]
        )

    positive_weight = positive_class_weight(
        y_outer
    )

    def objective(
        trial: optuna.Trial,
    ) -> float:
        parameters = suggest_model_parameters(
            trial,
            model_name=model_name,
            positive_weight=positive_weight,
        )

        fold_scores: List[float] = []

        for fold_index, fold in enumerate(
            fold_cache,
            start=1,
        ):
            X_train, X_validation = (
                model_arrays_from_fold(
                    fold,
                    model_name,
                )
            )
            estimator = build_estimator(
                model_name,
                parameters,
            )
            estimator.fit(
                X_train,
                fold["y_train"],
            )
            probability = estimator.predict_proba(
                X_validation
            )[:, 1]

            fold_scores.append(
                float(
                    average_precision_score(
                        fold["y_validation"],
                        probability,
                    )
                )
            )

            interim = float(
                np.mean(fold_scores)
            )
            trial.report(
                interim,
                step=fold_index,
            )

            if trial.should_prune():
                raise optuna.TrialPruned()

        return float(np.mean(fold_scores))

    study_seed = stable_seed(
        validation,
        outcome_label,
        outer_key,
        variant,
        model_name,
    )

    study = optuna.create_study(
        direction="maximize",
        sampler=optuna.samplers.TPESampler(
            seed=study_seed,
        ),
        pruner=optuna.pruners.MedianPruner(
            n_startup_trials=5,
            n_warmup_steps=1,
            interval_steps=1,
        ),
    )

    warm_path = previous_parameter_path(
        validation=validation,
        outcome_label=outcome_label,
        outer_key=outer_key,
        variant=variant,
        model_name=model_name,
    )

    if warm_path is not None:
        try:
            warm_payload = load_json(warm_path)
            warm_parameters = clamp_warm_start_parameters(
                model_name=model_name,
                parameters=warm_payload.get(
                    "best_parameters",
                    {},
                ),
                positive_weight=positive_weight,
            )
            if warm_parameters:
                study.enqueue_trial(
                    warm_parameters
                )
        except Exception as exc:
            print(
                f"Warm startをスキップしました: "
                f"{warm_path.name}: {exc}"
            )

    study.optimize(
        objective,
        n_trials=n_trials,
        n_jobs=OPTUNA_N_JOBS,
        show_progress_bar=False,
        gc_after_trial=True,
    )

    complete_trials = [
        trial
        for trial in study.trials
        if (
            trial.state == optuna.trial.TrialState.COMPLETE
            and trial.value is not None
        )
    ]
    complete_values = [float(trial.value) for trial in complete_trials]
    convergence_window = min(
        OPTUNA_CONVERGENCE_WINDOW,
        max(len(complete_values) - 1, 0),
    )
    if convergence_window > 0:
        prior_best = max(complete_values[:-convergence_window])
        final_best = max(complete_values)
        late_improvement = float(final_best - prior_best)
    else:
        late_improvement = np.nan

    best_trial_in_last_window = bool(
        study.best_trial.number
        >= max(0, len(study.trials) - OPTUNA_CONVERGENCE_WINDOW)
    )

    if (
        np.isfinite(late_improvement)
        and late_improvement > OPTUNA_LATE_IMPROVEMENT_WARNING
    ):
        warnings.warn(
            f"Optuna objective improved by {late_improvement:.4f} in the "
            f"last {convergence_window} completed trials for "
            f"{validation}/{outer_key}/{variant}/{model_name}. "
            "Review the optimization history before freezing results.",
            RuntimeWarning,
        )

    parameter_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    payload = {
        "validation": validation,
        "outcome": outcome_label,
        "outer_key": outer_key,
        "variant": variant,
        "model": model_name,
        "objective": "Mean average precision in stratified cross-validation",
        "best_value": float(study.best_value),
        "best_parameters": study.best_params,
        "n_trials_requested": n_trials,
        "n_trials_completed": len(study.trials),
        "n_trials_complete": len(complete_trials),
        "best_trial_number": int(study.best_trial.number),
        "best_trial_in_last_window": best_trial_in_last_window,
        "late_objective_improvement": late_improvement,
        "study_seed": study_seed,
        "n_pruned": int(
            sum(
                trial.state
                == optuna.trial.TrialState.PRUNED
                for trial in study.trials
            )
        ),
    }
    save_json(payload, parameter_path)

    trials_path = current_trials_path(
        validation,
        outcome_label,
        outer_key,
        variant,
        model_name,
    )
    trials_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    study.trials_dataframe().to_csv(
        trials_path,
        index=False,
        encoding="utf-8-sig",
    )

    return dict(study.best_params)


def fixed_sensitivity_parameters(
    primary_parameters: Mapping[str, Any],
    model_name: str,
    y_development: np.ndarray,
) -> Dict[str, Any]:
    parameters = dict(primary_parameters)

    if model_name in {
        "XGBoost",
        "LightGBM",
    }:
        parameters["scale_pos_weight"] = (
            positive_class_weight(y_development)
        )

    return parameters


def fit_isotonic_calibrated_bundle(
    *,
    model_name: str,
    variant: str,
    parameters: Mapping[str, Any],
    calibration_folds: Sequence[Mapping[str, Any]],
    final_preprocessed: Mapping[str, Any],
    y_development: np.ndarray,
    y_validation: np.ndarray,
    metadata: Mapping[str, Any],
) -> Tuple[Dict[str, Any], np.ndarray, np.ndarray]:
    """
    calibration foldごとのout-of-fold raw probabilityからisotonic mappingを学習し、
    development全体で再fitしたbase modelのvalidation予測へ適用する。
    """
    y_development = np.asarray(
        y_development,
        dtype=int,
    )
    oof_raw = np.full(
        len(y_development),
        np.nan,
        dtype=float,
    )

    for fold in calibration_folds:
        X_train, X_validation = (
            model_arrays_from_fold(
                fold,
                model_name,
            )
        )
        estimator = build_estimator(
            model_name,
            parameters,
        )
        estimator.fit(
            X_train,
            fold["y_train"],
        )
        fold_probability = estimator.predict_proba(
            X_validation
        )[:, 1]
        oof_raw[
            fold["validation_position"]
        ] = fold_probability

    if np.isnan(oof_raw).any():
        raise RuntimeError(
            "較正用out-of-fold予測に欠測があります。"
        )

    calibrator = IsotonicRegression(
        out_of_bounds="clip",
        y_min=0.0,
        y_max=1.0,
    )
    calibrator.fit(
        oof_raw,
        y_development,
    )

    X_development, X_validation = (
        model_arrays_from_final(
            final_preprocessed,
            model_name,
        )
    )
    final_estimator = build_estimator(
        model_name,
        parameters,
    )
    final_estimator.fit(
        X_development,
        y_development,
    )

    validation_raw = final_estimator.predict_proba(
        X_validation
    )[:, 1]
    validation_calibrated = np.asarray(
        calibrator.transform(validation_raw),
        dtype=float,
    )

    for label, values in {
        "raw": validation_raw,
        "calibrated": validation_calibrated,
    }.items():
        values = np.asarray(values, dtype=float)
        if not np.isfinite(values).all():
            raise ValueError(f"{label}予測確率に非有限値があります。")
        if ((values < 0) | (values > 1)).any():
            raise ValueError(f"{label}予測確率が0–1の範囲外です。")

    bundle = {
        "metadata": dict(metadata),
        "variant": variant,
        "model_name": model_name,
        "parameters": dict(parameters),
        "feature_names": list(
            final_preprocessed["feature_names"]
        ),
        "preprocessor": final_preprocessed[
            "preprocessor"
        ],
        "logistic_scaler": final_preprocessed[
            "logistic_scaler"
        ],
        "estimator": final_estimator,
        "calibrator": calibrator,
    }

    return (
        bundle,
        validation_raw,
        validation_calibrated,
    )


def transform_bundle_input(
    bundle: Mapping[str, Any],
    X_raw: pd.DataFrame,
) -> np.ndarray:
    model_input = bundle[
        "preprocessor"
    ].transform(X_raw).astype(np.float32)

    if bundle["model_name"] == "Logistic":
        indices = scaled_feature_indices(
            bundle["variant"]
        )
        model_input = model_input.copy()
        model_input[:, indices] = bundle[
            "logistic_scaler"
        ].transform(
            model_input[:, indices]
        ).astype(np.float32)

    return model_input


def predict_bundle(
    bundle: Mapping[str, Any],
    X_raw: pd.DataFrame,
) -> Tuple[np.ndarray, np.ndarray]:
    model_input = transform_bundle_input(
        bundle,
        X_raw,
    )
    raw_probability = bundle[
        "estimator"
    ].predict_proba(model_input)[:, 1]
    calibrated_probability = np.asarray(
        bundle["calibrator"].transform(
            raw_probability
        ),
        dtype=float,
    )
    return raw_probability, calibrated_probability

# %% [markdown]
# Cell 18
# ## 11. Outer validation split

# %%
# Cell 19
# ============================================================
# 11. 固定hold-out split・IECV outer split
# ============================================================

def fixed_holdout_indices(
    data: pd.DataFrame,
) -> Tuple[np.ndarray, np.ndarray]:
    split_path = (
        OUTPUT_DIR
        / "fixed_holdout_split.csv"
    )

    if split_path.exists():
        split_df = pd.read_csv(split_path)

        required = {ID_COL, "Set"}
        missing = required - set(split_df.columns)
        if missing:
            raise KeyError(
                f"既存splitファイルに必要列がありません: {missing}"
            )

        merged = data[[ID_COL]].merge(
            split_df[[ID_COL, "Set"]],
            on=ID_COL,
            how="left",
            validate="one_to_one",
        )

        if merged["Set"].isna().any():
            raise ValueError(
                "既存splitファイルに含まれない症例があります。"
            )

        train_indices = np.flatnonzero(
            merged["Set"].eq("Development").to_numpy()
        )
        holdout_indices = np.flatnonzero(
            merged["Set"].eq("Holdout").to_numpy()
        )
        return train_indices, holdout_indices

    all_indices = np.arange(
        len(data),
        dtype=int,
    )
    train_indices, holdout_indices = train_test_split(
        all_indices,
        test_size=0.20,
        stratify=data[
            PRIMARY_OUTCOME_COL
        ].to_numpy(dtype=int),
        random_state=RANDOM_STATE,
    )

    split_labels = np.full(
        len(data),
        "Development",
        dtype=object,
    )
    split_labels[holdout_indices] = "Holdout"

    pd.DataFrame(
        {
            ID_COL: data[ID_COL].to_numpy(),
            "Set": split_labels,
        }
    ).to_csv(
        split_path,
        index=False,
        encoding="utf-8-sig",
    )

    return (
        np.sort(train_indices),
        np.sort(holdout_indices),
    )


def fixed_temporal_indices(
    data: pd.DataFrame,
) -> Tuple[np.ndarray, np.ndarray]:
    if TEMPORAL_PERIOD_COL not in data.columns:
        raise KeyError(
            f"時間的検証列がありません: {TEMPORAL_PERIOD_COL}"
        )

    period = data[TEMPORAL_PERIOD_COL].to_numpy(dtype=int)
    development_indices = np.flatnonzero(
        period == TEMPORAL_DEVELOPMENT_VALUE
    )
    validation_indices = np.flatnonzero(
        period == TEMPORAL_VALIDATION_VALUE
    )

    if len(development_indices) == 0 or len(validation_indices) == 0:
        raise ValueError(
            "時間的検証のdevelopmentまたはvalidation cohortが空です。"
        )

    return development_indices, validation_indices


def build_outer_splits(
    data: pd.DataFrame,
) -> Tuple[
    List[Dict[str, Any]],
    List[Dict[str, Any]],
    List[Dict[str, Any]],
]:
    facilities = sorted(
        data[FACILITY_COL]
        .astype(str)
        .unique()
        .tolist()
    )

    iecv_splits: List[Dict[str, Any]] = []

    for facility in facilities:
        test_mask = (
            data[FACILITY_COL]
            .astype(str)
            .eq(str(facility))
            .to_numpy()
        )
        train_indices = np.flatnonzero(~test_mask)
        validation_indices = np.flatnonzero(test_mask)

        iecv_splits.append(
            {
                "validation": "IECV",
                "outer_key": (
                    f"IECV_heldout_{facility}"
                ),
                "held_out_facility": str(facility),
                "train_indices": train_indices,
                "validation_indices": validation_indices,
            }
        )

    development_indices, holdout_indices = (
        fixed_holdout_indices(data)
    )
    holdout_splits = [
        {
            "validation": "Holdout",
            "outer_key": "Random_holdout",
            "held_out_facility": "Random hold-out",
            "train_indices": development_indices,
            "validation_indices": holdout_indices,
        }
    ]

    temporal_splits: List[Dict[str, Any]] = []
    if RUN_PRIMARY_TEMPORAL:
        temporal_development_indices, temporal_validation_indices = (
            fixed_temporal_indices(data)
        )
        temporal_splits.append(
            {
                "validation": "Temporal",
                "outer_key": "Temporal_2021_2023_to_2024_2025Q1",
                "held_out_facility": TEMPORAL_VALIDATION_LABEL,
                "train_indices": temporal_development_indices,
                "validation_indices": temporal_validation_indices,
            }
        )

    return iecv_splits, holdout_splits, temporal_splits


IECV_SPLITS, HOLDOUT_SPLITS, TEMPORAL_SPLITS = build_outer_splits(
    df_raw
)

print(
    "IECV held-out facilities:",
    [
        split["held_out_facility"]
        for split in IECV_SPLITS
    ],
)
print(
    "Random hold-out N:",
    len(HOLDOUT_SPLITS[0]["validation_indices"]),
)
if TEMPORAL_SPLITS:
    print(
        "Temporal validation N:",
        len(TEMPORAL_SPLITS[0]["validation_indices"]),
        f"({TEMPORAL_VALIDATION_LABEL})",
    )


def validate_and_export_split_summary(
    data: pd.DataFrame,
    iecv_splits: Sequence[Mapping[str, Any]],
    holdout_splits: Sequence[Mapping[str, Any]],
    temporal_splits: Sequence[Mapping[str, Any]],
) -> Path:
    rows: List[Dict[str, Any]] = []
    all_indices = set(range(len(data)))

    for split in [*iecv_splits, *holdout_splits, *temporal_splits]:
        train_indices = np.asarray(split["train_indices"], dtype=int)
        validation_indices = np.asarray(split["validation_indices"], dtype=int)

        train_set = set(train_indices.tolist())
        validation_set = set(validation_indices.tolist())
        if train_set & validation_set:
            raise AssertionError(f"Split overlap detected: {split['outer_key']}")
        if train_set | validation_set != all_indices:
            raise AssertionError(f"Split does not cover the cohort: {split['outer_key']}")

        y_train = data.iloc[train_indices][PRIMARY_OUTCOME_COL].to_numpy(dtype=int)
        y_validation = data.iloc[validation_indices][PRIMARY_OUTCOME_COL].to_numpy(dtype=int)
        if np.unique(y_train).size < 2 or np.unique(y_validation).size < 2:
            raise ValueError(f"Both classes are required in {split['outer_key']}")

        rows.append(
            {
                "Validation": split["validation"],
                "Outer_key": split["outer_key"],
                "Held_out_facility": split["held_out_facility"],
                "Development_N": len(train_indices),
                "Development_events": int(y_train.sum()),
                "Development_prevalence": float(y_train.mean()),
                "Validation_N": len(validation_indices),
                "Validation_events": int(y_validation.sum()),
                "Validation_prevalence": float(y_validation.mean()),
            }
        )

    summary = pd.DataFrame(rows)
    path = LOG_DIR / "Validation_Split_Summary.xlsx"
    summary.to_excel(path, index=False)
    style_excel_workbook(path)
    return path


SPLIT_SUMMARY_PATH = validate_and_export_split_summary(
    df_raw,
    IECV_SPLITS,
    HOLDOUT_SPLITS,
    TEMPORAL_SPLITS,
)
print(f"Split summary: {SPLIT_SUMMARY_PATH}")

# %% [markdown]
# Cell 20
# ## 12. 主解析

# %%
# Cell 21
# ============================================================
# 12. 主解析：Event
#     - IECV：主要検証
#     - Random hold-out：二次的検証
#     - 3モデル × 2特徴量セット
# ============================================================

def bundle_path(
    *,
    analysis_group: str,
    validation: str,
    outcome_label: str,
    outer_key: str,
    variant: str,
    model_name: str,
) -> Path:
    return (
        MODEL_DIR
        / sanitize_filename(analysis_group)
        / validation.lower()
        / sanitize_filename(outcome_label)
        / sanitize_filename(outer_key)
        / (
            f"{sanitize_filename(variant)}__"
            f"{sanitize_filename(model_name)}.pkl"
        )
    )


def prediction_path(
    *,
    analysis_group: str,
    validation: str,
    outcome_label: str,
    outer_key: str,
    variant: str,
    model_name: str,
) -> Path:
    return (
        PREDICTION_DIR
        / sanitize_filename(analysis_group)
        / validation.lower()
        / sanitize_filename(outcome_label)
        / sanitize_filename(outer_key)
        / (
            f"{sanitize_filename(variant)}__"
            f"{sanitize_filename(model_name)}.csv"
        )
    )


def run_primary_outer_split(
    data: pd.DataFrame,
    split: Mapping[str, Any],
    *,
    n_trials_by_model: Mapping[str, int],
) -> pd.DataFrame:
    validation = str(split["validation"])
    outer_key = str(split["outer_key"])
    held_out_facility = str(
        split["held_out_facility"]
    )
    train_indices = np.asarray(
        split["train_indices"],
        dtype=int,
    )
    validation_indices = np.asarray(
        split["validation_indices"],
        dtype=int,
    )

    y_development = data.iloc[
        train_indices
    ][PRIMARY_OUTCOME_COL].to_numpy(dtype=int)
    y_validation = data.iloc[
        validation_indices
    ][PRIMARY_OUTCOME_COL].to_numpy(dtype=int)

    prediction_frames: List[pd.DataFrame] = []

    for variant in VARIANTS:
        print(
            f"[Primary {validation}] {outer_key} | {variant}"
        )

        final_preprocessed = (
            build_or_load_final_preprocessing(
                data=data,
                train_indices=train_indices,
                validation_indices=validation_indices,
                variant=variant,
                outer_key=outer_key,
                imputer_kind="lgbm_single",
            )
        )

        tuning_folds = (
            build_or_load_cv_preprocessing(
                data=data,
                outer_train_indices=train_indices,
                outcome_column=PRIMARY_OUTCOME_COL,
                variant=variant,
                outer_key=outer_key,
                purpose="tuning",
                n_splits=N_SPLITS_TUNING,
                seed=RANDOM_STATE,
                imputer_kind="lgbm_single",
            )
        )

        calibration_folds = (
            build_or_load_cv_preprocessing(
                data=data,
                outer_train_indices=train_indices,
                outcome_column=PRIMARY_OUTCOME_COL,
                variant=variant,
                outer_key=outer_key,
                purpose="calibration",
                n_splits=N_SPLITS_CALIBRATION,
                seed=RANDOM_STATE + 1000,
                imputer_kind="lgbm_single",
            )
        )

        for model_name in PRIMARY_MODELS:
            model_bundle_path = bundle_path(
                analysis_group="primary",
                validation=validation,
                outcome_label=PRIMARY_OUTCOME_LABEL,
                outer_key=outer_key,
                variant=variant,
                model_name=model_name,
            )
            model_prediction_path = prediction_path(
                analysis_group="primary",
                validation=validation,
                outcome_label=PRIMARY_OUTCOME_LABEL,
                outer_key=outer_key,
                variant=variant,
                model_name=model_name,
            )

            if (
                SKIP_COMPLETED
                and model_bundle_path.exists()
                and model_prediction_path.exists()
            ):
                prediction_frames.append(
                    pd.read_csv(
                        model_prediction_path
                    )
                )
                print(
                    f"  reuse: {model_name}"
                )
                continue

            best_parameters = (
                tune_model_parameters_cached(
                    validation=validation,
                    outcome_label=PRIMARY_OUTCOME_LABEL,
                    outer_key=outer_key,
                    variant=variant,
                    model_name=model_name,
                    fold_cache=tuning_folds,
                    y_outer=y_development,
                    n_trials=int(
                        n_trials_by_model[model_name]
                    ),
                )
            )

            metadata = {
                "analysis_group": "primary",
                "validation": validation,
                "outcome_label": PRIMARY_OUTCOME_LABEL,
                "outcome_column": PRIMARY_OUTCOME_COL,
                "outer_key": outer_key,
                "held_out_facility": held_out_facility,
            }

            (
                bundle,
                raw_probability,
                calibrated_probability,
            ) = fit_isotonic_calibrated_bundle(
                model_name=model_name,
                variant=variant,
                parameters=best_parameters,
                calibration_folds=calibration_folds,
                final_preprocessed=final_preprocessed,
                y_development=y_development,
                y_validation=y_validation,
                metadata=metadata,
            )

            dump_object(
                bundle,
                model_bundle_path,
            )

            prediction_frame = pd.DataFrame(
                {
                    ID_COL: data.iloc[
                        validation_indices
                    ][ID_COL].to_numpy(),
                    FACILITY_COL: data.iloc[
                        validation_indices
                    ][FACILITY_COL].astype(str).to_numpy(),
                    "Analysis_group": "Primary",
                    "Validation": validation,
                    "Outer_key": outer_key,
                    "Held_out_facility": held_out_facility,
                    "Outcome": PRIMARY_OUTCOME_LABEL,
                    "Outcome_column": PRIMARY_OUTCOME_COL,
                    "Variant": variant,
                    "Model": model_name,
                    "y_true": y_validation,
                    "y_pred_raw": raw_probability,
                    "y_pred": calibrated_probability,
                }
            )

            model_prediction_path.parent.mkdir(
                parents=True,
                exist_ok=True,
            )
            prediction_frame.to_csv(
                model_prediction_path,
                index=False,
                encoding="utf-8-sig",
            )
            prediction_frames.append(
                prediction_frame
            )

            print(
                f"  completed: {model_name}"
            )

    return pd.concat(
        prediction_frames,
        ignore_index=True,
    )



def validate_primary_prediction_integrity(
    predictions: pd.DataFrame,
    data: pd.DataFrame,
) -> None:
    required_columns = {
        ID_COL,
        "Validation",
        "Variant",
        "Model",
        "y_true",
        "y_pred",
    }
    missing = required_columns - set(predictions.columns)
    if missing:
        raise KeyError(
            f"Primary predictions are missing columns: {sorted(missing)}"
        )

    if not np.isfinite(
        predictions["y_pred"].to_numpy(dtype=float)
    ).all():
        raise ValueError(
            "Primary calibrated predictions contain nonfinite values."
        )

    def validate_group(
        *,
        validation_name: str,
        splits: Sequence[Mapping[str, Any]],
        enabled: bool,
    ) -> None:
        if not enabled:
            return
        if not splits:
            raise AssertionError(
                f"{validation_name} is enabled but no split was configured."
            )

        subset = predictions.loc[
            predictions["Validation"].eq(validation_name)
        ].copy()
        expected_indices = np.concatenate(
            [
                np.asarray(split["validation_indices"], dtype=int)
                for split in splits
            ]
        )
        expected_ids = set(
            data.iloc[expected_indices][ID_COL].tolist()
        )
        expected_rows = (
            len(expected_indices)
            * len(VARIANTS)
            * len(PRIMARY_MODELS)
        )
        if len(subset) != expected_rows:
            raise AssertionError(
                f"Unexpected {validation_name} prediction rows: "
                f"{len(subset)} vs {expected_rows}"
            )

        counts = subset.groupby(
            [ID_COL, "Variant", "Model"],
            dropna=False,
        ).size()
        if not counts.eq(1).all():
            raise AssertionError(
                f"Each {validation_name} patient must have exactly one "
                "prediction per model/variant."
            )

        if set(subset[ID_COL].tolist()) != expected_ids:
            raise AssertionError(
                f"{validation_name} predictions do not cover the configured cohort."
            )

    validate_group(
        validation_name="IECV",
        splits=IECV_SPLITS,
        enabled=RUN_PRIMARY_IECV,
    )
    validate_group(
        validation_name="Holdout",
        splits=HOLDOUT_SPLITS,
        enabled=RUN_PRIMARY_HOLDOUT,
    )
    validate_group(
        validation_name="Temporal",
        splits=TEMPORAL_SPLITS,
        enabled=RUN_PRIMARY_TEMPORAL,
    )


def run_primary_analysis(
    data: pd.DataFrame,
) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []

    if RUN_PRIMARY_IECV:
        for split in IECV_SPLITS:
            frames.append(
                run_primary_outer_split(
                    data,
                    split,
                    n_trials_by_model=N_TRIALS_IECV_BY_MODEL,
                )
            )

    if RUN_PRIMARY_HOLDOUT:
        for split in HOLDOUT_SPLITS:
            frames.append(
                run_primary_outer_split(
                    data,
                    split,
                    n_trials_by_model=N_TRIALS_HOLDOUT_BY_MODEL,
                )
            )

    if RUN_PRIMARY_TEMPORAL:
        for split in TEMPORAL_SPLITS:
            frames.append(
                run_primary_outer_split(
                    data,
                    split,
                    n_trials_by_model=N_TRIALS_TEMPORAL_BY_MODEL,
                )
            )

    if not frames:
        return pd.DataFrame()

    predictions = pd.concat(
        frames,
        ignore_index=True,
    )
    validate_primary_prediction_integrity(
        predictions,
        data,
    )

    output_path = (
        PREDICTION_DIR
        / "Primary_all_predictions.csv"
    )
    predictions.to_csv(
        output_path,
        index=False,
        encoding="utf-8-sig",
    )
    return predictions

# %% [markdown]
# Cell 22
# ## 13. 感度分析

# %%
# Cell 23
# ============================================================
# 13. 感度分析：PIおよびHard_endpoint
#     - LightGBMのみ
#     - 主解析で選択した設定を固定
#     - scale_pos_weightのみ各アウトカムのdevelopment prevalenceに合わせる
# ============================================================

SENSITIVITY_OUTCOMES = {
    INFECTION_OUTCOME_LABEL: INFECTION_OUTCOME_COL,
    NONINFECTION_OUTCOME_LABEL: NONINFECTION_OUTCOME_COL,
}


def load_primary_parameters(
    *,
    validation: str,
    outer_key: str,
    variant: str,
    model_name: str,
) -> Dict[str, Any]:
    path = current_parameter_path(
        validation=validation,
        outcome_label=PRIMARY_OUTCOME_LABEL,
        outer_key=outer_key,
        variant=variant,
        model_name=model_name,
    )

    if not path.exists():
        raise FileNotFoundError(
            f"主解析のハイパーパラメータがありません: {path}"
        )

    return dict(
        load_json(path)["best_parameters"]
    )


def save_fixed_sensitivity_parameter_record(
    *,
    validation: str,
    outcome_label: str,
    outcome_column: str,
    outer_key: str,
    variant: str,
    model_name: str,
    parameters: Mapping[str, Any],
) -> Path:
    path = (
        PARAMETER_DIR
        / "sensitivity_fixed"
        / validation.lower()
        / sanitize_filename(outcome_label)
        / sanitize_filename(outer_key)
        / (
            f"{sanitize_filename(variant)}__"
            f"{sanitize_filename(model_name)}.json"
        )
    )

    save_json(
        {
            "validation": validation,
            "outcome": outcome_label,
            "outcome_column": outcome_column,
            "outer_key": outer_key,
            "variant": variant,
            "model": model_name,
            "selection": (
                "Primary-outcome hyperparameters fixed; "
                "scale_pos_weight recalculated in the "
                "corresponding development dataset"
            ),
            "parameters": dict(parameters),
        },
        path,
    )
    return path


def run_sensitivity_outer_split(
    data: pd.DataFrame,
    split: Mapping[str, Any],
    outcome_label: str,
    outcome_column: str,
) -> pd.DataFrame:
    validation = str(split["validation"])
    outer_key = str(split["outer_key"])
    held_out_facility = str(
        split["held_out_facility"]
    )
    train_indices = np.asarray(
        split["train_indices"],
        dtype=int,
    )
    validation_indices = np.asarray(
        split["validation_indices"],
        dtype=int,
    )

    y_development = data.iloc[
        train_indices
    ][outcome_column].to_numpy(dtype=int)
    y_validation = data.iloc[
        validation_indices
    ][outcome_column].to_numpy(dtype=int)

    if np.unique(y_development).size < 2:
        raise ValueError(
            f"{outcome_label} | {outer_key}: "
            "development dataに両クラスがありません。"
        )

    prediction_frames: List[pd.DataFrame] = []

    for variant in VARIANTS:
        print(
            f"[Sensitivity {validation}] "
            f"{outcome_label} | {outer_key} | {variant}"
        )

        final_preprocessed = (
            build_or_load_final_preprocessing(
                data=data,
                train_indices=train_indices,
                validation_indices=validation_indices,
                variant=variant,
                outer_key=outer_key,
                imputer_kind="lgbm_single",
            )
        )

        calibration_folds = (
            build_or_load_cv_preprocessing(
                data=data,
                outer_train_indices=train_indices,
                outcome_column=outcome_column,
                variant=variant,
                outer_key=outer_key,
                purpose=(
                    f"calibration_{outcome_label}"
                ),
                n_splits=N_SPLITS_CALIBRATION,
                seed=RANDOM_STATE + 2000,
                imputer_kind="lgbm_single",
            )
        )

        for model_name in SENSITIVITY_MODELS:
            model_bundle_path = bundle_path(
                analysis_group="sensitivity",
                validation=validation,
                outcome_label=outcome_label,
                outer_key=outer_key,
                variant=variant,
                model_name=model_name,
            )
            model_prediction_path = prediction_path(
                analysis_group="sensitivity",
                validation=validation,
                outcome_label=outcome_label,
                outer_key=outer_key,
                variant=variant,
                model_name=model_name,
            )

            if (
                SKIP_COMPLETED
                and model_bundle_path.exists()
                and model_prediction_path.exists()
            ):
                prediction_frames.append(
                    pd.read_csv(
                        model_prediction_path
                    )
                )
                print(
                    f"  reuse: {model_name}"
                )
                continue

            primary_parameters = (
                load_primary_parameters(
                    validation=validation,
                    outer_key=outer_key,
                    variant=variant,
                    model_name=model_name,
                )
            )
            parameters = (
                fixed_sensitivity_parameters(
                    primary_parameters=primary_parameters,
                    model_name=model_name,
                    y_development=y_development,
                )
            )

            save_fixed_sensitivity_parameter_record(
                validation=validation,
                outcome_label=outcome_label,
                outcome_column=outcome_column,
                outer_key=outer_key,
                variant=variant,
                model_name=model_name,
                parameters=parameters,
            )

            metadata = {
                "analysis_group": "sensitivity",
                "validation": validation,
                "outcome_label": outcome_label,
                "outcome_column": outcome_column,
                "outer_key": outer_key,
                "held_out_facility": held_out_facility,
            }

            (
                bundle,
                raw_probability,
                calibrated_probability,
            ) = fit_isotonic_calibrated_bundle(
                model_name=model_name,
                variant=variant,
                parameters=parameters,
                calibration_folds=calibration_folds,
                final_preprocessed=final_preprocessed,
                y_development=y_development,
                y_validation=y_validation,
                metadata=metadata,
            )

            dump_object(
                bundle,
                model_bundle_path,
            )

            prediction_frame = pd.DataFrame(
                {
                    ID_COL: data.iloc[
                        validation_indices
                    ][ID_COL].to_numpy(),
                    FACILITY_COL: data.iloc[
                        validation_indices
                    ][FACILITY_COL].astype(str).to_numpy(),
                    "Analysis_group": "Sensitivity",
                    "Validation": validation,
                    "Outer_key": outer_key,
                    "Held_out_facility": held_out_facility,
                    "Outcome": outcome_label,
                    "Outcome_column": outcome_column,
                    "Variant": variant,
                    "Model": model_name,
                    "y_true": y_validation,
                    "y_pred_raw": raw_probability,
                    "y_pred": calibrated_probability,
                }
            )

            model_prediction_path.parent.mkdir(
                parents=True,
                exist_ok=True,
            )
            prediction_frame.to_csv(
                model_prediction_path,
                index=False,
                encoding="utf-8-sig",
            )
            prediction_frames.append(
                prediction_frame
            )

            print(
                f"  completed: {model_name}"
            )

    return pd.concat(
        prediction_frames,
        ignore_index=True,
    )


def run_sensitivity_analysis(
    data: pd.DataFrame,
) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []

    for outcome_label, outcome_column in (
        SENSITIVITY_OUTCOMES.items()
    ):
        if RUN_SENSITIVITY_IECV:
            for split in IECV_SPLITS:
                frames.append(
                    run_sensitivity_outer_split(
                        data=data,
                        split=split,
                        outcome_label=outcome_label,
                        outcome_column=outcome_column,
                    )
                )

        if RUN_SENSITIVITY_HOLDOUT:
            for split in HOLDOUT_SPLITS:
                frames.append(
                    run_sensitivity_outer_split(
                        data=data,
                        split=split,
                        outcome_label=outcome_label,
                        outcome_column=outcome_column,
                    )
                )

    if not frames:
        return pd.DataFrame()

    predictions = pd.concat(
        frames,
        ignore_index=True,
    )
    output_path = (
        PREDICTION_DIR
        / "Sensitivity_all_predictions.csv"
    )
    predictions.to_csv(
        output_path,
        index=False,
        encoding="utf-8-sig",
    )
    return predictions

# %% [markdown]
# Cell 24
# ## 14. 多重代入感度分析

# %%
# Cell 25
# ============================================================
# 14. 多重代入感度分析
#     - Eventのみ
#     - LightGBMのみ
#     - Preoperative / Perioperative
#     - IECV
#     - 主解析で選択したハイパーパラメータを固定
# ============================================================

def mi_prediction_path(
    *,
    outer_key: str,
    variant: str,
    imputation_index: int,
) -> Path:
    return (
        MI_DIR
        / "predictions"
        / sanitize_filename(outer_key)
        / sanitize_filename(variant)
        / f"imputation_{imputation_index:02d}.csv"
    )


def run_single_mi_fold(
    data: pd.DataFrame,
    split: Mapping[str, Any],
    *,
    variant: str,
    imputation_index: int,
) -> pd.DataFrame:
    validation = "IECV"
    outer_key = str(split["outer_key"])
    held_out_facility = str(
        split["held_out_facility"]
    )
    train_indices = np.asarray(
        split["train_indices"],
        dtype=int,
    )
    validation_indices = np.asarray(
        split["validation_indices"],
        dtype=int,
    )

    output_path = mi_prediction_path(
        outer_key=outer_key,
        variant=variant,
        imputation_index=imputation_index,
    )

    if (
        SKIP_COMPLETED
        and output_path.exists()
    ):
        return pd.read_csv(output_path)

    y_development = data.iloc[
        train_indices
    ][PRIMARY_OUTCOME_COL].to_numpy(dtype=int)
    y_validation = data.iloc[
        validation_indices
    ][PRIMARY_OUTCOME_COL].to_numpy(dtype=int)

    primary_parameters = load_primary_parameters(
        validation=validation,
        outer_key=outer_key,
        variant=variant,
        model_name="LightGBM",
    )

    final_preprocessed = (
        build_or_load_final_preprocessing(
            data=data,
            train_indices=train_indices,
            validation_indices=validation_indices,
            variant=variant,
            outer_key=outer_key,
            imputer_kind="bayesian_multiple",
            imputation_index=imputation_index,
        )
    )

    calibration_folds = (
        build_or_load_cv_preprocessing(
            data=data,
            outer_train_indices=train_indices,
            outcome_column=PRIMARY_OUTCOME_COL,
            variant=variant,
            outer_key=outer_key,
            purpose="mi_calibration",
            n_splits=N_SPLITS_MI_CALIBRATION,
            seed=RANDOM_STATE + 3000,
            imputer_kind="bayesian_multiple",
            imputation_index=imputation_index,
        )
    )

    metadata = {
        "analysis_group": "multiple_imputation",
        "validation": validation,
        "outcome_label": PRIMARY_OUTCOME_LABEL,
        "outcome_column": PRIMARY_OUTCOME_COL,
        "outer_key": outer_key,
        "held_out_facility": held_out_facility,
        "imputation_index": imputation_index,
    }

    (
        _,
        raw_probability,
        calibrated_probability,
    ) = fit_isotonic_calibrated_bundle(
        model_name="LightGBM",
        variant=variant,
        parameters=primary_parameters,
        calibration_folds=calibration_folds,
        final_preprocessed=final_preprocessed,
        y_development=y_development,
        y_validation=y_validation,
        metadata=metadata,
    )

    prediction_frame = pd.DataFrame(
        {
            ID_COL: data.iloc[
                validation_indices
            ][ID_COL].to_numpy(),
            FACILITY_COL: data.iloc[
                validation_indices
            ][FACILITY_COL].astype(str).to_numpy(),
            "Validation": "IECV",
            "Outer_key": outer_key,
            "Held_out_facility": held_out_facility,
            "Outcome": PRIMARY_OUTCOME_LABEL,
            "Variant": variant,
            "Model": "LightGBM",
            "Imputation": imputation_index,
            "y_true": y_validation,
            "y_pred_raw": raw_probability,
            "y_pred": calibrated_probability,
        }
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    prediction_frame.to_csv(
        output_path,
        index=False,
        encoding="utf-8-sig",
    )
    return prediction_frame


def summarize_multiple_imputation(
    mi_predictions: pd.DataFrame,
    primary_predictions: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    per_imputation_rows: List[Dict[str, Any]] = []

    group_columns = [
        "Imputation",
        "Variant",
    ]
    for keys, group in mi_predictions.groupby(
        group_columns,
        sort=True,
    ):
        imputation_index, variant = keys
        metrics = compute_metrics(
            group["y_true"].to_numpy(dtype=int),
            group["y_pred"].to_numpy(dtype=float),
        )
        per_imputation_rows.append(
            {
                "Imputation": imputation_index,
                "Variant": variant,
                "Model": "LightGBM",
                **metrics,
            }
        )

    per_imputation = pd.DataFrame(
        per_imputation_rows
    )

    metric_summary_rows: List[Dict[str, Any]] = []
    for variant, group in per_imputation.groupby(
        "Variant",
        sort=True,
    ):
        for metric in METRIC_NAMES:
            values = pd.to_numeric(
                group[metric],
                errors="coerce",
            ).dropna()

            metric_summary_rows.append(
                {
                    "Variant": variant,
                    "Model": "LightGBM",
                    "Metric": metric,
                    "M": int(group["Imputation"].nunique()),
                    "Mean": float(values.mean())
                    if len(values)
                    else np.nan,
                    "SD_between_imputations": float(
                        values.std(ddof=1)
                    )
                    if len(values) > 1
                    else np.nan,
                    "Minimum": float(values.min())
                    if len(values)
                    else np.nan,
                    "Maximum": float(values.max())
                    if len(values)
                    else np.nan,
                    "Percentile_2.5": float(
                        np.percentile(values, 2.5)
                    )
                    if len(values)
                    else np.nan,
                    "Percentile_97.5": float(
                        np.percentile(values, 97.5)
                    )
                    if len(values)
                    else np.nan,
                }
            )

    between_imputation_summary = pd.DataFrame(
        metric_summary_rows
    )

    pooled_predictions = (
        mi_predictions.groupby(
            [
                ID_COL,
                FACILITY_COL,
                "Held_out_facility",
                "Variant",
                "Model",
                "y_true",
            ],
            as_index=False,
        )
        .agg(
            y_pred=("y_pred", "mean"),
            y_pred_raw=("y_pred_raw", "mean"),
            Prediction_SD=("y_pred", "std"),
            M=("Imputation", "nunique"),
        )
    )

    pooled_metric_rows: List[Dict[str, Any]] = []

    for variant, group in pooled_predictions.groupby(
        "Variant",
        sort=True,
    ):
        summary = bootstrap_metric_summary(
            y_true=group["y_true"].to_numpy(dtype=int),
            y_prob=group["y_pred"].to_numpy(dtype=float),
            n_bootstrap=N_BOOTSTRAP_MI_POOLED,
            seed=RANDOM_STATE,
        )
        pooled_metric_rows.append(
            {
                "Validation": "Pooled IECV",
                "Outcome": PRIMARY_OUTCOME_LABEL,
                "Variant": variant,
                "Model": "LightGBM",
                "M": N_IMPUTATIONS,
                **summary,
            }
        )

    pooled_metrics = pd.DataFrame(
        pooled_metric_rows
    )

    comparison_rows: List[pd.DataFrame] = []

    single_iecv = primary_predictions.loc[
        (primary_predictions["Validation"] == "IECV")
        & (primary_predictions["Model"] == "LightGBM")
    ].copy()

    for variant in VARIANTS:
        single_variant = single_iecv.loc[
            single_iecv["Variant"] == variant,
            [ID_COL, "y_true", "y_pred"],
        ].rename(
            columns={
                "y_pred": "y_pred_single",
            }
        )

        mi_variant = pooled_predictions.loc[
            pooled_predictions["Variant"] == variant,
            [ID_COL, "y_true", "y_pred"],
        ].rename(
            columns={
                "y_pred": "y_pred_mi",
            }
        )

        merged = single_variant.merge(
            mi_variant,
            on=[ID_COL, "y_true"],
            how="inner",
            validate="one_to_one",
        )

        if len(merged) == 0:
            continue

        delong_result = paired_delong_test(
            y_true=merged["y_true"].to_numpy(dtype=int),
            prediction_1=merged["y_pred_single"].to_numpy(dtype=float),
            prediction_2=merged["y_pred_mi"].to_numpy(dtype=float),
        )
        comparison_rows.append(
            pd.DataFrame(
                [
                    {
                        "Variant": variant,
                        "Model": "LightGBM",
                        "Metric": "AUROC",
                        "Comparison": "Multiple minus single imputation",
                        "Method": "Paired DeLong",
                        **delong_result,
                    }
                ]
            )
        )

        for metric in (
            "AUPRC",
            "Brier",
            "Scaled_Brier",
        ):
            result = paired_bootstrap_difference(
                y_true=merged["y_true"].to_numpy(dtype=int),
                prediction_1=merged["y_pred_single"].to_numpy(dtype=float),
                prediction_2=merged["y_pred_mi"].to_numpy(dtype=float),
                metric=metric,
                n_bootstrap=N_BOOTSTRAP_MI_POOLED,
            )
            comparison_rows.append(
                pd.DataFrame(
                    [
                        {
                            "Variant": variant,
                            "Model": "LightGBM",
                            "Metric": metric,
                            "Comparison": (
                                "Multiple minus single imputation"
                            ),
                            **result,
                        }
                    ]
                )
            )

    if comparison_rows:
        comparisons = pd.concat(
            comparison_rows,
            ignore_index=True,
        )
    else:
        comparisons = pd.DataFrame()

    return {
        "per_imputation": per_imputation,
        "between_imputation_summary": between_imputation_summary,
        "pooled_predictions": pooled_predictions,
        "pooled_metrics": pooled_metrics,
        "comparisons": comparisons,
    }


def run_multiple_imputation_analysis(
    data: pd.DataFrame,
    primary_predictions: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    if not RUN_MULTIPLE_IMPUTATION:
        return {}

    all_frames: List[pd.DataFrame] = []

    for imputation_index in range(
        1,
        N_IMPUTATIONS + 1,
    ):
        print(
            f"[Multiple imputation] "
            f"{imputation_index}/{N_IMPUTATIONS}"
        )

        for split in IECV_SPLITS:
            for variant in VARIANTS:
                all_frames.append(
                    run_single_mi_fold(
                        data=data,
                        split=split,
                        variant=variant,
                        imputation_index=imputation_index,
                    )
                )

    mi_predictions = pd.concat(
        all_frames,
        ignore_index=True,
    )

    imputation_counts = (
        mi_predictions.groupby([ID_COL, "Variant", "Model"])["Imputation"]
        .nunique()
    )
    if not imputation_counts.eq(N_IMPUTATIONS).all():
        raise AssertionError(
            "Some patient/model/variant combinations do not contain all imputations."
        )

    mi_predictions.to_csv(
        MI_DIR / "MI_all_predictions.csv",
        index=False,
        encoding="utf-8-sig",
    )

    results = summarize_multiple_imputation(
        mi_predictions=mi_predictions,
        primary_predictions=primary_predictions,
    )

    workbook_path = (
        MI_DIR
        / "Multiple_Imputation_Sensitivity_Results.xlsx"
    )
    with pd.ExcelWriter(
        workbook_path,
        engine="openpyxl",
    ) as writer:
        results["pooled_metrics"].to_excel(
            writer,
            sheet_name="Pooled_metrics",
            index=False,
        )
        results["per_imputation"].to_excel(
            writer,
            sheet_name="Per_imputation",
            index=False,
        )
        results[
            "between_imputation_summary"
        ].to_excel(
            writer,
            sheet_name="Between_imputation",
            index=False,
        )
        results["comparisons"].to_excel(
            writer,
            sheet_name="Vs_single_imputation",
            index=False,
        )

    style_excel_workbook(workbook_path)

    results["pooled_predictions"].to_csv(
        MI_DIR / "MI_pooled_predictions.csv",
        index=False,
        encoding="utf-8-sig",
    )

    return results

# %% [markdown]
# Cell 26
# ## 15. 性能指標・比較表

# %%
# Cell 27
# ============================================================
# 15. 性能指標表・比較表
# ============================================================

def metric_table_from_predictions(
    predictions: pd.DataFrame,
    *,
    group_columns: Sequence[str],
    n_bootstrap: int,
    seed_offset: int = 0,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    grouped = predictions.groupby(
        list(group_columns),
        sort=True,
        dropna=False,
    )

    for group_index, (keys, group) in enumerate(
        grouped,
        start=1,
    ):
        if not isinstance(keys, tuple):
            keys = (keys,)

        identifiers = dict(
            zip(group_columns, keys)
        )
        summary = bootstrap_metric_summary(
            y_true=group["y_true"].to_numpy(dtype=int),
            y_prob=group["y_pred"].to_numpy(dtype=float),
            n_bootstrap=n_bootstrap,
            seed=RANDOM_STATE + seed_offset + group_index,
        )
        rows.append(
            {
                **identifiers,
                **summary,
            }
        )

    return pd.DataFrame(rows)


def pooled_iecv_metric_table(
    predictions: pd.DataFrame,
    *,
    n_bootstrap: int,
) -> pd.DataFrame:
    iecv = predictions.loc[
        predictions["Validation"] == "IECV"
    ].copy()

    pooled = metric_table_from_predictions(
        iecv,
        group_columns=[
            "Outcome",
            "Variant",
            "Model",
        ],
        n_bootstrap=n_bootstrap,
        seed_offset=100,
    )
    pooled.insert(
        0,
        "Validation_summary",
        "Pooled IECV",
    )
    pooled["Held_out_facility"] = "Pooled"
    return pooled


def site_iecv_metric_table(
    predictions: pd.DataFrame,
    *,
    n_bootstrap: int,
) -> pd.DataFrame:
    iecv = predictions.loc[
        predictions["Validation"] == "IECV"
    ].copy()

    site = metric_table_from_predictions(
        iecv,
        group_columns=[
            "Outcome",
            "Held_out_facility",
            "Variant",
            "Model",
        ],
        n_bootstrap=n_bootstrap,
        seed_offset=1000,
    )
    site.insert(
        0,
        "Validation_summary",
        "Held-out institution",
    )
    return site


def holdout_metric_table(
    predictions: pd.DataFrame,
    *,
    n_bootstrap: int,
) -> pd.DataFrame:
    holdout = predictions.loc[
        predictions["Validation"] == "Holdout"
    ].copy()

    if holdout.empty:
        return pd.DataFrame()

    table = metric_table_from_predictions(
        holdout,
        group_columns=[
            "Outcome",
            "Variant",
            "Model",
        ],
        n_bootstrap=n_bootstrap,
        seed_offset=2000,
    )
    table.insert(
        0,
        "Validation_summary",
        "Random hold-out",
    )
    table["Held_out_facility"] = "Random hold-out"
    return table


def temporal_metric_table(
    predictions: pd.DataFrame,
    *,
    n_bootstrap: int,
) -> pd.DataFrame:
    temporal = predictions.loc[
        predictions["Validation"] == "Temporal"
    ].copy()

    if temporal.empty:
        return pd.DataFrame()

    table = metric_table_from_predictions(
        temporal,
        group_columns=[
            "Outcome",
            "Variant",
            "Model",
        ],
        n_bootstrap=n_bootstrap,
        seed_offset=3000,
    )
    table.insert(
        0,
        "Validation_summary",
        "Temporal validation",
    )
    table["Held_out_facility"] = TEMPORAL_VALIDATION_LABEL
    table["Development_period"] = TEMPORAL_DEVELOPMENT_LABEL
    table["Validation_period"] = TEMPORAL_VALIDATION_LABEL
    return table


def pre_vs_peri_comparison_table(
    predictions: pd.DataFrame,
    *,
    n_bootstrap: int,
) -> pd.DataFrame:
    rows: List[pd.DataFrame] = []

    comparison_groups = [
        "Outcome",
        "Validation",
        "Held_out_facility",
        "Model",
    ]

    for keys, group in predictions.groupby(
        comparison_groups,
        sort=True,
        dropna=False,
    ):
        outcome, validation, held_out_facility, model_name = keys

        pre = group.loc[
            group["Variant"] == "Preoperative",
            [ID_COL, "y_true", "y_pred"],
        ].rename(
            columns={
                "y_pred": "y_pred_pre",
            }
        )
        peri = group.loc[
            group["Variant"] == "Perioperative",
            [ID_COL, "y_true", "y_pred"],
        ].rename(
            columns={
                "y_pred": "y_pred_peri",
            }
        )

        merged = pre.merge(
            peri,
            on=[ID_COL, "y_true"],
            how="inner",
            validate="one_to_one",
        )

        if merged.empty:
            continue

        result = compare_preoperative_vs_perioperative(
            y_true=merged["y_true"].to_numpy(dtype=int),
            preoperative_prediction=merged[
                "y_pred_pre"
            ].to_numpy(dtype=float),
            perioperative_prediction=merged[
                "y_pred_peri"
            ].to_numpy(dtype=float),
            n_bootstrap=n_bootstrap,
        )
        result.insert(0, "Model", model_name)
        result.insert(
            0,
            "Held_out_facility",
            held_out_facility,
        )
        result.insert(0, "Validation", validation)
        result.insert(0, "Outcome", outcome)
        rows.append(result)

    # pooled IECV comparison
    iecv = predictions.loc[
        predictions["Validation"] == "IECV"
    ].copy()

    for (outcome, model_name), group in iecv.groupby(
        ["Outcome", "Model"],
        sort=True,
    ):
        pre = group.loc[
            group["Variant"] == "Preoperative",
            [ID_COL, "y_true", "y_pred"],
        ].rename(
            columns={
                "y_pred": "y_pred_pre",
            }
        )
        peri = group.loc[
            group["Variant"] == "Perioperative",
            [ID_COL, "y_true", "y_pred"],
        ].rename(
            columns={
                "y_pred": "y_pred_peri",
            }
        )

        merged = pre.merge(
            peri,
            on=[ID_COL, "y_true"],
            how="inner",
            validate="one_to_one",
        )

        if merged.empty:
            continue

        result = compare_preoperative_vs_perioperative(
            y_true=merged["y_true"].to_numpy(dtype=int),
            preoperative_prediction=merged[
                "y_pred_pre"
            ].to_numpy(dtype=float),
            perioperative_prediction=merged[
                "y_pred_peri"
            ].to_numpy(dtype=float),
            n_bootstrap=n_bootstrap,
        )
        result.insert(0, "Model", model_name)
        result.insert(
            0,
            "Held_out_facility",
            "Pooled",
        )
        result.insert(0, "Validation", "Pooled IECV")
        result.insert(0, "Outcome", outcome)
        rows.append(result)

    if not rows:
        return pd.DataFrame()

    return pd.concat(
        rows,
        ignore_index=True,
    )



def build_and_save_metric_outputs(
    primary_predictions: pd.DataFrame,
    sensitivity_predictions: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    outputs: Dict[str, pd.DataFrame] = {}

    primary_site = site_iecv_metric_table(
        primary_predictions,
        n_bootstrap=N_BOOTSTRAP_PRIMARY,
    )
    primary_pooled = pooled_iecv_metric_table(
        primary_predictions,
        n_bootstrap=N_BOOTSTRAP_PRIMARY,
    )
    primary_holdout = holdout_metric_table(
        primary_predictions,
        n_bootstrap=N_BOOTSTRAP_PRIMARY,
    )
    primary_temporal = temporal_metric_table(
        primary_predictions,
        n_bootstrap=N_BOOTSTRAP_PRIMARY,
    )
    primary_comparisons = (
        pre_vs_peri_comparison_table(
            primary_predictions,
            n_bootstrap=N_BOOTSTRAP_PRIMARY,
        )
    )

    outputs.update(
        {
            "primary_site": primary_site,
            "primary_pooled": primary_pooled,
            "primary_holdout": primary_holdout,
            "primary_temporal": primary_temporal,
            "primary_comparisons": primary_comparisons,
        }
    )

    primary_workbook = (
        METRIC_DIR
        / "Primary_Validation_Results.xlsx"
    )
    with pd.ExcelWriter(
        primary_workbook,
        engine="openpyxl",
    ) as writer:
        primary_pooled.to_excel(
            writer,
            sheet_name="Pooled_IECV",
            index=False,
        )
        primary_site.to_excel(
            writer,
            sheet_name="Site_IECV",
            index=False,
        )
        primary_holdout.to_excel(
            writer,
            sheet_name="Random_holdout",
            index=False,
        )
        primary_temporal.to_excel(
            writer,
            sheet_name="Temporal_validation",
            index=False,
        )
        primary_comparisons.to_excel(
            writer,
            sheet_name="Pre_vs_Peri",
            index=False,
        )

    style_excel_workbook(primary_workbook)

    if not sensitivity_predictions.empty:
        sensitivity_site = site_iecv_metric_table(
            sensitivity_predictions,
            n_bootstrap=N_BOOTSTRAP_SENSITIVITY,
        )
        sensitivity_pooled = pooled_iecv_metric_table(
            sensitivity_predictions,
            n_bootstrap=N_BOOTSTRAP_SENSITIVITY,
        )
        sensitivity_holdout = holdout_metric_table(
            sensitivity_predictions,
            n_bootstrap=N_BOOTSTRAP_SENSITIVITY,
        )
        sensitivity_comparisons = (
            pre_vs_peri_comparison_table(
                sensitivity_predictions,
                n_bootstrap=N_BOOTSTRAP_SENSITIVITY,
            )
        )

        outputs.update(
            {
                "sensitivity_site": sensitivity_site,
                "sensitivity_pooled": sensitivity_pooled,
                "sensitivity_holdout": sensitivity_holdout,
                "sensitivity_comparisons": sensitivity_comparisons,
            }
        )

        sensitivity_workbook = (
            METRIC_DIR
            / "Sensitivity_Outcome_Results.xlsx"
        )
        with pd.ExcelWriter(
            sensitivity_workbook,
            engine="openpyxl",
        ) as writer:
            sensitivity_pooled.to_excel(
                writer,
                sheet_name="Pooled_IECV",
                index=False,
            )
            sensitivity_site.to_excel(
                writer,
                sheet_name="Site_IECV",
                index=False,
            )
            sensitivity_holdout.to_excel(
                writer,
                sheet_name="Random_holdout",
                index=False,
            )
            sensitivity_comparisons.to_excel(
                writer,
                sheet_name="Pre_vs_Peri",
                index=False,
            )

        style_excel_workbook(
            sensitivity_workbook
        )


    return outputs

# %% [markdown]
# Cell 28
# ## 16. 投稿用Tables

# %%
# Cell 29
# ============================================================
# 16. 投稿用Tableおよびハイパーパラメータ表
# ============================================================

def format_estimate_ci(
    row: pd.Series,
    metric: str,
    digits: int = 3,
) -> str:
    estimate = row.get(metric, np.nan)
    lower = row.get(f"{metric}_lo", np.nan)
    upper = row.get(f"{metric}_hi", np.nan)

    if not np.isfinite(estimate):
        return "NA"

    if np.isfinite(lower) and np.isfinite(upper):
        return (
            f"{estimate:.{digits}f} "
            f"({lower:.{digits}f}–{upper:.{digits}f})"
        )

    return f"{estimate:.{digits}f}"


def create_publication_validation_tables(
    metric_outputs: Mapping[str, pd.DataFrame],
    mi_outputs: Mapping[str, pd.DataFrame],
) -> Dict[str, Path]:
    outputs: Dict[str, Path] = {}

    primary_site = metric_outputs[
        "primary_site"
    ].copy()
    primary_pooled = metric_outputs[
        "primary_pooled"
    ].copy()

    main_table = pd.concat(
        [
            primary_site,
            primary_pooled,
        ],
        ignore_index=True,
        sort=False,
    )

    main_table["Institution"] = np.where(
        main_table["Held_out_facility"].eq("Pooled"),
        "Pooled IECV",
        main_table["Held_out_facility"].map(facility_display_label),
    )

    formatted = pd.DataFrame(
        {
            "Institution": main_table["Institution"],
            "Feature set": main_table["Variant"],
            "Model": main_table["Model"],
            "N": main_table["N"],
            "Events, n (%)": (
                main_table["Events"].astype(int).astype(str)
                + " ("
                + (
                    main_table["Prevalence"] * 100
                ).map(lambda value: f"{value:.1f}")
                + "%)"
            ),
            "AUROC (95% CI)": main_table.apply(
                format_estimate_ci,
                axis=1,
                metric="AUROC",
            ),
            "PR-AUC (average precision; 95% CI)": main_table.apply(
                format_estimate_ci,
                axis=1,
                metric="AUPRC",
            ),
            "Brier score (95% CI)": main_table.apply(
                format_estimate_ci,
                axis=1,
                metric="Brier",
            ),
            "Scaled Brier score (95% CI)": main_table.apply(
                format_estimate_ci,
                axis=1,
                metric="Scaled_Brier",
            ),
            "Calibration slope (95% CI)": main_table.apply(
                format_estimate_ci,
                axis=1,
                metric="Calibration_slope",
            ),
            "Calibration intercept (95% CI)": main_table.apply(
                format_estimate_ci,
                axis=1,
                metric="Calibration_intercept",
            ),
        }
    )

    main_path = (
        PUBLICATION_TABLE_DIR
        / "Table2_Primary_IECV_Validation.xlsx"
    )
    with pd.ExcelWriter(
        main_path,
        engine="openpyxl",
    ) as writer:
        formatted.to_excel(
            writer,
            sheet_name="Formatted",
            index=False,
        )
        main_table.to_excel(
            writer,
            sheet_name="Numeric",
            index=False,
        )
        metric_outputs[
            "primary_comparisons"
        ].to_excel(
            writer,
            sheet_name="Pre_vs_Peri",
            index=False,
        )

    style_excel_workbook(main_path)
    outputs["Table2"] = main_path

    holdout_path = (
        PUBLICATION_TABLE_DIR
        / "TableSx_Random_Holdout_Validation.xlsx"
    )
    metric_outputs[
        "primary_holdout"
    ].to_excel(
        holdout_path,
        index=False,
    )
    style_excel_workbook(holdout_path)
    outputs["Holdout"] = holdout_path

    temporal_path = (
        PUBLICATION_TABLE_DIR
        / "TableSx_Temporal_Validation.xlsx"
    )
    temporal_comparisons = metric_outputs[
        "primary_comparisons"
    ].loc[
        metric_outputs["primary_comparisons"]["Validation"].eq("Temporal")
    ].copy()
    with pd.ExcelWriter(
        temporal_path,
        engine="openpyxl",
    ) as writer:
        metric_outputs[
            "primary_temporal"
        ].to_excel(
            writer,
            sheet_name="Temporal_metrics",
            index=False,
        )
        temporal_comparisons.to_excel(
            writer,
            sheet_name="Pre_vs_Peri",
            index=False,
        )
    style_excel_workbook(temporal_path)
    outputs["Temporal_validation"] = temporal_path

    if "sensitivity_pooled" in metric_outputs:
        sensitivity_path = (
            PUBLICATION_TABLE_DIR
            / "TableSx_Sensitivity_Outcomes.xlsx"
        )
        with pd.ExcelWriter(
            sensitivity_path,
            engine="openpyxl",
        ) as writer:
            metric_outputs[
                "sensitivity_pooled"
            ].to_excel(
                writer,
                sheet_name="Pooled_IECV",
                index=False,
            )
            metric_outputs[
                "sensitivity_site"
            ].to_excel(
                writer,
                sheet_name="Site_IECV",
                index=False,
            )
            metric_outputs[
                "sensitivity_holdout"
            ].to_excel(
                writer,
                sheet_name="Random_holdout",
                index=False,
            )
            metric_outputs[
                "sensitivity_comparisons"
            ].to_excel(
                writer,
                sheet_name="Pre_vs_Peri",
                index=False,
            )

        style_excel_workbook(
            sensitivity_path
        )
        outputs["Sensitivity"] = sensitivity_path

    if mi_outputs:
        mi_path = (
            PUBLICATION_TABLE_DIR
            / "TableSx_Multiple_Imputation.xlsx"
        )
        with pd.ExcelWriter(
            mi_path,
            engine="openpyxl",
        ) as writer:
            mi_outputs[
                "pooled_metrics"
            ].to_excel(
                writer,
                sheet_name="Pooled_metrics",
                index=False,
            )
            mi_outputs[
                "between_imputation_summary"
            ].to_excel(
                writer,
                sheet_name="Between_imputation",
                index=False,
            )
            mi_outputs[
                "comparisons"
            ].to_excel(
                writer,
                sheet_name="Vs_single_imputation",
                index=False,
            )

        style_excel_workbook(mi_path)
        outputs["Multiple_imputation"] = mi_path

    return outputs


def hyperparameter_search_space_table() -> pd.DataFrame:
    rows = [
        {
            "Model": "Logistic regression",
            "Hyperparameter": "C",
            "Search space": "0.001–10",
            "Sampling scale": "Log-uniform",
        },
        {
            "Model": "Logistic regression",
            "Hyperparameter": "class_weight",
            "Search space": "None; balanced",
            "Sampling scale": "Categorical",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "n_estimators",
            "Search space": "300–800 in increments of 100",
            "Sampling scale": "Discrete integer",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "max_depth",
            "Search space": "3–9",
            "Sampling scale": "Integer",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "learning_rate",
            "Search space": "0.001–0.03",
            "Sampling scale": "Log-uniform",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "min_child_weight",
            "Search space": "1–10",
            "Sampling scale": "Integer",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "gamma",
            "Search space": "0–5",
            "Sampling scale": "Uniform",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "subsample",
            "Search space": "0.5–0.9",
            "Sampling scale": "Uniform",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "colsample_bytree",
            "Search space": "0.5–0.9",
            "Sampling scale": "Uniform",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "reg_alpha",
            "Search space": "0.001–10",
            "Sampling scale": "Log-uniform",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "reg_lambda",
            "Search space": "0.001–10",
            "Sampling scale": "Log-uniform",
        },
        {
            "Model": "XGBoost",
            "Hyperparameter": "scale_pos_weight",
            "Search space": "0.5w–1.5w",
            "Sampling scale": "Uniform",
        },
        {
            "Model": "LightGBM",
            "Hyperparameter": "n_estimators",
            "Search space": "300–800 in increments of 100",
            "Sampling scale": "Discrete integer",
        },
        {
            "Model": "LightGBM",
            "Hyperparameter": "num_leaves",
            "Search space": "31–255",
            "Sampling scale": "Integer",
        },
        {
            "Model": "LightGBM",
            "Hyperparameter": "learning_rate",
            "Search space": "0.001–0.03",
            "Sampling scale": "Log-uniform",
        },
        {
            "Model": "LightGBM",
            "Hyperparameter": "feature_fraction",
            "Search space": "0.5–0.9",
            "Sampling scale": "Uniform",
        },
        {
            "Model": "LightGBM",
            "Hyperparameter": "bagging_fraction",
            "Search space": "0.5–0.9",
            "Sampling scale": "Uniform",
        },
        {
            "Model": "LightGBM",
            "Hyperparameter": "bagging_freq",
            "Search space": "1–7",
            "Sampling scale": "Integer",
        },
        {
            "Model": "LightGBM",
            "Hyperparameter": "lambda_l1",
            "Search space": "0.001–10",
            "Sampling scale": "Log-uniform",
        },
        {
            "Model": "LightGBM",
            "Hyperparameter": "lambda_l2",
            "Search space": "0.001–10",
            "Sampling scale": "Log-uniform",
        },
        {
            "Model": "LightGBM",
            "Hyperparameter": "scale_pos_weight",
            "Search space": "0.5w–1.5w",
            "Sampling scale": "Uniform",
        },
    ]
    table = pd.DataFrame(rows)
    table["Footnote"] = (
        "w = number of negative cases / number of positive cases "
        "within the corresponding development dataset."
    )
    return table


def fixed_setting_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Procedure": "Hyperparameter optimization",
                "Setting": "Sampler",
                "Value": "Optuna tree-structured Parzen estimator",
            },
            {
                "Procedure": "Hyperparameter optimization",
                "Setting": "Objective",
                "Value": "Mean average precision in five-fold stratified cross-validation",
            },
            {
                "Procedure": "Hyperparameter optimization",
                "Setting": "Trials: IECV",
                "Value": "; ".join(
                    f"{model}={trials}"
                    for model, trials in N_TRIALS_IECV_BY_MODEL.items()
                ),
            },
            {
                "Procedure": "Hyperparameter optimization",
                "Setting": "Trials: random hold-out",
                "Value": "; ".join(
                    f"{model}={trials}"
                    for model, trials in N_TRIALS_HOLDOUT_BY_MODEL.items()
                ),
            },
            {
                "Procedure": "Hyperparameter optimization",
                "Setting": "Trials: temporal validation",
                "Value": "; ".join(
                    f"{model}={trials}"
                    for model, trials in N_TRIALS_TEMPORAL_BY_MODEL.items()
                ),
            },
            {
                "Procedure": "Temporal validation",
                "Setting": "Development and validation periods",
                "Value": (
                    f"{TEMPORAL_DEVELOPMENT_LABEL} to "
                    f"{TEMPORAL_VALIDATION_LABEL}"
                ),
            },
            {
                "Procedure": "Probability calibration",
                "Setting": "Method",
                "Value": (
                    "Isotonic regression fitted to out-of-fold "
                    "predictions from three-fold stratified cross-validation"
                ),
            },
            {
                "Procedure": "Missing-data preprocessing",
                "Setting": "Primary analysis",
                "Value": (
                    "IterativeImputer with LGBMRegressor; "
                    "mean initialization; maximum five iterations"
                ),
            },
            {
                "Procedure": "Sensitivity outcomes",
                "Setting": "Hyperparameters",
                "Value": (
                    "Primary-outcome LightGBM parameters fixed; "
                    "scale_pos_weight recalculated"
                ),
            },
            {
                "Procedure": "All models",
                "Setting": "Random seed",
                "Value": RANDOM_STATE,
            },
        ]
    )


def collect_selected_hyperparameters() -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    for path in (
        sorted(PARAMETER_DIR.glob("iecv/**/*.json"))
        + sorted(PARAMETER_DIR.glob("holdout/**/*.json"))
        + sorted(PARAMETER_DIR.glob("temporal/**/*.json"))
    ):
        payload = load_json(path)
        parameters = payload.get(
            "best_parameters",
            {},
        )
        base = {
            "Validation": payload.get("validation"),
            "Outcome": payload.get("outcome"),
            "Outer_key": payload.get("outer_key"),
            "Variant": payload.get("variant"),
            "Model": payload.get("model"),
            "Best_mean_CV_average_precision": payload.get(
                "best_value"
            ),
            "N_trials_completed": payload.get(
                "n_trials_completed"
            ),
            "N_trials_complete": payload.get("n_trials_complete"),
            "Best_trial_number": payload.get("best_trial_number"),
            "Best_trial_in_last_window": payload.get("best_trial_in_last_window"),
            "Late_objective_improvement": payload.get("late_objective_improvement"),
            "N_pruned": payload.get("n_pruned"),
        }

        for parameter, value in parameters.items():
            rows.append(
                {
                    **base,
                    "Hyperparameter": parameter,
                    "Selected_value": value,
                }
            )

    return pd.DataFrame(rows)


def collect_sensitivity_fixed_parameters() -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    for path in sorted(
        (
            PARAMETER_DIR
            / "sensitivity_fixed"
        ).glob("**/*.json")
    ):
        payload = load_json(path)
        base = {
            "Validation": payload.get("validation"),
            "Outcome": payload.get("outcome"),
            "Outer_key": payload.get("outer_key"),
            "Variant": payload.get("variant"),
            "Model": payload.get("model"),
            "Selection": payload.get("selection"),
        }

        for parameter, value in payload.get(
            "parameters",
            {},
        ).items():
            rows.append(
                {
                    **base,
                    "Hyperparameter": parameter,
                    "Applied_value": value,
                }
            )

    return pd.DataFrame(rows)


def create_hyperparameter_workbook() -> Path:
    path = (
        PUBLICATION_TABLE_DIR
        / "TableS2_Hyperparameters.xlsx"
    )

    selected = collect_selected_hyperparameters()
    sensitivity_fixed = (
        collect_sensitivity_fixed_parameters()
    )

    with pd.ExcelWriter(
        path,
        engine="openpyxl",
    ) as writer:
        hyperparameter_search_space_table().to_excel(
            writer,
            sheet_name="Search_spaces",
            index=False,
        )
        fixed_setting_table().to_excel(
            writer,
            sheet_name="Fixed_settings",
            index=False,
        )
        selected.to_excel(
            writer,
            sheet_name="Primary_selected_long",
            index=False,
        )
        sensitivity_fixed.to_excel(
            writer,
            sheet_name="Sensitivity_fixed",
            index=False,
        )

        if not selected.empty:
            selected_wide = (
                selected.pivot_table(
                    index=[
                        "Validation",
                        "Outcome",
                        "Outer_key",
                        "Variant",
                        "Model",
                    ],
                    columns="Hyperparameter",
                    values="Selected_value",
                    aggfunc="first",
                )
                .reset_index()
            )
            selected_wide.columns.name = None
            selected_wide.to_excel(
                writer,
                sheet_name="Primary_selected_wide",
                index=False,
            )

    style_excel_workbook(path)
    return path

# %% [markdown]
# Cell 30
# ## 17. 投稿用Figure（最終セルで生成）

# %%
# Cell 31
# ============================================================
# 17. 投稿用Figure
# ============================================================
# 旧Figure生成コードは使用しません。
# 投稿用Figure 2–4、random hold-out Figure、追加SHAP Figureは、
# SHAP解析完了後にNotebook最終セルの統一コードで生成します。

# %% [markdown]
# Cell 32
# ## 18. IECV out-of-sample SHAP（3モデル）

# %%
# Cell 33
# ============================================================
# 18. IECV out-of-sample SHAP：3モデル
#     - Primary composite outcome
#     - Preoperative and perioperative predictor sets
#     - Logistic / XGBoost / LightGBM
#     - Isotonic calibration前のbase modelを説明
# ============================================================

SHAP_FEATURE_LABELS = {
    "ASA": "ASA-PS",
    "DeliMed": "Delirium-associated medication",
    "AntiCa": "Calcium-channel blocker",
    "ResectNum": "Number of resections",
    "HighRiskProc": "High-risk procedure",
    "RBC Tx": "RBC transfusion",
    "FFP Tx": "FFP transfusion",
    "PLT Tx": "Platelet transfusion",
    "FluidBal": "Intraoperative fluid balance",
    "OpTime": "Operative duration",
    "HR at 6h": "Heart rate at 6 h",
    "MAP at 6h": "Mean arterial pressure at 6 h",
}


def shap_feature_label(
    feature: str,
) -> str:
    return SHAP_FEATURE_LABELS.get(
        feature,
        feature,
    )


def stratified_sample_positions(
    y: np.ndarray,
    max_n: int,
    seed: int,
) -> np.ndarray:
    y_array = np.asarray(y, dtype=int)
    n = len(y_array)

    if n <= max_n:
        return np.arange(n, dtype=int)

    rng = np.random.default_rng(seed)
    positive = np.flatnonzero(y_array == 1)
    negative = np.flatnonzero(y_array == 0)

    if len(positive) == 0 or len(negative) == 0:
        return np.sort(
            rng.choice(
                n,
                size=max_n,
                replace=False,
            )
        )

    target_positive = int(
        round(
            max_n * len(positive) / n
        )
    )
    target_positive = max(
        1,
        min(target_positive, len(positive)),
    )
    target_negative = max_n - target_positive

    if target_negative > len(negative):
        target_negative = len(negative)
        target_positive = min(
            max_n - target_negative,
            len(positive),
        )

    sampled_positive = rng.choice(
        positive,
        size=target_positive,
        replace=False,
    )
    sampled_negative = rng.choice(
        negative,
        size=target_negative,
        replace=False,
    )

    return np.sort(
        np.concatenate(
            [
                sampled_positive,
                sampled_negative,
            ]
        ).astype(int)
    )


def normalize_binary_shap_values(
    raw_values: Any,
    n_samples: int,
    n_features: int,
) -> np.ndarray:
    if hasattr(raw_values, "values"):
        raw_values = raw_values.values

    if isinstance(raw_values, list):
        if len(raw_values) == 0:
            raise ValueError("SHAP値が空です。")
        raw_values = raw_values[-1]

    array = np.asarray(raw_values)

    if array.ndim == 2:
        if array.shape != (
            n_samples,
            n_features,
        ):
            raise ValueError(
                f"SHAP値の形状が不正です: {array.shape}"
            )
        return array.astype(float)

    if array.ndim == 3:
        if array.shape == (
            n_samples,
            n_features,
            2,
        ):
            return array[:, :, 1].astype(float)

        if array.shape == (
            n_samples,
            2,
            n_features,
        ):
            return array[:, 1, :].astype(float)

        if array.shape == (
            2,
            n_samples,
            n_features,
        ):
            return array[1, :, :].astype(float)

    raise ValueError(
        f"解釈できないSHAP値の形状です: {array.shape}"
    )


def calculate_model_shap_values(
    model_name: str,
    estimator: BaseEstimator,
    X_model: np.ndarray,
    X_background: Optional[np.ndarray],
) -> np.ndarray:
    n_samples, n_features = X_model.shape

    if model_name == "Logistic":
        if X_background is None:
            raise ValueError(
                "Logistic SHAPにはbackground dataが必要です。"
            )

        explainer = shap.LinearExplainer(
            estimator,
            X_background,
        )

        try:
            raw_values = explainer.shap_values(
                X_model
            )
        except AttributeError:
            raw_values = explainer(
                X_model
            )

    elif model_name in {
        "XGBoost",
        "LightGBM",
    }:
        explainer = shap.TreeExplainer(
            estimator,
            model_output="raw",
        )

        try:
            raw_values = explainer.shap_values(
                X_model,
                check_additivity=False,
            )
        except TypeError:
            raw_values = explainer.shap_values(
                X_model
            )

    else:
        raise KeyError(
            f"未対応のSHAPモデルです: {model_name}"
        )

    return normalize_binary_shap_values(
        raw_values,
        n_samples=n_samples,
        n_features=n_features,
    )


def calculate_iecv_shap_for_model(
    data: pd.DataFrame,
    model_name: str,
    variant: str,
) -> Dict[str, Any]:
    if variant not in VARIANT_FEATURES:
        raise KeyError(f"Unknown SHAP feature set: {variant}")

    model_output_dir = (
        SHAP_DIR
        / sanitize_filename(model_name)
        / sanitize_filename(variant)
    )
    model_output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    importance_path = (
        model_output_dir
        / "SHAP_importance.xlsx"
    )
    values_path = (
        model_output_dir
        / "SHAP_values.csv.gz"
    )
    inputs_path = (
        model_output_dir
        / "SHAP_model_inputs.csv.gz"
    )

    if (
        SKIP_COMPLETED
        and importance_path.exists()
        and values_path.exists()
        and inputs_path.exists()
    ):
        pooled_importance = pd.read_excel(
            importance_path,
            sheet_name="Pooled_importance",
        )
        site_importance = pd.read_excel(
            importance_path,
            sheet_name="By_site_importance",
        )
        sample_summary = pd.read_excel(
            importance_path,
            sheet_name="Sample_summary",
        )
        shap_values_df = pd.read_csv(
            values_path,
            compression="gzip",
        )
        model_inputs_df = pd.read_csv(
            inputs_path,
            compression="gzip",
        )

        metadata_columns = [
            ID_COL,
            FACILITY_COL,
            "Held_out_facility",
            "y_true",
        ]
        feature_names = list(
            VARIANT_FEATURES[
                variant
            ]
        )

        return {
            "model": model_name,
            "variant": variant,
            "feature_names": feature_names,
            "shap_values": shap_values_df[
                [
                    f"SHAP__{feature}"
                    for feature in feature_names
                ]
            ].to_numpy(dtype=float),
            "model_inputs": model_inputs_df[
                [
                    f"MODEL_INPUT__{feature}"
                    for feature in feature_names
                ]
            ].to_numpy(dtype=float),
            "metadata": shap_values_df[
                metadata_columns
            ].copy(),
            "pooled_importance": pooled_importance,
            "site_importance": site_importance,
            "sample_summary": sample_summary,
        }

    feature_names = list(
        VARIANT_FEATURES[
            variant
        ]
    )

    shap_blocks: List[np.ndarray] = []
    input_blocks: List[np.ndarray] = []
    metadata_blocks: List[pd.DataFrame] = []
    sample_summary_rows: List[Dict[str, Any]] = []

    for facility_index, split in enumerate(
        IECV_SPLITS,
        start=1,
    ):
        outer_key = str(split["outer_key"])
        held_out_facility = str(
            split["held_out_facility"]
        )
        train_indices = np.asarray(
            split["train_indices"],
            dtype=int,
        )
        validation_indices = np.asarray(
            split["validation_indices"],
            dtype=int,
        )

        model_bundle_path = bundle_path(
            analysis_group="primary",
            validation="IECV",
            outcome_label=PRIMARY_OUTCOME_LABEL,
            outer_key=outer_key,
            variant=variant,
            model_name=model_name,
        )
        if not model_bundle_path.exists():
            raise FileNotFoundError(
                f"SHAP対象モデルがありません: {model_bundle_path}"
            )

        bundle = load_object(
            model_bundle_path
        )

        y_validation_full = data.iloc[
            validation_indices
        ][PRIMARY_OUTCOME_COL].to_numpy(dtype=int)

        sampled_positions = stratified_sample_positions(
            y=y_validation_full,
            max_n=SHAP_MAX_PER_FACILITY,
            seed=RANDOM_STATE + facility_index,
        )
        sampled_indices = validation_indices[
            sampled_positions
        ]

        X_sample_raw = data.iloc[
            sampled_indices
        ][feature_names]
        X_model = transform_bundle_input(
            bundle,
            X_sample_raw,
        )

        if model_name == "Logistic":
            rng = np.random.default_rng(
                RANDOM_STATE + facility_index
            )
            background_n = min(
                SHAP_BACKGROUND_N,
                len(train_indices),
            )
            background_indices = rng.choice(
                train_indices,
                size=background_n,
                replace=False,
            )
            X_background_raw = data.iloc[
                background_indices
            ][feature_names]
            X_background = transform_bundle_input(
                bundle,
                X_background_raw,
            )
        else:
            X_background = None

        shap_values = calculate_model_shap_values(
            model_name=model_name,
            estimator=bundle["estimator"],
            X_model=X_model,
            X_background=X_background,
        )

        shap_blocks.append(
            shap_values
        )
        input_blocks.append(
            X_model
        )
        metadata_blocks.append(
            pd.DataFrame(
                {
                    ID_COL: data.iloc[
                        sampled_indices
                    ][ID_COL].to_numpy(),
                    FACILITY_COL: data.iloc[
                        sampled_indices
                    ][FACILITY_COL].astype(str).to_numpy(),
                    "Held_out_facility": held_out_facility,
                    "Model": model_name,
                    "Variant": variant,
                    "y_true": data.iloc[
                        sampled_indices
                    ][PRIMARY_OUTCOME_COL].to_numpy(dtype=int),
                }
            )
        )
        sample_summary_rows.append(
            {
                "Model": model_name,
                "Variant": variant,
                "Held_out_facility": held_out_facility,
                "Source_N": len(validation_indices),
                "Source_events": int(
                    y_validation_full.sum()
                ),
                "Source_prevalence": float(
                    y_validation_full.mean()
                ),
                "SHAP_sample_N": len(
                    sampled_indices
                ),
                "SHAP_sample_events": int(
                    data.iloc[
                        sampled_indices
                    ][PRIMARY_OUTCOME_COL].sum()
                ),
            }
        )

        print(
            f"[SHAP] {model_name} | {variant} | "
            f"held-out {held_out_facility} | "
            f"N={len(sampled_indices):,}"
        )

    pooled_shap = np.vstack(
        shap_blocks
    )
    pooled_input = np.vstack(
        input_blocks
    )
    pooled_metadata = pd.concat(
        metadata_blocks,
        ignore_index=True,
    )

    mean_abs = np.abs(
        pooled_shap
    ).mean(axis=0)
    pooled_importance = pd.DataFrame(
        {
            "Rank": np.arange(
                1,
                len(feature_names) + 1,
            ),
            "Model": model_name,
            "Variant": variant,
            "Predictor": feature_names,
            "Feature": [
                shap_feature_label(feature)
                for feature in feature_names
            ],
            "Mean_abs_SHAP": mean_abs,
        }
    ).sort_values(
        "Mean_abs_SHAP",
        ascending=False,
    ).reset_index(drop=True)
    pooled_importance["Rank"] = np.arange(
        1,
        len(pooled_importance) + 1,
    )
    pooled_total = float(pooled_importance["Mean_abs_SHAP"].sum())
    pooled_importance["Relative_importance_percent"] = (
        100 * pooled_importance["Mean_abs_SHAP"] / pooled_total
        if pooled_total > 0
        else np.nan
    )

    site_rows: List[Dict[str, Any]] = []
    for facility in sorted(
        pooled_metadata[
            "Held_out_facility"
        ].astype(str).unique()
    ):
        mask = (
            pooled_metadata[
                "Held_out_facility"
            ].astype(str).to_numpy()
            == facility
        )
        site_mean = np.abs(
            pooled_shap[mask]
        ).mean(axis=0)
        site_total = float(
            site_mean.sum()
        )

        for feature, value in zip(
            feature_names,
            site_mean,
        ):
            site_rows.append(
                {
                    "Model": model_name,
                    "Variant": variant,
                    "Held_out_facility": facility,
                    "Predictor": feature,
                    "Feature": shap_feature_label(
                        feature
                    ),
                    "Mean_abs_SHAP": float(
                        value
                    ),
                    "Relative_importance_percent": (
                        100
                        * float(value)
                        / site_total
                        if site_total > 0
                        else np.nan
                    ),
                }
            )

    site_importance = pd.DataFrame(
        site_rows
    )
    site_importance[
        "Rank_within_facility"
    ] = (
        site_importance.groupby(
            "Held_out_facility"
        )["Mean_abs_SHAP"]
        .rank(
            method="first",
            ascending=False,
        )
        .astype(int)
    )
    sample_summary = pd.DataFrame(
        sample_summary_rows
    )

    with pd.ExcelWriter(
        importance_path,
        engine="openpyxl",
    ) as writer:
        pooled_importance.to_excel(
            writer,
            sheet_name="Pooled_importance",
            index=False,
        )
        site_importance.to_excel(
            writer,
            sheet_name="By_site_importance",
            index=False,
        )
        sample_summary.to_excel(
            writer,
            sheet_name="Sample_summary",
            index=False,
        )
        pd.DataFrame(
            {
                "Item": [
                    "Outcome",
                    "Feature set",
                    "Validation",
                    "Explained model",
                    "Scale",
                    "Cross-model comparison",
                ],
                "Description": [
                    PRIMARY_OUTCOME_LABEL,
                    variant,
                    (
                        "Out-of-sample SHAP values calculated "
                        "in each IECV held-out institution"
                    ),
                    (
                        "Base model before isotonic probability "
                        "calibration"
                    ),
                    (
                        "Logistic regression: log-odds scale; "
                        "XGBoost and LightGBM: raw margin scale"
                    ),
                    (
                        "Absolute SHAP magnitudes should not be "
                        "directly compared across algorithms; "
                        "rankings are reported separately."
                    ),
                ],
            }
        ).to_excel(
            writer,
            sheet_name="Notes",
            index=False,
        )

    style_excel_workbook(
        importance_path
    )

    shap_values_output = pd.concat(
        [
            pooled_metadata,
            pd.DataFrame(
                pooled_shap,
                columns=[
                    f"SHAP__{feature}"
                    for feature in feature_names
                ],
            ),
        ],
        axis=1,
    )
    model_inputs_output = pd.concat(
        [
            pooled_metadata,
            pd.DataFrame(
                pooled_input,
                columns=[
                    f"MODEL_INPUT__{feature}"
                    for feature in feature_names
                ],
            ),
        ],
        axis=1,
    )

    if SHAP_SAVE_VALUES:
        shap_values_output.to_csv(
            values_path,
            index=False,
            compression="gzip",
            encoding="utf-8-sig",
        )
        model_inputs_output.to_csv(
            inputs_path,
            index=False,
            compression="gzip",
            encoding="utf-8-sig",
        )

    return {
        "model": model_name,
        "variant": variant,
        "feature_names": feature_names,
        "shap_values": pooled_shap,
        "model_inputs": pooled_input,
        "metadata": pooled_metadata,
        "pooled_importance": pooled_importance,
        "site_importance": site_importance,
        "sample_summary": sample_summary,
    }



def create_shap_rank_comparison(
    results: Mapping[str, Mapping[str, Any]],
    variant: str,
) -> Path:
    frames: List[pd.DataFrame] = []

    for model_name, result in results.items():
        frame = result[
            "pooled_importance"
        ][
            [
                "Predictor",
                "Feature",
                "Rank",
                "Relative_importance_percent",
            ]
        ].copy()
        frame = frame.rename(
            columns={
                "Rank": f"{model_name}_rank",
                "Relative_importance_percent": (
                    f"{model_name}_relative_importance_percent"
                ),
            }
        )
        frames.append(frame)

    comparison = frames[0]
    for frame in frames[1:]:
        comparison = comparison.merge(
            frame,
            on=[
                "Predictor",
                "Feature",
            ],
            how="outer",
            validate="one_to_one",
        )

    path = (
        PUBLICATION_TABLE_DIR
        / f"TableSx_SHAP_Rank_Comparison_{sanitize_filename(variant)}.xlsx"
    )
    comparison.to_excel(
        path,
        index=False,
    )
    style_excel_workbook(path)
    return path


def run_shap_analysis(
    data: pd.DataFrame,
) -> Dict[str, Any]:
    if not RUN_SHAP:
        return {}

    all_results: Dict[
        str,
        Dict[str, Dict[str, Any]],
    ] = {}
    rank_tables: Dict[str, Path] = {}

    for variant in SHAP_VARIANTS:
        variant_results: Dict[str, Dict[str, Any]] = {}

        for model_name in SHAP_MODELS:
            variant_results[model_name] = (
                calculate_iecv_shap_for_model(
                    data=data,
                    model_name=model_name,
                    variant=variant,
                )
            )

        all_results[variant] = variant_results
        rank_tables[variant] = (
            create_shap_rank_comparison(
                variant_results,
                variant=variant,
            )
        )

    return {
        "results": all_results,
        "rank_tables": rank_tables,
    }

# %% [markdown]
# Cell 34
# ## 19. 一括実行

# %%
# Cell 35
# ============================================================
# 19. 一括実行：解析・Table・SHAP
# ============================================================

analysis_start_time = time.time()
manifest_path = write_run_manifest()

print("=" * 72)
print("JMIR revision submission-ready analysis")
print("=" * 72)
print(f"Manifest: {manifest_path}")
print(f"Output:   {OUTPUT_DIR}")

# 1. 主解析（IECV、random hold-out、temporal validation）
primary_predictions = run_primary_analysis(df_raw)

# 2. アウトカム感度分析
sensitivity_predictions = run_sensitivity_analysis(df_raw)

# 3. 多重代入感度分析
mi_outputs = run_multiple_imputation_analysis(
    data=df_raw,
    primary_predictions=primary_predictions,
)

# 4. 性能指標
metric_outputs = build_and_save_metric_outputs(
    primary_predictions=primary_predictions,
    sensitivity_predictions=sensitivity_predictions,
)

# 5. 投稿用Tables
publication_table_outputs: Dict[str, Path] = {}
if RUN_PUBLICATION_TABLES:
    publication_table_outputs.update(
        create_publication_validation_tables(
            metric_outputs=metric_outputs,
            mi_outputs=mi_outputs,
        )
    )
    publication_table_outputs["Hyperparameters"] = (
        create_hyperparameter_workbook()
    )

# 6. SHAP：3モデル × 2特徴量セット
shap_outputs = run_shap_analysis(df_raw)

# 7. Core output一覧
output_rows: List[Dict[str, Any]] = []
for label, path in publication_table_outputs.items():
    output_rows.append(
        {
            "Category": "Publication table",
            "Label": label,
            "Path": str(path),
        }
    )

if shap_outputs:
    for variant, path in shap_outputs["rank_tables"].items():
        output_rows.append(
            {
                "Category": "SHAP table",
                "Label": f"Rank comparison: {variant}",
                "Path": str(path),
            }
        )

core_output_manifest = pd.DataFrame(output_rows)
core_output_manifest_path = OUTPUT_DIR / "Core_Output_File_Manifest.xlsx"
core_output_manifest.to_excel(core_output_manifest_path, index=False)
style_excel_workbook(core_output_manifest_path)

elapsed_seconds = time.time() - analysis_start_time
save_json(
    {
        "core_analysis_elapsed_seconds": elapsed_seconds,
        "core_analysis_elapsed_hours": elapsed_seconds / 3600,
        "core_analysis_completed_at_local": pd.Timestamp.now().isoformat(),
        "output_directory": str(OUTPUT_DIR),
    },
    LOG_DIR / "core_runtime_summary.json",
)

print("")
print("=" * 72)
print("Core analysis completed")
print("=" * 72)
print(f"Elapsed time: {elapsed_seconds / 3600:.2f} hours")
print(f"Core output manifest: {core_output_manifest_path}")
print("Run the final cell to create the submission figures and final manifest.")

# %% [markdown]
# Cell 36
# ## 主な出力
# 
# ```text
# C:\Users\tears\OneDrive\Revice_JMIR
# ├─ 00_feature_dictionary
# ├─ 01_cache
# ├─ 02_hyperparameters
# ├─ 03_models
# ├─ 04_predictions
# ├─ 05_metrics
# ├─ 06_multiple_imputation
# ├─ 07_shap
# ├─ 08_publication_tables
# ├─ 09_publication_figures
# ├─ 10_logs
# ├─ fixed_holdout_split.csv
# ├─ TableSx_Temporal_Validation.xlsx
# └─ Output_File_Manifest.xlsx
# ```
# 
# ### SHAP
# 
# - `Figure4A_SHAP_Logistic.png`
# - `Figure4B_SHAP_XGBoost.png`
# - `Figure4C_SHAP_LightGBM.png`
# - `TableSx_SHAP_Rank_Comparison.xlsx`
# 
# SHAP値は各IECV foldのheld-out施設で算出し、isotonic calibration前のbase modelを説明します。  
# モデル間でSHAP値の絶対量のスケールが異なるため、アルゴリズム間では順位を中心に比較します。
# 
# ### 再実行
# 
# `SKIP_COMPLETED=True`では、既存の前処理キャッシュ、ハイパーパラメータ、モデル、予測値を再利用します。  
# 完全に再実行する場合のみ、以下を変更してください。
# 
# ```python
# SKIP_COMPLETED = False
# FORCE_RETUNE = True
# FORCE_REBUILD_PREPROCESSING = True
# ```

# %%
# Cell 37
# ============================================================
# JMIR Revision: JMIR-compliant publication figures
#
# This standalone script reuses previously generated outputs:
#   1) Table2_Primary_IECV_Validation.xlsx
#   2) Primary_all_predictions.csv
#   3) SHAP_values.csv.gz / SHAP_model_inputs.csv.gz
#
# It does NOT refit any model.
#
# Main manuscript figures:
#   - Figure 2: IECV performance summary
#   - Figure 3: IECV decision curves
#   - Figure 4: LightGBM SHAP summary
#
# Multimedia appendix figures:
#   - Random hold-out performance
#   - Temporal validation performance
#   - Logistic regression SHAP summary
#   - XGBoost SHAP summary
#
# JMIR-oriented output rules implemented:
#   - PNG only
#   - white, nontransparent background
#   - maximum 1200 × 1200 pixels
#   - maximum 5 MB
#   - no figure number or full caption embedded in the artwork
#   - color is supplemented by marker shape, fill, or line type
#   - captions are exported separately for entry into JMIR metadata
# ============================================================


from pathlib import Path
from typing import Dict, List, Tuple
import math
import warnings

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from PIL import Image
from IPython.display import display

from sklearn.calibration import calibration_curve
from sklearn.metrics import (
    precision_recall_curve,
    roc_curve,
)

import shap


# ============================================================
# 0. Paths and output settings
# ============================================================

ANALYSIS_ROOT_DIR = OUTPUT_DIR
ROOT_DIR = ANALYSIS_ROOT_DIR

TABLE2_PATH = (
    ROOT_DIR
    / "08_publication_tables"
    / "Table2_Primary_IECV_Validation.xlsx"
)

PRIMARY_PREDICTION_PATH = (
    ROOT_DIR
    / "04_predictions"
    / "Primary_all_predictions.csv"
)

SHAP_DIR = ROOT_DIR / "07_shap"

MAIN_FIGURE_DIR = (
    ROOT_DIR
    / "09_publication_figures_JMIR"
)
MULTIMEDIA_APPENDIX_DIR = (
    ROOT_DIR
    / "10_multimedia_appendix_figures_JMIR"
)

MAIN_FIGURE_DIR.mkdir(
    parents=True,
    exist_ok=True,
)
MULTIMEDIA_APPENDIX_DIR.mkdir(
    parents=True,
    exist_ok=True,
)

CAPTION_FILE = (
    ROOT_DIR
    / "JMIR_Figure_Captions_Submission_Ready.txt"
)

# Submission restrictions
PNG_MAX_WIDTH = 1200
PNG_MAX_HEIGHT = 1200
PNG_MAX_FILE_SIZE_MB = 5.0
PNG_DPI = 180

# Figure settings
DCA_THRESHOLD_MIN = 0.01
DCA_THRESHOLD_MAX = 0.30
DCA_N_THRESHOLDS = 100

SHAP_TOP_N = 15
SHAP_PANEL_WIDTH = 1100
SHAP_PANEL_HEIGHT = 535
SHAP_PANEL_GAP = 20

# Model and feature-set display
MODEL_ORDER = [
    "Logistic",
    "XGBoost",
    "LightGBM",
]

MODEL_DISPLAY = {
    "Logistic": "Logistic regression",
    "XGBoost": "XGBoost",
    "LightGBM": "LightGBM",
}

# Color represents the algorithm
MODEL_COLORS = {
    "Logistic": "#0072B2",
    "XGBoost": "#D55E00",
    "LightGBM": "#009E73",
}

# Marker/fill or line style represents the feature set
VARIANT_ORDER = [
    "Preoperative",
    "Perioperative",
]

VARIANT_DISPLAY = {
    "Preoperative": "Preoperative",
    "Perioperative": "Perioperative",
}

VARIANT_MARKERS = {
    "Preoperative": "o",
    "Perioperative": "s",
}

VARIANT_LINESTYLES = {
    "Preoperative": ":",
    "Perioperative": "-",
}

# Internal facility values are mapped to the anonymous labels fixed above.
FACILITY_DISPLAY = {
    **FACILITY_LABEL_MAP,
    "A": "A",
    "B": "B",
    "C": "C",
    "Pooled": "Pooled IECV",
    "Pooled IECV": "Pooled IECV",
    "Random hold-out": "Random hold-out",
    TEMPORAL_VALIDATION_LABEL: TEMPORAL_VALIDATION_LABEL,
}

# Expanded feature labels for figures
FEATURE_LABELS = {
    "ASA": "ASA-PS",
    "Age": "Age",
    "Male": "Male sex",
    "BMI": "Body mass index",
    "Dialysis": "Maintenance dialysis",
    "CHF": "Congestive heart failure",
    "Malig": "Malignancy",
    "Alb": "Serum albumin",
    "BUN": "Blood urea nitrogen",
    "CRP": "C-reactive protein",
    "Cre": "Serum creatinine",
    "Hb": "Hemoglobin",
    "K": "Serum potassium",
    "Na": "Serum sodium",
    "PLT": "Platelet count",
    "T-Bil": "Total bilirubin",
    "WBC": "White blood cell count",
    "DeliMed": "Delirium-associated medication",
    "β-blocker": "Beta-blocker",
    "Oral steroids": "Oral corticosteroid",
    "Antiplatelet": "Antiplatelet agent",
    "Anticoag": "Anticoagulant",
    "AntiCa": "Calcium-channel blocker",
    "Opioid": "Opioid",
    "Proc-Eye": "Ophthalmic procedure",
    "Proc-Face/Neck": "Face or neck procedure",
    "Proc-Thorax": "Thoracic procedure",
    "Proc-MSK": "Musculoskeletal procedure",
    "Proc-ENT": "Otorhinolaryngologic procedure",
    "Proc-Neuro": "Neurosurgical procedure",
    "Proc-Genital": "Genital procedure",
    "Proc-Urinary": "Urinary tract procedure",
    "Proc-Skin": "Skin or soft-tissue procedure",
    "Proc-Abd": "Abdominal procedure",
    "ResectNum": "Number of resections",
    "HighRiskProc": "High-risk procedure",
    "OpTime": "Operative duration",
    "RBC Tx": "Red blood cell transfusion",
    "FFP Tx": "Fresh frozen plasma transfusion",
    "PLT Tx": "Platelet transfusion",
    "FluidBal": "Intraoperative fluid balance",
    "HR at 6h": "Heart rate at 6 h",
    "MAP at 6h": "Mean arterial pressure at 6 h",
}


# ============================================================
# 1. General utilities
# ============================================================

def require_file(path: Path, description: str) -> None:
    if not path.exists():
        raise FileNotFoundError(
            f"{description} was not found:\n{path}"
        )


def facility_label(value: object) -> str:
    text = str(value)
    return FACILITY_DISPLAY.get(text, text)


def image_has_transparency(
    image: Image.Image,
) -> bool:
    """
    Return True if the image contains any transparent pixels.
    """
    if image.mode in {"RGBA", "LA"}:
        alpha = image.getchannel("A")
        minimum_alpha, _ = alpha.getextrema()
        return minimum_alpha < 255

    if image.mode == "P" and "transparency" in image.info:
        return True

    return False


def save_pil_png_compliant(
    image: Image.Image,
    output_path: Path,
    *,
    dpi: int = PNG_DPI,
    max_width: int = PNG_MAX_WIDTH,
    max_height: int = PNG_MAX_HEIGHT,
    max_file_size_mb: float = PNG_MAX_FILE_SIZE_MB,
) -> Path:
    """
    Save a lossless, nontransparent PNG within the JMIR upload limits:
      - width <= 1200 px
      - height <= 1200 px
      - file size <= 5 MB

    PNG compression is lossless. The function first preserves the largest
    possible dimensions and increases compression before reducing dimensions.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    max_bytes = int(
        max_file_size_mb * 1024 * 1024
    )

    # Explicit white background; no transparent areas.
    source = image.convert("RGBA")
    white_background = Image.new(
        "RGBA",
        source.size,
        (255, 255, 255, 255),
    )
    white_background.alpha_composite(source)
    img = white_background.convert("RGB")

    img.thumbnail(
        (max_width, max_height),
        Image.Resampling.LANCZOS,
    )

    compression_levels = [4, 6, 9]

    while True:
        saved = False

        for compression_level in compression_levels:
            img.save(
                output_path,
                format="PNG",
                optimize=False,
                compress_level=compression_level,
                dpi=(dpi, dpi),
            )

            if output_path.stat().st_size <= max_bytes:
                saved = True
                break

        if saved:
            break

        new_width = max(
            300,
            int(round(img.width * 0.95)),
        )
        new_height = max(
            300,
            int(round(img.height * 0.95)),
        )

        if (
            new_width == img.width
            and new_height == img.height
        ):
            raise RuntimeError(
                f"Could not reduce the PNG to the required size: "
                f"{output_path}"
            )

        img = img.resize(
            (new_width, new_height),
            Image.Resampling.LANCZOS,
        )

    return output_path


def save_figure_png_compliant(
    fig: plt.Figure,
    output_path: Path,
    *,
    dpi: int = PNG_DPI,
    bbox_inches: str = "tight",
    pad_inches: float = 0.05,
) -> Path:
    """
    Render a matplotlib figure on a white background and save it as a
    JMIR-compliant PNG.
    """
    output_path = Path(output_path)
    temporary_path = output_path.with_name(
        f"{output_path.stem}__temporary.png"
    )

    fig.savefig(
        temporary_path,
        dpi=dpi,
        bbox_inches=bbox_inches,
        pad_inches=pad_inches,
        facecolor="white",
        edgecolor="white",
        transparent=False,
    )

    with Image.open(temporary_path) as image:
        image.load()
        copied = image.copy()

    temporary_path.unlink(
        missing_ok=True
    )

    return save_pil_png_compliant(
        copied,
        output_path,
        dpi=dpi,
    )


def validate_png(path: Path) -> Dict[str, object]:
    """
    Validate the JMIR figure upload requirements.
    """
    with Image.open(path) as image:
        image.load()
        width, height = image.size
        file_format = image.format
        mode = image.mode
        has_transparency = image_has_transparency(image)

    file_size_mb = (
        path.stat().st_size
        / (1024 ** 2)
    )

    extension_ok = (
        path.suffix.lower() == ".png"
    )
    format_ok = (
        file_format == "PNG"
    )
    width_ok = (
        width <= PNG_MAX_WIDTH
    )
    height_ok = (
        height <= PNG_MAX_HEIGHT
    )
    size_ok = (
        file_size_mb <= PNG_MAX_FILE_SIZE_MB
    )
    transparency_ok = (
        not has_transparency
    )

    valid = all(
        [
            extension_ok,
            format_ok,
            width_ok,
            height_ok,
            size_ok,
            transparency_ok,
        ]
    )

    return {
        "File": path.name,
        "Format": file_format,
        "Mode": mode,
        "Width_px": width,
        "Height_px": height,
        "File_size_MB": round(
            file_size_mb,
            3,
        ),
        "Has_transparency": has_transparency,
        "Extension_OK": extension_ok,
        "Format_OK": format_ok,
        "Width_OK": width_ok,
        "Height_OK": height_ok,
        "File_size_OK": size_ok,
        "No_transparency_OK": transparency_ok,
        "All_requirements_met": valid,
    }


def net_benefit(
    y_true: np.ndarray,
    y_probability: np.ndarray,
    threshold: float,
) -> float:
    y = np.asarray(
        y_true,
        dtype=int,
    )
    p = np.asarray(
        y_probability,
        dtype=float,
    )

    n = len(y)
    if n == 0:
        return np.nan

    predicted_positive = p >= threshold
    true_positive = int(
        np.sum(
            predicted_positive
            & (y == 1)
        )
    )
    false_positive = int(
        np.sum(
            predicted_positive
            & (y == 0)
        )
    )

    return (
        true_positive / n
        - false_positive / n
        * threshold
        / (1.0 - threshold)
    )


def create_model_feature_legend_handles() -> List[Line2D]:
    handles: List[Line2D] = []

    # Algorithm colors
    for model in MODEL_ORDER:
        handles.append(
            Line2D(
                [0],
                [0],
                color=MODEL_COLORS[model],
                linewidth=2.5,
                label=MODEL_DISPLAY[model],
            )
        )

    # Feature-set marker/fill
    handles.extend(
        [
            Line2D(
                [0],
                [0],
                marker=VARIANT_MARKERS["Preoperative"],
                markerfacecolor="white",
                markeredgecolor="black",
                color="none",
                markersize=7,
                label="Preoperative feature set",
            ),
            Line2D(
                [0],
                [0],
                marker=VARIANT_MARKERS["Perioperative"],
                markerfacecolor="black",
                markeredgecolor="black",
                color="none",
                markersize=7,
                label="Perioperative feature set",
            ),
        ]
    )

    return handles


# ============================================================
# 2. Read existing results
# ============================================================

require_file(
    TABLE2_PATH,
    "Table 2 workbook",
)
require_file(
    PRIMARY_PREDICTION_PATH,
    "Primary prediction file",
)

table2_numeric = pd.read_excel(
    TABLE2_PATH,
    sheet_name="Numeric",
)

primary_predictions = pd.read_csv(
    PRIMARY_PREDICTION_PATH,
)

required_prediction_columns = {
    ID_COL,
    "Validation",
    "Held_out_facility",
    "Variant",
    "Model",
    "y_true",
    "y_pred",
}

missing_prediction_columns = (
    required_prediction_columns
    - set(primary_predictions.columns)
)

if missing_prediction_columns:
    raise KeyError(
        "The prediction file is missing columns: "
        f"{sorted(missing_prediction_columns)}"
    )


# ============================================================
# 3. Revised Figure 2
#    - no lines connecting institutions
#    - colors = algorithms
#    - marker/fill = feature set
#    - panel D = calibration slope for all models
# ============================================================

def create_revised_figure2(
    metric_table: pd.DataFrame,
) -> Path:
    data = metric_table.copy()

    if "Outcome" in data.columns:
        data = data.loc[
            data["Outcome"]
            .astype(str)
            .eq("Primary_composite")
        ].copy()

    institution_order = [
        institution
        for institution in [
            "A",
            "B",
            "C",
            "Pooled IECV",
        ]
        if institution in set(
            data["Institution"].astype(str)
        )
    ]

    if not institution_order:
        raise ValueError(
            "No IECV institutions were found in the Numeric sheet."
        )

    # A larger gap separates pooled IECV from the held-out institutions.
    base_position = {
        "A": 0.0,
        "B": 1.0,
        "C": 2.0,
        "Pooled IECV": 3.35,
    }

    model_offset = {
        "Logistic": -0.18,
        "XGBoost": 0.00,
        "LightGBM": 0.18,
    }

    variant_offset = {
        "Preoperative": -0.035,
        "Perioperative": 0.035,
    }

    metric_specs = [
        (
            "AUROC",
            "AUROC_lo",
            "AUROC_hi",
            "(A) AUROC",
        ),
        (
            "AUPRC",
            "AUPRC_lo",
            "AUPRC_hi",
            "(B) PR-AUC",
        ),
        (
            "Scaled_Brier",
            "Scaled_Brier_lo",
            "Scaled_Brier_hi",
            "(C) Scaled Brier score",
        ),
        (
            "Calibration_slope",
            "Calibration_slope_lo",
            "Calibration_slope_hi",
            "(D) Calibration slope",
        ),
    ]

    fig, axes = plt.subplots(
        2,
        2,
        figsize=(11.2, 9.3),
    )
    axes = axes.ravel()

    for ax, (
        metric,
        lower_column,
        upper_column,
        title,
    ) in zip(
        axes,
        metric_specs,
    ):
        for model in MODEL_ORDER:
            for variant in VARIANT_ORDER:
                subset = data.loc[
                    data["Model"].eq(model)
                    & data["Variant"].eq(variant)
                    & data["Institution"].isin(
                        institution_order
                    )
                ].copy()

                if subset.empty:
                    continue

                subset["Institution"] = pd.Categorical(
                    subset["Institution"],
                    categories=institution_order,
                    ordered=True,
                )
                subset = subset.sort_values(
                    "Institution"
                )

                x = np.asarray(
                    [
                        base_position[str(institution)]
                        + model_offset[model]
                        + variant_offset[variant]
                        for institution in subset["Institution"]
                    ],
                    dtype=float,
                )

                estimate = subset[
                    metric
                ].to_numpy(dtype=float)
                lower = subset[
                    lower_column
                ].to_numpy(dtype=float)
                upper = subset[
                    upper_column
                ].to_numpy(dtype=float)

                y_error = np.vstack(
                    [
                        estimate - lower,
                        upper - estimate,
                    ]
                )

                is_preoperative = (
                    variant == "Preoperative"
                )

                ax.errorbar(
                    x,
                    estimate,
                    yerr=y_error,
                    fmt=VARIANT_MARKERS[variant],
                    linestyle="none",
                    color=MODEL_COLORS[model],
                    markeredgecolor=MODEL_COLORS[model],
                    markerfacecolor=(
                        "white"
                        if is_preoperative
                        else MODEL_COLORS[model]
                    ),
                    markersize=5.5,
                    capsize=2.5,
                    elinewidth=0.9,
                    markeredgewidth=1.2,
                    zorder=3,
                )


        if metric == "Scaled_Brier":
            ax.axhline(
                0.0,
                color="gray",
                linestyle="--",
                linewidth=0.9,
                zorder=1,
            )

        if metric == "Calibration_slope":
            ax.axhline(
                1.0,
                color="gray",
                linestyle="--",
                linewidth=0.9,
                zorder=1,
            )

        ax.axvline(
            2.67,
            color="#BDBDBD",
            linestyle=":",
            linewidth=0.8,
            zorder=1,
        )

        ax.set_xticks(
            [
                base_position[
                    institution
                ]
                for institution in institution_order
            ]
        )
        ax.set_xticklabels(
            institution_order,
            rotation=20,
            ha="right",
        )
        ax.set_title(
            title,
            fontweight="bold",
        )
        ax.grid(
            axis="y",
            alpha=0.22,
        )

        finite_lower = pd.to_numeric(
            data[lower_column],
            errors="coerce",
        ).dropna()
        finite_upper = pd.to_numeric(
            data[upper_column],
            errors="coerce",
        ).dropna()

        if len(finite_lower) and len(finite_upper):
            minimum = float(
                finite_lower.min()
            )
            maximum = float(
                finite_upper.max()
            )
            span = max(
                maximum - minimum,
                0.05,
            )
            margin = 0.10 * span

            if metric == "Scaled_Brier":
                minimum = min(
                    -0.01,
                    minimum - margin,
                )
            else:
                minimum -= margin

            maximum += margin
            ax.set_ylim(
                minimum,
                maximum,
            )

    legend_handles = (
        create_model_feature_legend_handles()
    )

    fig.legend(
        handles=legend_handles,
        loc="lower center",
        bbox_to_anchor=(0.5, 0.01),
        ncol=3,
        frameon=False,
        fontsize=8.5,
    )

    fig.tight_layout(
        rect=[0, 0.10, 1, 1]
    )

    output_path = (
        MAIN_FIGURE_DIR
        / "Figure2_IECV_performance.png"
    )
    save_figure_png_compliant(
        fig,
        output_path,
    )
    plt.close(fig)

    return output_path


# ============================================================
# 4. Revised Figure 3
#    - all 3 algorithms
#    - both preoperative and perioperative feature sets
#    - six model curves in each panel
#    - A, B, C, and pooled IECV panels
#    - threshold range 0.01–0.30
# ============================================================

def create_revised_figure3_dca(
    predictions: pd.DataFrame,
) -> Path:
    """
    Create IECV decision curves for all six model-feature-set combinations:

      - Preoperative—Logistic regression
      - Perioperative—Logistic regression
      - Preoperative—XGBoost
      - Perioperative—XGBoost
      - Preoperative—LightGBM
      - Perioperative—LightGBM

    Color identifies the algorithm.
    Line style identifies the feature set.
    """
    iecv = predictions.loc[
        predictions["Validation"]
        .astype(str)
        .eq("IECV")
        & predictions["Model"]
        .astype(str)
        .isin(MODEL_ORDER)
        & predictions["Variant"]
        .astype(str)
        .isin(VARIANT_ORDER)
    ].copy()

    if iecv.empty:
        raise ValueError(
            "No IECV predictions for the specified models and feature sets "
            "were found."
        )

    facility_values = sorted(
        iecv["Held_out_facility"]
        .astype(str)
        .unique()
        .tolist(),
        key=lambda value: facility_label(value),
    )

    panel_specs: List[
        Tuple[str, pd.DataFrame]
    ] = []

    for facility in facility_values:
        panel_specs.append(
            (
                facility_label(facility),
                iecv.loc[
                    iecv["Held_out_facility"]
                    .astype(str)
                    .eq(facility)
                ].copy(),
            )
        )

    panel_specs.append(
        (
            "Pooled IECV",
            iecv.copy(),
        )
    )

    if len(panel_specs) != 4:
        raise ValueError(
            "Figure 3 expects three held-out institutions plus pooled IECV. "
            f"Detected {len(panel_specs) - 1} held-out institutions."
        )

    thresholds = np.linspace(
        DCA_THRESHOLD_MIN,
        DCA_THRESHOLD_MAX,
        DCA_N_THRESHOLDS,
    )

    fig, axes = plt.subplots(
        2,
        2,
        figsize=(10.6, 9.4),
        sharex=True,
        sharey=True,
    )
    axes = axes.ravel()

    all_model_net_benefits: List[float] = []

    for ax, (
        panel_label,
        panel_data,
    ) in zip(
        axes,
        panel_specs,
    ):
        # Use one model-feature-set combination only to count patients/events.
        reference = panel_data.loc[
            panel_data["Model"].eq(MODEL_ORDER[0])
            & panel_data["Variant"].eq(VARIANT_ORDER[0])
        ].copy()

        if reference.empty:
            reference = panel_data.drop_duplicates(
                subset=[ID_COL]
            ).copy()

        y_reference = reference[
            "y_true"
        ].to_numpy(dtype=int)

        prevalence = float(
            np.mean(y_reference)
        )

        # Six model curves: 3 algorithms × 2 feature sets.
        for model in MODEL_ORDER:
            for variant in VARIANT_ORDER:
                group = panel_data.loc[
                    panel_data["Model"].eq(model)
                    & panel_data["Variant"].eq(variant)
                ].copy()

                if group.empty:
                    warnings.warn(
                        (
                            f"No DCA predictions found for panel={panel_label}, "
                            f"model={model}, variant={variant}."
                        ),
                        RuntimeWarning,
                    )
                    continue

                # Ensure one prediction per patient for this combination.
                if group[ID_COL].duplicated().any():
                    duplicate_n = int(
                        group[ID_COL].duplicated().sum()
                    )
                    raise ValueError(
                        (
                            f"Duplicate predictions were found for "
                            f"panel={panel_label}, model={model}, "
                            f"variant={variant}: {duplicate_n} duplicates."
                        )
                    )

                y = group[
                    "y_true"
                ].to_numpy(dtype=int)
                p = group[
                    "y_pred"
                ].to_numpy(dtype=float)

                benefits = np.asarray(
                    [
                        net_benefit(
                            y,
                            p,
                            threshold,
                        )
                        for threshold in thresholds
                    ],
                    dtype=float,
                )

                all_model_net_benefits.extend(
                    benefits[
                        np.isfinite(benefits)
                    ].tolist()
                )

                ax.plot(
                    thresholds,
                    benefits,
                    color=MODEL_COLORS[model],
                    linestyle=VARIANT_LINESTYLES[variant],
                    linewidth=(
                        1.8
                        if variant == "Perioperative"
                        else 1.55
                    ),
                    alpha=(
                        1.0
                        if variant == "Perioperative"
                        else 0.92
                    ),
                    label=(
                        f"{VARIANT_DISPLAY[variant]}—"
                        f"{MODEL_DISPLAY[model]}"
                    ),
                    zorder=3,
                )

        treat_none = np.zeros_like(
            thresholds
        )
        treat_all = (
            prevalence
            - (1.0 - prevalence)
            * thresholds
            / (1.0 - thresholds)
        )

        ax.plot(
            thresholds,
            treat_none,
            color="gray",
            linestyle="--",
            linewidth=1.0,
            label="Treat none",
            zorder=2,
        )
        ax.plot(
            thresholds,
            treat_all,
            color="black",
            linestyle="-.",
            linewidth=1.0,
            label="Treat all",
            zorder=2,
        )

        n_unique = int(
            reference[ID_COL].nunique()
        )
        event_count = int(
            reference["y_true"].sum()
        )

        ax.set_title(
            (
                f"{panel_label}\n"
                f"N={n_unique:,}, events={event_count:,} "
                f"({prevalence * 100:.1f}%)"
            ),
            fontweight="bold",
            fontsize=10.3,
        )
        ax.grid(
            alpha=0.20,
        )
        ax.set_xlim(
            DCA_THRESHOLD_MIN,
            DCA_THRESHOLD_MAX,
        )

    # Keep the clinically relevant region readable.
    if all_model_net_benefits:
        finite_benefits = np.asarray(
            all_model_net_benefits,
            dtype=float,
        )
        finite_benefits = finite_benefits[
            np.isfinite(finite_benefits)
        ]

        upper = max(
            0.12,
            float(
                np.nanpercentile(
                    finite_benefits,
                    99,
                )
            )
            + 0.01,
        )
    else:
        upper = 0.12

    for ax in axes:
        ax.set_ylim(
            -0.02,
            upper,
        )
        ax.set_xlabel(
            "Threshold probability"
        )

    axes[0].set_ylabel(
        "Net benefit"
    )
    axes[2].set_ylabel(
        "Net benefit"
    )

    # Six model-feature-set entries plus two reference strategies.
    shared_handles: List[Line2D] = []

    for model in MODEL_ORDER:
        for variant in VARIANT_ORDER:
            shared_handles.append(
                Line2D(
                    [0],
                    [0],
                    color=MODEL_COLORS[model],
                    linestyle=VARIANT_LINESTYLES[variant],
                    linewidth=(
                        2.1
                        if variant == "Perioperative"
                        else 1.8
                    ),
                    label=(
                        f"{VARIANT_DISPLAY[variant]}—"
                        f"{MODEL_DISPLAY[model]}"
                    ),
                )
            )

    shared_handles.extend(
        [
            Line2D(
                [0],
                [0],
                color="gray",
                linestyle="--",
                linewidth=1.2,
                label="Treat none",
            ),
            Line2D(
                [0],
                [0],
                color="black",
                linestyle="-.",
                linewidth=1.2,
                label="Treat all",
            ),
        ]
    )

    fig.legend(
        handles=shared_handles,
        loc="lower center",
        bbox_to_anchor=(0.5, 0.045),
        ncol=4,
        frameon=False,
        fontsize=7.8,
        columnspacing=1.2,
        handlelength=3.0,
    )


    fig.tight_layout(
        rect=[0, 0.105, 1, 1]
    )

    output_path = (
        MAIN_FIGURE_DIR
        / "Figure3_IECV_decision_curves.png"
    )
    save_figure_png_compliant(
        fig,
        output_path,
    )
    plt.close(fig)

    return output_path


# ============================================================
# 5. Revised random hold-out supplementary figure
# ============================================================

def create_secondary_validation_figure(
    predictions: pd.DataFrame,
    *,
    validation_name: str,
    output_filename: str,
    empty_error_message: str,
) -> Path:
    validation_predictions = predictions.loc[
        predictions[
            "Validation"
        ].astype(str).eq(validation_name)
    ].copy()

    if validation_predictions.empty:
        raise ValueError(empty_error_message)

    # Use one reference group to obtain prevalence.
    reference = validation_predictions.loc[
        validation_predictions["Model"].eq(
            MODEL_ORDER[0]
        )
        & validation_predictions["Variant"].eq(
            VARIANT_ORDER[0]
        )
    ]

    if reference.empty:
        reference = validation_predictions

    prevalence = float(
        reference["y_true"].mean()
    )

    fig, axes = plt.subplots(
        2,
        2,
        figsize=(10.4, 9.0),
    )

    # Six model-feature-set combinations
    for model in MODEL_ORDER:
        for variant in VARIANT_ORDER:
            group = validation_predictions.loc[
                validation_predictions["Model"].eq(model)
                & validation_predictions["Variant"].eq(
                    variant
                )
            ].copy()

            if group.empty:
                continue

            y = group[
                "y_true"
            ].to_numpy(dtype=int)
            p = group[
                "y_pred"
            ].to_numpy(dtype=float)

            label = (
                f"{VARIANT_DISPLAY[variant]}—"
                f"{MODEL_DISPLAY[model]}"
            )

            precision, recall, _ = (
                precision_recall_curve(
                    y,
                    p,
                )
            )
            axes[0, 0].plot(
                recall,
                precision,
                color=MODEL_COLORS[model],
                linestyle=VARIANT_LINESTYLES[
                    variant
                ],
                linewidth=1.5,
                label=label,
            )

            false_positive_rate, true_positive_rate, _ = (
                roc_curve(
                    y,
                    p,
                )
            )
            axes[0, 1].plot(
                false_positive_rate,
                true_positive_rate,
                color=MODEL_COLORS[model],
                linestyle=VARIANT_LINESTYLES[
                    variant
                ],
                linewidth=1.5,
                label=label,
            )

            observed, predicted = (
                calibration_curve(
                    y,
                    p,
                    n_bins=10,
                    strategy="quantile",
                )
            )
            axes[1, 0].plot(
                predicted,
                observed,
                color=MODEL_COLORS[model],
                linestyle=VARIANT_LINESTYLES[
                    variant
                ],
                marker=VARIANT_MARKERS[
                    variant
                ],
                markersize=3.2,
                linewidth=1.3,
                label=label,
            )

            thresholds = np.linspace(
                DCA_THRESHOLD_MIN,
                DCA_THRESHOLD_MAX,
                DCA_N_THRESHOLDS,
            )
            benefits = [
                net_benefit(
                    y,
                    p,
                    threshold,
                )
                for threshold in thresholds
            ]
            axes[1, 1].plot(
                thresholds,
                benefits,
                color=MODEL_COLORS[model],
                linestyle=VARIANT_LINESTYLES[
                    variant
                ],
                linewidth=1.5,
                label=label,
            )

    # PR baseline
    axes[0, 0].axhline(
        prevalence,
        color="gray",
        linestyle="--",
        linewidth=1.0,
    )
    axes[0, 0].text(
        0.98,
        prevalence + 0.015,
        f"Prevalence={prevalence:.3f}",
        ha="right",
        va="bottom",
        fontsize=8,
        color="gray",
    )

    # ROC chance line
    axes[0, 1].plot(
        [0, 1],
        [0, 1],
        color="gray",
        linestyle="--",
        linewidth=1.0,
    )

    # Perfect calibration line
    axes[1, 0].plot(
        [0, 1],
        [0, 1],
        color="gray",
        linestyle="--",
        linewidth=1.0,
    )

    # DCA reference strategies
    thresholds = np.linspace(
        DCA_THRESHOLD_MIN,
        DCA_THRESHOLD_MAX,
        DCA_N_THRESHOLDS,
    )
    axes[1, 1].plot(
        thresholds,
        np.zeros_like(
            thresholds
        ),
        color="gray",
        linestyle="--",
        linewidth=1.0,
        label="Treat none",
    )
    axes[1, 1].plot(
        thresholds,
        (
            prevalence
            - (1.0 - prevalence)
            * thresholds
            / (1.0 - thresholds)
        ),
        color="black",
        linestyle="-.",
        linewidth=1.0,
        label="Treat all",
    )

    axes[0, 0].set_title(
        "(A) Precision–recall",
        fontweight="bold",
    )
    axes[0, 1].set_title(
        "(B) ROC",
        fontweight="bold",
    )
    axes[1, 0].set_title(
        "(C) Calibration",
        fontweight="bold",
    )
    axes[1, 1].set_title(
        "(D) Decision curve",
        fontweight="bold",
    )

    axes[0, 0].set_xlabel(
        "Recall"
    )
    axes[0, 0].set_ylabel(
        "Precision"
    )
    axes[0, 0].set_xlim(
        0,
        1,
    )
    axes[0, 0].set_ylim(
        0,
        1.02,
    )

    axes[0, 1].set_xlabel(
        "False-positive rate"
    )
    axes[0, 1].set_ylabel(
        "True-positive rate"
    )
    axes[0, 1].set_xlim(
        0,
        1,
    )
    axes[0, 1].set_ylim(
        0,
        1.02,
    )

    axes[1, 0].set_xlabel(
        "Predicted probability"
    )
    axes[1, 0].set_ylabel(
        "Observed event rate"
    )
    axes[1, 0].set_xlim(
        0,
        1,
    )
    axes[1, 0].set_ylim(
        0,
        1.02,
    )

    axes[1, 1].set_xlabel(
        "Threshold probability"
    )
    axes[1, 1].set_ylabel(
        "Net benefit"
    )
    axes[1, 1].set_xlim(
        DCA_THRESHOLD_MIN,
        DCA_THRESHOLD_MAX,
    )
    axes[1, 1].set_ylim(
        -0.02,
        0.12,
    )
    axes[1, 1].legend(
        handles=[
            Line2D(
                [0],
                [0],
                color="gray",
                linestyle="--",
                linewidth=1.2,
                label="Treat none",
            ),
            Line2D(
                [0],
                [0],
                color="black",
                linestyle="-.",
                linewidth=1.2,
                label="Treat all",
            ),
        ],
        loc="upper right",
        frameon=False,
        fontsize=8,
    )

    for ax in axes.ravel():
        ax.grid(
            alpha=0.20,
        )

    shared_handles = []
    for model in MODEL_ORDER:
        for variant in VARIANT_ORDER:
            shared_handles.append(
                Line2D(
                    [0],
                    [0],
                    color=MODEL_COLORS[
                        model
                    ],
                    linestyle=VARIANT_LINESTYLES[
                        variant
                    ],
                    linewidth=1.8,
                    label=(
                        f"{VARIANT_DISPLAY[variant]}—"
                        f"{MODEL_DISPLAY[model]}"
                    ),
                )
            )

    fig.legend(
        handles=shared_handles,
        loc="lower center",
        bbox_to_anchor=(0.5, 0.015),
        ncol=3,
        frameon=False,
        fontsize=8.2,
    )

    fig.tight_layout(
        rect=[0, 0.10, 1, 1]
    )

    output_path = (
        MULTIMEDIA_APPENDIX_DIR
        / output_filename
    )
    save_figure_png_compliant(
        fig,
        output_path,
    )
    plt.close(fig)

    return output_path


def create_revised_holdout_figure(
    predictions: pd.DataFrame,
) -> Path:
    return create_secondary_validation_figure(
        predictions,
        validation_name="Holdout",
        output_filename="Random_holdout_performance.png",
        empty_error_message="No random hold-out predictions were found.",
    )


def create_revised_temporal_figure(
    predictions: pd.DataFrame,
) -> Path:
    return create_secondary_validation_figure(
        predictions,
        validation_name="Temporal",
        output_filename="Temporal_validation_performance.png",
        empty_error_message="No temporal-validation predictions were found.",
    )


# ============================================================
# 6. Revised SHAP figures
#    - existing out-of-sample IECV SHAP values are reused
#    - expanded clinical feature names
#    - common x-axis range within each algorithm
# ============================================================

def read_existing_shap_result(
    model: str,
    variant: str,
) -> Dict[str, object]:
    result_dir = (
        SHAP_DIR
        / model
        / variant
    )

    shap_path = (
        result_dir
        / "SHAP_values.csv.gz"
    )
    input_path = (
        result_dir
        / "SHAP_model_inputs.csv.gz"
    )

    require_file(
        shap_path,
        f"{model} {variant} SHAP values",
    )
    require_file(
        input_path,
        f"{model} {variant} SHAP model inputs",
    )

    shap_frame = pd.read_csv(
        shap_path,
        compression="gzip",
    )
    input_frame = pd.read_csv(
        input_path,
        compression="gzip",
    )

    shap_columns = [
        column
        for column in shap_frame.columns
        if column.startswith(
            "SHAP__"
        )
    ]

    feature_names = [
        column.replace(
            "SHAP__",
            "",
            1,
        )
        for column in shap_columns
    ]

    input_columns = [
        f"MODEL_INPUT__{feature}"
        for feature in feature_names
    ]

    missing_input_columns = [
        column
        for column in input_columns
        if column not in input_frame.columns
    ]

    if missing_input_columns:
        raise KeyError(
            f"{model} {variant}: missing model-input columns: "
            f"{missing_input_columns}"
        )

    return {
        "model": model,
        "variant": variant,
        "feature_names": feature_names,
        "shap_values": shap_frame[
            shap_columns
        ].to_numpy(dtype=float),
        "model_inputs": input_frame[
            input_columns
        ].to_numpy(dtype=float),
    }


def create_shap_panel(
    result: Dict[str, object],
    output_path: Path,
    panel_label: str,
    common_xlim: Tuple[float, float],
) -> Path:
    model = str(
        result["model"]
    )
    variant = str(
        result["variant"]
    )
    feature_names = list(
        result["feature_names"]
    )

    display_names = [
        FEATURE_LABELS.get(
            feature,
            feature,
        )
        for feature in feature_names
    ]

    input_data = pd.DataFrame(
        result["model_inputs"],
        columns=display_names,
    )

    plt.figure(
        figsize=(10.3, 5.2)
    )

    shap.summary_plot(
        result["shap_values"],
        input_data,
        max_display=SHAP_TOP_N,
        show=False,
        plot_size=None,
    )

    figure = plt.gcf()
    axis = plt.gca()
    axis.set_xlim(
        common_xlim
    )

    model_label = (
        MODEL_DISPLAY.get(
            model,
            model,
        )
    )

    scale_label = (
        "log-odds scale"
        if model == "Logistic"
        else "raw margin scale"
    )

    axis.set_title(
        (
            f"{panel_label} "
            f"{VARIANT_DISPLAY[variant]} model — "
            f"{model_label}"
        ),
        fontweight="bold",
        fontsize=11,
        pad=8,
    )
    axis.set_xlabel(
        f"SHAP value ({scale_label})"
    )
    axis.tick_params(
        axis="y",
        labelsize=9.5,
    )

    figure.subplots_adjust(
        left=0.30,
        right=0.91,
        top=0.91,
        bottom=0.14,
    )

    raw_path = output_path.with_name(
        f"{output_path.stem}__raw.png"
    )

    figure.savefig(
        raw_path,
        dpi=130,
        bbox_inches="tight",
        pad_inches=0.04,
        facecolor="white",
    )
    plt.close(
        figure
    )

    with Image.open(
        raw_path
    ) as image:
        panel = image.convert(
            "RGB"
        )
        panel.thumbnail(
            (
                SHAP_PANEL_WIDTH,
                SHAP_PANEL_HEIGHT,
            ),
            Image.Resampling.LANCZOS,
        )

        canvas = Image.new(
            "RGB",
            (
                SHAP_PANEL_WIDTH,
                SHAP_PANEL_HEIGHT,
            ),
            "white",
        )

        canvas.paste(
            panel,
            (
                (
                    SHAP_PANEL_WIDTH
                    - panel.width
                )
                // 2,
                (
                    SHAP_PANEL_HEIGHT
                    - panel.height
                )
                // 2,
            ),
        )

        canvas.save(
            output_path,
            format="PNG",
            optimize=True,
            compress_level=9,
            dpi=(
                PNG_DPI,
                PNG_DPI,
            ),
        )

    raw_path.unlink(
        missing_ok=True
    )

    return output_path


def create_revised_shap_figures() -> List[Path]:
    output_paths: List[Path] = []

    figure_prefix = {
        "Logistic": "Figure4A",
        "XGBoost": "Figure4B",
        "LightGBM": "Figure4C",
    }

    for model in MODEL_ORDER:
        preoperative = (
            read_existing_shap_result(
                model,
                "Preoperative",
            )
        )
        perioperative = (
            read_existing_shap_result(
                model,
                "Perioperative",
            )
        )

        pooled_values = np.concatenate(
            [
                np.asarray(
                    preoperative[
                        "shap_values"
                    ],
                    dtype=float,
                ).ravel(),
                np.asarray(
                    perioperative[
                        "shap_values"
                    ],
                    dtype=float,
                ).ravel(),
            ]
        )
        pooled_values = pooled_values[
            np.isfinite(
                pooled_values
            )
        ]

        if len(
            pooled_values
        ) == 0:
            raise ValueError(
                f"No finite SHAP values were found for {model}."
            )

        lower = float(
            pooled_values.min()
        )
        upper = float(
            pooled_values.max()
        )
        span = max(
            upper - lower,
            0.1,
        )
        common_xlim = (
            lower - 0.04 * span,
            upper + 0.04 * span,
        )

        panel_dir = (
            ROOT_DIR
            / "_figure_working_files"
            / "shap_panels"
            / model
        )
        panel_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

        pre_path = create_shap_panel(
            preoperative,
            panel_dir
            / "A_Preoperative.png",
            "(A)",
            common_xlim,
        )
        peri_path = create_shap_panel(
            perioperative,
            panel_dir
            / "B_Perioperative.png",
            "(B)",
            common_xlim,
        )

        top = Image.open(
            pre_path
        ).convert("RGB")
        bottom = Image.open(
            peri_path
        ).convert("RGB")

        canvas = Image.new(
            "RGB",
            (
                SHAP_PANEL_WIDTH,
                top.height
                + bottom.height
                + SHAP_PANEL_GAP,
            ),
            "white",
        )
        canvas.paste(
            top,
            (0, 0),
        )
        canvas.paste(
            bottom,
            (
                0,
                top.height
                + SHAP_PANEL_GAP,
            ),
        )

        if model == "LightGBM":
            output_path = (
                MAIN_FIGURE_DIR
                / "Figure4_SHAP_LightGBM.png"
            )
        elif model == "Logistic":
            output_path = (
                MULTIMEDIA_APPENDIX_DIR
                / "SHAP_Logistic_regression.png"
            )
        elif model == "XGBoost":
            output_path = (
                MULTIMEDIA_APPENDIX_DIR
                / "SHAP_XGBoost.png"
            )
        else:
            raise ValueError(
                f"Unexpected model for SHAP output: {model}"
            )

        save_pil_png_compliant(
            canvas,
            output_path,
        )

        output_paths.append(
            output_path
        )

    return output_paths


# ============================================================
# 7. JMIR caption and submission-note export
# ============================================================

def write_jmir_caption_file(
    output_path: Path,
) -> Path:
    """
    Write captions separately from the image files.

    The caption text intentionally does not begin with a figure number or
    multimedia appendix number because JMIR adds those labels automatically.
    """
    captions = [
        (
            "Figure2_IECV_performance.png",
            (
                "Predictive performance in internal-external cross-validation. "
                "Point estimates and 95% confidence intervals are shown for "
                "(A) the area under the receiver operating characteristic curve, "
                "(B) the area under the precision-recall curve, calculated as average precision, "
                "(C) the scaled Brier score, and (D) the calibration slope. "
                "Open circles indicate preoperative models, filled squares indicate "
                "perioperative models, and colors identify the modeling algorithms. "
                "Pooled estimates combine out-of-sample predictions from the three "
                "held-out institutions. A scaled Brier score of 0 indicates performance "
                "equivalent to a prevalence-based null model, and a calibration slope "
                "of 1 indicates ideal calibration. "
                "AUROC: area under the receiver operating characteristic curve; "
                "IECV: internal-external cross-validation; "
                "PR-AUC was calculated using average precision."
            ),
        ),
        (
            "Figure3_IECV_decision_curves.png",
            (
                "Decision curve analysis in internal-external cross-validation. "
                "Decision curves are shown for the preoperative and perioperative "
                "logistic regression, XGBoost, and LightGBM models in each held-out "
                "institution and in the pooled IECV population. Colors identify the "
                "modeling algorithms, and dotted and solid lines indicate the "
                "preoperative and perioperative feature sets, respectively. "
                "In this exploratory analysis, a positive prediction was interpreted "
                "as prompting prioritized postoperative clinical reassessment rather "
                "than a specific intervention. Treat-all and treat-none represent "
                "reassessment of all patients and no model-triggered reassessment, "
                "respectively. IECV: internal-external cross-validation; "
                "LightGBM: Light Gradient-Boosting Machine; "
                "XGBoost: Extreme Gradient Boosting."
            ),
        ),
        (
            "Figure4_SHAP_LightGBM.png",
            (
                "Out-of-sample SHAP interpretation of the LightGBM models. "
                "SHAP summary plots are shown for the (A) preoperative and "
                "(B) perioperative LightGBM models. SHAP values were calculated for "
                "patients in each held-out institution using models trained exclusively "
                "on the remaining institutions and were subsequently pooled across "
                "IECV folds. The color scale indicates low to high feature values. "
                "SHAP values represent contributions to the uncalibrated model output; "
                "feature importance reflects predictive contribution and should not be "
                "interpreted causally. IECV: internal-external cross-validation; "
                "LightGBM: Light Gradient-Boosting Machine; "
                "SHAP: Shapley Additive Explanations."
            ),
        ),
        (
            "Random_holdout_performance.png",
            (
                "Model performance in the random hold-out validation cohort. "
                "Panels show (A) precision-recall curves, (B) receiver operating "
                "characteristic curves, (C) calibration plots, and (D) decision curves "
                "for the preoperative and perioperative logistic regression, XGBoost, "
                "and LightGBM models. Colors identify the modeling algorithms, and "
                "dotted and solid lines indicate the preoperative and perioperative "
                "feature sets, respectively. LightGBM: Light Gradient-Boosting Machine; "
                "XGBoost: Extreme Gradient Boosting."
            ),
        ),
        (
            "Temporal_validation_performance.png",
            (
                "Model performance in the temporal validation cohort. "
                f"Models were developed using data from {TEMPORAL_DEVELOPMENT_LABEL} "
                f"and evaluated using data from {TEMPORAL_VALIDATION_LABEL}. "
                "Panels show (A) precision-recall curves, (B) receiver operating "
                "characteristic curves, (C) calibration plots, and (D) decision curves "
                "for the preoperative and perioperative logistic regression, XGBoost, "
                "and LightGBM models. Colors identify the modeling algorithms, and "
                "dotted and solid lines indicate the preoperative and perioperative "
                "feature sets, respectively. LightGBM: Light Gradient-Boosting Machine; "
                "XGBoost: Extreme Gradient Boosting."
            ),
        ),
        (
            "SHAP_Logistic_regression.png",
            (
                "Out-of-sample SHAP interpretation of the logistic regression models. "
                "SHAP summary plots are shown for the (A) preoperative and "
                "(B) perioperative models. SHAP values were calculated for patients "
                "in each held-out institution using models trained exclusively on the "
                "remaining institutions and were subsequently pooled across IECV folds. "
                "The color scale indicates low to high feature values. "
                "IECV: internal-external cross-validation; "
                "SHAP: Shapley Additive Explanations."
            ),
        ),
        (
            "SHAP_XGBoost.png",
            (
                "Out-of-sample SHAP interpretation of the XGBoost models. "
                "SHAP summary plots are shown for the (A) preoperative and "
                "(B) perioperative models. SHAP values were calculated for patients "
                "in each held-out institution using models trained exclusively on the "
                "remaining institutions and were subsequently pooled across IECV folds. "
                "The color scale indicates low to high feature values. "
                "IECV: internal-external cross-validation; "
                "SHAP: Shapley Additive Explanations; "
                "XGBoost: Extreme Gradient Boosting."
            ),
        ),
    ]

    lines: List[str] = [
        "JMIR figure captions for submission metadata",
        "=" * 60,
        "",
        (
            "Paste only the caption text into the JMIR Caption field. "
            "Do not add a figure number or multimedia appendix number."
        ),
        "",
    ]

    for file_name, caption in captions:
        lines.extend(
            [
                f"File: {file_name}",
                caption,
                "",
            ]
        )

    output_path.write_text(
        "\n".join(lines),
        encoding="utf-8",
    )

    return output_path


def write_jmir_submission_note(
    output_path: Path,
) -> Path:
    note = """JMIR figure submission checklist

1. Keep figures in the manuscript during peer review and upload each main figure
   separately as a PNG figure file.
2. Enter each caption in the JMIR metadata Caption field.
3. Do not place the figure number or full caption inside the image.
4. Mention every main figure explicitly in the manuscript text.
5. Upload the random hold-out, temporal validation, and additional SHAP figures as
   Multimedia Appendices or other supplementary material according to the editorial decision.
6. Confirm that every PNG is no larger than 1200 × 1200 pixels, no larger than
   5 MB, and contains no transparent sections.
7. Define abbreviations, colors, symbols, marker shapes, and line types in the
   corresponding caption.
"""
    output_path.write_text(
        note,
        encoding="utf-8",
    )
    return output_path


# ============================================================
# 8. Run all figure revisions
# ============================================================

figure_paths: List[Path] = []

figure_paths.append(
    create_revised_figure2(
        table2_numeric
    )
)

figure_paths.append(
    create_revised_figure3_dca(
        primary_predictions
    )
)

figure_paths.append(
    create_revised_holdout_figure(
        primary_predictions
    )
)

if RUN_PRIMARY_TEMPORAL:
    figure_paths.append(
        create_revised_temporal_figure(
            primary_predictions
        )
    )

figure_paths.extend(
    create_revised_shap_figures()
)

caption_path = write_jmir_caption_file(
    CAPTION_FILE
)
submission_note_path = write_jmir_submission_note(
    ROOT_DIR / "JMIR_Figure_Submission_Checklist.txt"
)

# PNG compliance report
compliance_report = pd.DataFrame(
    [
        validate_png(
            path
        )
        for path in figure_paths
    ]
)

compliance_report["Submission_role"] = (
    compliance_report["File"].map(
        {
            "Figure2_IECV_performance.png": "Main figure",
            "Figure3_IECV_decision_curves.png": "Main figure",
            "Figure4_SHAP_LightGBM.png": "Main figure",
            "Random_holdout_performance.png": "Multimedia appendix",
            "Temporal_validation_performance.png": "Multimedia appendix",
            "SHAP_Logistic_regression.png": "Multimedia appendix",
            "SHAP_XGBoost.png": "Multimedia appendix",
        }
    )
)

compliance_path = (
    ROOT_DIR
    / "JMIR_PNG_Compliance_Report.xlsx"
)
compliance_report.to_excel(
    compliance_path,
    index=False,
)

final_manifest_rows = []
for path in figure_paths:
    final_manifest_rows.append(
        {
            "Category": "Publication figure",
            "Label": path.stem,
            "Path": str(path),
        }
    )
for label, path in {
    "Figure captions": caption_path,
    "Figure submission checklist": submission_note_path,
    "PNG compliance report": compliance_path,
}.items():
    final_manifest_rows.append(
        {
            "Category": "Submission support",
            "Label": label,
            "Path": str(path),
        }
    )

final_manifest = pd.DataFrame(final_manifest_rows)
if "core_output_manifest" in globals() and not core_output_manifest.empty:
    final_manifest = pd.concat(
        [core_output_manifest, final_manifest],
        ignore_index=True,
    )
final_manifest_path = ROOT_DIR / "Final_Output_File_Manifest.xlsx"
final_manifest.to_excel(final_manifest_path, index=False)
style_excel_workbook(final_manifest_path)

violations = compliance_report.loc[
    ~compliance_report[
        "All_requirements_met"
    ]
]

if not violations.empty:
    raise RuntimeError(
        "Some PNG files did not meet the submission requirements:\n"
        + violations.to_string(
            index=False
        )
    )

print("=" * 72)
print("Revised figures were created.")
print("=" * 72)

for path in figure_paths:
    print(path)

print("")
print("JMIR caption file:")
print(caption_path)

print("")
print("JMIR submission checklist:")
print(submission_note_path)

print("")
print("PNG compliance report:")
print(compliance_path)

print("")
print("Final output manifest:")
print(final_manifest_path)

display(
    compliance_report[
        [
            "File",
            "Submission_role",
            "Width_px",
            "Height_px",
            "File_size_MB",
            "Has_transparency",
            "All_requirements_met",
        ]
    ]
)
